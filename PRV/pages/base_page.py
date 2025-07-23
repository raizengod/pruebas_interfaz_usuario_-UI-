import re # Importa el módulo de expresiones regulares
import time # Importa el módulo para funciones relacionadas con el tiempo
import random # Importa el módulo para generar números aleatorios
from playwright.sync_api import Page, expect, Error, TimeoutError, sync_playwright, Response, Dialog, Locator # Importa clases y excepciones necesarias de Playwright
from datetime import datetime # Importa la clase datetime para trabajar con fechas y horas
import os # Importa el módulo os para interactuar con el sistema operativo (rutas de archivos, directorios)
from typing import List, Dict, Union, Callable # Importa tipos para mejorar la legibilidad y validación del código
from PRV.utils.config import LOGGER_DIR # Importa la ruta del directorio de logs desde config.py
from PRV.utils.logger import setup_logger # Importa la función setup_logger desde logger.py
import logging # Importa el módulo logging para configurar y usar loggers
import openpyxl # Librería para hacer uso del excel (para archivos .xlsx)
import csv # Importa la librería csv para manejar archivos CSV (para archivos .csv)
import json # Importa la librería json para manejar archivos JSON
import xml.etree.ElementTree as ET # Importa el módulo para trabajar con XML

class Funciones_Globales:
    
    #1- Creamos una función incial 'Constructor'-----ES IMPORTANTE TENER ESTE INICIADOR-----
    def __init__(self, page):
        self.page = page
        self._alerta_detectada = False
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""
        self._alerta_input_capturado = ""
        self._dialog_handler_registered = False # <--- ¡Esta línea es crucial!

        # --- Nuevas variables para el manejo de pestañas (popups) ---
        self._popup_detectado = False
        self._popup_page = None # Para almacenar el objeto Page de la nueva pestaña
        self._popup_url_capturado = ""
        self._popup_title_capturado = ""  
        
        # Nueva lista para almacenar todas las nuevas páginas abiertas durante una interacción
        self._all_new_pages_opened_by_click: List[Page] = []
        
        # Registramos el manejador de eventos para nuevas páginas
        # Limpiamos la lista al registrar para evitar resagos de pruebas anteriores
        self.page.context.on("page", self._on_new_page)
        # Esto es importante: Si se va a usar _all_new_pages_opened_by_click,
        # necesitamos una forma de reiniciarla o asegurarnos de que solo contenga
        # las páginas relevantes para la acción actual.
        # Una estrategia es limpiar la lista antes de la acción que abre la nueva ventana,
        # y luego recopilar las páginas.
        
        # Configurar el logger para esta clase
        self.logger = setup_logger(name='Funciones_Globales', console_level=logging.INFO, file_level=logging.DEBUG)
        
    #2- Función para generar el nombre de archivo con marca de tiempo
    def _generar_nombre_archivo_con_timestamp(self, prefijo):
        now = datetime.now()
        timestamp = now.strftime("%Y-%m-%d_%H-%M-%S-%f")[:-3] # Quita los últimos 3 dígitos para milisegundos más precisos
        return f"{timestamp}_{prefijo}"
    
    #3- Función para tomar captura de pantalla
    def tomar_captura(self, nombre_base, directorio):
        """
        Toma una captura de pantalla de la página y la guarda en el directorio especificado.
        Por defecto, usa SCREENSHOT_DIR de config.py.

        Args:
            nombre_base (str): El nombre base para el archivo de la captura de pantalla.
            directorio (str): El directorio donde se guardará la captura. Por defecto, SCREENSHOT_DIR.
        """
        try:
            if not os.path.exists(directorio):
                os.makedirs(directorio)
                self.logger.info(f"\n Directorio creado para capturas de pantalla: {directorio}") #

            nombre_archivo = self._generar_nombre_archivo_con_timestamp(nombre_base) #
            ruta_completa = os.path.join(directorio, f"{nombre_archivo}.png") # Cambiado a .png para mejor calidad
            self.page.screenshot(path=ruta_completa) #
            self.logger.info(f"\n 📸 Captura de pantalla guardada en: {ruta_completa}") #
        except Exception as e:
            self.logger.error(f"\n ❌ Error al tomar captura de pantalla '{nombre_base}': {e}") #
        
    #4- unción basica para tiempo de espera que espera recibir el parametro tiempo
    #En caso de no pasar el tiempo por parametro, el mismo tendra un valor de medio segundo
    def esperar_fijo(self, tiempo=0.5):
        """
        Espera un tiempo fijo en segundos.

        Args:
            tiempo (Union[int, float]): El tiempo en segundos a esperar. Por defecto, 0.5 segundos.
        """
        self.logger.debug(f"\n Esperando fijo por {tiempo} segundos...") #
        try:
            time.sleep(tiempo) #
            self.logger.info(f"Espera fija de {tiempo} segundos completada.") #
        except TypeError:
            self.logger.error(f"\n ❌ Error: El tiempo de espera debe ser un número. Se recibió: {tiempo}") #
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado durante la espera fija: {e}") #
        
    #5- Función para indicar el tiempo que se tardará en hacer el scroll
    def scroll_pagina(self, horz, vert, tiempo: Union[int, float] = 0.5):
        """
        Realiza un scroll en la página.

        Args:
            horz (int): Cantidad de scroll horizontal. Por defecto, 0.
            vert (int): Cantidad de scroll vertical. Por defecto, 0.
            tiempo (Union[int, float]): Tiempo de espera después del scroll en segundos. Por defecto, 0.5.
        """
        self.logger.debug(f"Realizando scroll - Horizontal: {horz}, Vertical: {vert}. Espera: {tiempo} segundos.") #
        try:
            self.page.mouse.wheel(horz, vert) #
            self.esperar_fijo(tiempo) # Reutiliza la función esperar_fijo para el log y manejo de errores
            self.logger.info(f"Scroll completado (H: {horz}, V: {vert}).") #
        except Exception as e:
            self.logger.error(f"❌ Error al realizar scroll en la página: {e}") #
            
    #6- Función para validar que un elemento es visible
    def validar_elemento_visible(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5, resaltar: bool = True) -> bool:
        """
        Valida que un elemento sea visible en la página.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para que el elemento sea visible.
            resaltar (bool): Si es True, resaltará el elemento si es visible.

        Returns:
            bool: True si el elemento es visible, False en caso contrario.
        """
        self.logger.info(f"\n Validando visibilidad del elemento con selector: '{selector}'") # Usar logger.info en lugar de print

        try:
            # Espera explícita para que el elemento sea visible.
            # Se usa self.page.locator(selector) si 'selector' es una string,
            # o se asume que 'selector' ya es un Locator si viene de otro lado.
            # Aquí, la suposición es que 'selector' es un Locator, pero si es un string,
            # deberías cambiarlo a expect(self.page.locator(selector)).to_be_visible(timeout=tiempo*1000)
            # Para fines de este ejemplo, asumo que 'selector' es un Locator o se maneja la conversión externamente.
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            expect(locator).to_be_visible(timeout=tiempo * 10000) # El tiempo en expect es en milisegundos.

            if resaltar:
                locator.highlight() # Resaltar el elemento visible
                self.logger.debug(f"Elemento '{selector}' resaltado.")

            self.tomar_captura(f"{nombre_base}_visible", directorio)
            self.logger.info(f"\n ✔ ÉXITO: El elemento '{selector}' es visible en la página.") # Usar logger.info
            self.esperar_fijo(tiempo) # Reutiliza la función esperar_fijo

            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El elemento con selector '{selector}' NO es visible "
                f"después de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.warning(error_msg) # Usar logger.warning para timeouts
            self.tomar_captura(f"{nombre_base}_NO_visible_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar la visibilidad de '{selector}'. "
                f"Posibles causas: Selector inválido, elemento desprendido del DOM. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Usar logger.error y exc_info para trazas completas
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al validar la visibilidad de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True) # Usar logger.critical para errores inesperados graves
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise

        finally:
            pass # No hay necesidad de un sleep final aquí, el timeout de expect() ya maneja la espera.

    #7- Función para validar que un elemento NO es visible
    def validar_elemento_no_visible(self, selector, nombre_base, directorio, tiempo=0.5):
        """
        Valida que un elemento NO es visible en la página.

        Args:
            selector: El selector del elemento (puede ser un string).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para que el elemento NO sea visible.
        """
        self.logger.info(f"\n Validando que el elemento con selector '{selector}' NO es visible.")
        try:
            expect(selector).to_be_hidden(timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: El elemento con selector '{selector}' NO es visible.")
            # No se toma captura en el try block si la aserción es para "no visible"
            # La captura de éxito se toma en finally.

        except AssertionError as e:
            error_msg = (
                f"\n ❌ Error: El elemento con selector '{selector}' aún es visible o no se ocultó a tiempo. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_no_visible", directorio)
            raise e
        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El elemento con selector '{selector}' aún es visible "
                f"después de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_no_visible", directorio)
            raise e
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al validar que '{selector}' NO es visible. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_no_visible", directorio)
            raise e
        finally:
            # Tomar una captura de pantalla cuando la función finaliza con éxito
            self.tomar_captura(nombre_base=f"{nombre_base}_exito_no_visible", directorio=directorio)

    #8- Función para verificar que un elemento (o elementos) localizado en una página web contiene un texto específico
    def verificar_texto_contenido(self, selector, texto_esperado, nombre_base, directorio, tiempo=0.5):
        """
        Verifica que un elemento localizado en una página web contiene un texto específico.

        Args:
            selector: El selector del elemento (puede ser un string).
            texto_esperado (str): El texto que se espera encontrar en el elemento.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para que el elemento sea visible y contenga el texto.
        """
        self.logger.info(f"Verificando que el elemento con selector '{selector}' contiene el texto: '{texto_esperado}'.")
        try:
            # Esperar visibilidad del elemento
            expect(selector).to_be_visible(timeout=tiempo * 1000)
            self.logger.debug(f"\n Elemento con selector '{selector}' es visible.")

            # Opcional: Resaltar el elemento para depuración visual
            selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_verificacion_texto", directorio)

            # Verificar que el elemento contenga el texto esperado
            expect(selector).to_contain_text(texto_esperado, timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: Elemento con selector '{selector}' contiene el texto esperado: '{texto_esperado}'.")

            self.tomar_captura(nombre_base=f"{nombre_base}_despues_verificacion_texto", directorio=directorio)
            self.esperar_fijo(tiempo) # Usa la función de espera fija con log

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El elemento con selector '{selector}' no se hizo visible o no contenía "
                f"el texto '{texto_esperado}' después de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_verificacion_texto_timeout", directorio)
            raise

        except AssertionError as e:
            error_msg = (
                f"\n ❌ FALLO (Aserción): El elemento con selector '{selector}' NO contiene el texto esperado: "
                f"'{texto_esperado}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_verificacion_texto_contenido", directorio)
            raise

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar texto para '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_verificacion_texto", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el texto para el selector '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificacion_texto", directorio)
            raise

    #9- Función para rellenar campo de texto y hacer capture la imagen
    def rellenar_campo_de_texto(self, selector, texto, nombre_base, directorio, tiempo=0.5):
        """
        Rellena un campo de texto con el valor especificado y toma capturas de pantalla.

        Args:
            selector: El selector del campo de texto (puede ser un string).
            texto (str): El texto a introducir en el campo.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después de rellenar el campo.
        """
        self.logger.info(f"\n Rellenando campo con selector '{selector}' con el texto: '{texto}'.")
        try:
            selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_de_rellenar_texto", directorio)

            selector.fill(texto)
            self.logger.info(f"\n ✔ ÉXITO: Campo '{selector}' rellenado con éxito con el texto: '{texto}'.")

            self.tomar_captura(f"{nombre_base}_despues_de_rellenar_texto", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al interactuar con '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/editable a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_rellenar", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al interactuar con '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_rellenar", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al interactuar con '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_rellenar", directorio)
            raise

        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo) # Usa la función de espera fija con log

    #10- Función para rellenar campo numérico y hacer capture la imagen
    def rellenar_campo_numerico_positivo(self, selector, valor_numerico: int | float, nombre_base, directorio, tiempo=0.5):
        """
        Rellena un campo de texto con un valor numérico positivo y toma capturas de pantalla.

        Args:
            selector: El selector del campo de texto (puede ser un string).
            valor_numerico (Union[int, float]): El valor numérico positivo a introducir.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después de rellenar el campo.
        """
        self.logger.info(f"\n Rellenando campo con selector '{selector}' con el valor numérico POSITIVO: '{valor_numerico}'.")

        if not isinstance(valor_numerico, (int, float)):
            error_msg = f"\n ❌ ERROR: El valor proporcionado '{valor_numerico}' no es un tipo numérico (int o float)."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_valor_no_numerico", directorio)
            raise ValueError(error_msg)

        if valor_numerico < 0:
            error_msg = f"\n ❌ ERROR: El valor numérico '{valor_numerico}' no es positivo. Se esperaba un número mayor o igual a cero."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_valor_negativo", directorio)
            raise ValueError(error_msg)

        valor_a_rellenar_str = str(valor_numerico)

        try:
            selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_de_rellenar_numerico", directorio)

            selector.fill(valor_a_rellenar_str)
            self.logger.info(f"\n ✔ ÉXITO: Campo '{selector}' rellenado con éxito con el valor: '{valor_a_rellenar_str}'.")

            self.tomar_captura(f"{nombre_base}_despues_de_rellenar_numerico", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al interactuar con '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/editable a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_numerico", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al interactuar con '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_numerico", directorio)
            raise

        except TypeError as e:
            error_msg = (
                f"\n ❌ ERROR (TypeError): El selector proporcionado no es un objeto Locator válido.\n"
                f"Asegúrate de pasar un objeto locator o una cadena para que sea convertido a locator.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_tipo_selector_numerico", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al interactuar con '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_numerico", directorio)
            raise

        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    #11- Función para validar titulo de una página
    def validar_titulo_de_web(self, titulo_esperado, nombre_base, directorio, tiempo=0.5):
        """
        Valida el título de la página web actual.

        Args:
            titulo_esperado (str): El título que se espera que tenga la página.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la validación del título.
        """
        self.logger.info(f"\n Validando que el título de la página sea: '{titulo_esperado}'.")
        try:
            expect(self.page).to_have_title(titulo_esperado, timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: Título de la página '{self.page.title()}' validado exitosamente.")
            self.tomar_captura(f"{nombre_base}_exito_titulo", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El título de la página no coincidió con '{titulo_esperado}' "
                f"después de {tiempo} segundos. Título actual: '{self.page.title()}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_titulo_timeout", directorio)
            raise

        except AssertionError as e:
            error_msg = (
                f"\n ❌ FALLO (Aserción): El título de la página NO coincide con '{titulo_esperado}'. "
                f"Título actual: '{self.page.title()}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_titulo_no_coincide", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al validar el título de la página. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_titulo", directorio)
            raise
        
    #12- Función para validar URL    
    def validar_url_actual(self, patron_url, tiempo=0.5):
        """
        Valida la URL actual de la página usando un patrón de expresión regular.

        Args:
            patron_url (str): El patrón de expresión regular que se espera que coincida con la URL.
            tiempo (Union[int, float]): Tiempo máximo de espera para la validación de la URL.
        """
        self.logger.info(f"\n Validando que la URL actual coincida con el patrón: '{patron_url}'.")
        try:
            expect(self.page).to_have_url(re.compile(patron_url), timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: URL '{self.page.url}' validada exitosamente con el patrón: '{patron_url}'.")
            # No se toma captura por defecto para URL, pero se puede añadir si es necesario visualmente
            
        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): La URL actual '{self.page.url}' no coincidió con el patrón "
                f"'{patron_url}' después de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # self.tomar_captura("fallo_url_timeout", directorio) # Considerar añadir si la captura es útil
            raise

        except AssertionError as e:
            error_msg = (
                f"\n ❌ FALLO (Aserción): La URL actual '{self.page.url}' NO coincide con el patrón: "
                f"'{patron_url}'. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            # self.tomar_captura("fallo_url_no_coincide", directorio) # Considerar añadir si la captura es útil
            raise
        
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al validar la URL. "
                f"URL actual: '{self.page.url}', Patrón esperado: '{patron_url}'. Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            raise
        
    #13- Función para hacer click
    def hacer_click_en_elemento(self, selector, nombre_base, directorio, texto_esperado=None, tiempo=0.5):
        """
        Realiza un click en un elemento y toma capturas de pantalla.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            texto_esperado (str, optional): Texto que se espera que el elemento contenga. Por defecto, None.
            tiempo (Union[int, float]): Tiempo de espera después del click.
        """
        self.logger.info(f"\n Intentando hacer click en el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector
            
            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_click", directorio)

            if texto_esperado:
                expect(locator).to_have_text(texto_esperado, timeout=tiempo * 1000)
                self.logger.info(f"\n ✅ El elemento con selector '{selector}' contiene el texto esperado: '{texto_esperado}'.")

            locator.click(timeout=tiempo * 1000) # Añadir timeout al click
            self.logger.info(f"\n ✔ ÉXITO: Click realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_click", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer click en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/clicable a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_click", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer click en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_click", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer click en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_click", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #14- Función para hacer doble click
    def hacer_doble_click_en_elemento(self, selector, nombre_base, directorio, texto_esperado=None, tiempo=1):
        """
        Realiza un doble click en un elemento y toma capturas de pantalla.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            texto_esperado (str, optional): Texto que se espera que el elemento contenga. Por defecto, None.
            tiempo (Union[int, float]): Tiempo de espera después del doble click.
        """
        self.logger.info(f"\n Intentando hacer doble click en el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_doble_click", directorio)

            if texto_esperado:
                expect(locator).to_have_text(texto_esperado, timeout=tiempo * 1000)
                self.logger.info(f"\n ✅ El elemento con selector '{selector}' contiene el texto esperado: '{texto_esperado}'.")

            locator.dblclick(timeout=tiempo * 1000) # Añadir timeout al dblclick
            self.logger.info(f"\n ✔ ÉXITO: Doble click realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_doble_click", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer doble click en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado/clicable a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_doble_click", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer doble click en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_doble_click", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer doble click en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_doble_click", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    #15- Función para hacer hover
    def hacer_hover_en_elemento(self, selector, nombre_base, directorio, tiempo=0.5):
        """
        Realiza una acción de hover sobre un elemento.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después del hover.
        """
        self.logger.info(f"\n Intentando hacer hover sobre el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_hover", directorio)

            locator.hover(timeout=tiempo * 1000) # Añadir timeout al hover
            self.logger.info(f"\n ✔ ÉXITO: Hover realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_hover", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer hover en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_hover", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer hover en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_hover", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer hover en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_hover", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #16- Función para verificar si un elemento está habilitado (enabled)
    def verificar_elemento_habilitado(self, selector, nombre_base, directorio, tiempo=0.5) -> bool:
        """
        Verifica si un elemento está habilitado (enabled) en la página.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para verificar la habilitación.

        Returns:
            bool: True si el elemento está habilitado, False en caso contrario.
        """
        self.logger.info(f"\n Verificando si el elemento con selector '{selector}' está habilitado.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector
            
            locator.highlight()
            expect(locator).to_be_enabled(timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: El elemento '{selector}' está habilitado.")
            self.tomar_captura(f"{nombre_base}_habilitado", directorio)
            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El elemento con selector '{selector}' NO está habilitado "
                f"después de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_NO_habilitado_timeout", directorio)
            return False

        except AssertionError as e:
            error_msg = (
                f"\n ❌ FALLO (Aserción): El elemento con selector '{selector}' NO está habilitado. "
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_NO_habilitado_fallo", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar habilitación de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_habilitado", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la habilitación de '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_habilitado", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #17- Función para mover el mouse a coordenadas y hacer clic
    def mouse_mueve_y_hace_clic_xy(self, x: int, y: int, nombre_base, directorio, tiempo=0.5):
        """
        Mueve el mouse a unas coordenadas X, Y y hace clic.

        Args:
            x (int): Coordenada X.
            y (int): Coordenada Y.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después del clic.
        """
        self.logger.info(f"\n Moviendo el mouse a X:{x}, Y:{y} y haciendo click.")
        try:
            self.tomar_captura(f"{nombre_base}_antes_mouse_click_xy", directorio)
            
            self.page.mouse.move(x, y, steps=5) # Mueve el mouse suavemente
            self.logger.debug(f"\n Mouse movido a X:{x}, Y:{y}.")
            
            self.page.mouse.click(x, y)
            self.logger.info(f"\n ✔ ÉXITO: Click realizado en X:{x}, Y:{y}.")
            self.tomar_captura(f"{nombre_base}_despues_mouse_click_xy", directorio)

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR: Fallo al mover el mouse y hacer clic en X:{x}, Y:{y}. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_mouse_click_xy", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #18- Función para marcar un checkbox
    def marcar_checkbox(self, selector, nombre_base, directorio, tiempo=0.5):
        """
        Marca un checkbox y verifica que está marcado.

        Args:
            selector: El selector del checkbox (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la acción y verificación.
        """
        self.logger.info(f"\n Intentando marcar el checkbox con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_marcar_checkbox", directorio)
            
            locator.check(timeout=tiempo * 1000)
            expect(locator).to_be_checked(timeout=tiempo * 1000) # Verifica que está marcado
            
            self.logger.info(f"\n ✔ ÉXITO: Checkbox con selector '{selector}' marcado exitosamente.")
            self.tomar_captura(f"{nombre_base}_despues_marcar_checkbox", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El checkbox con selector '{selector}' no pudo ser marcado "
                f"o verificado como marcado dentro de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_marcar", directorio)
            raise AssertionError(f"\n Checkbox no marcado/verificado (Timeout): {selector}") from e

        except Error as e: # Playwright-specific errors
            error_msg = (
                f"\n ❌ FALLO (Playwright Error): Problema al interactuar con el checkbox '{selector}'.\n"
                f"Posibles causas: Selector inválido, elemento no interactuable, DOM no estable.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_marcar", directorio)
            raise AssertionError(f"\n Error de Playwright con checkbox: {selector}") from e

        except Exception as e: # Catch-all for any other unexpected errors
            error_msg = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar marcar el checkbox '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_marcar", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #19- Función para desmarcar un checkbox
    def desmarcar_checkbox(self, selector, nombre_base, directorio, tiempo=0.5):
        """
        Desmarca un checkbox y verifica que está desmarcado.

        Args:
            selector: El selector del checkbox (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la acción y verificación.
        """
        self.logger.info(f"\n Intentando desmarcar el checkbox con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_desmarcar_checkbox", directorio)
            
            locator.uncheck(timeout=tiempo * 1000)
            expect(locator).not_to_be_checked(timeout=tiempo * 1000) # Verifica que no está marcado
            
            self.logger.info(f"\n ✔ ÉXITO: Checkbox con selector '{selector}' desmarcado exitosamente.")
            self.tomar_captura(f"{nombre_base}_despues_desmarcar_checkbox", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El checkbox con selector '{selector}' no pudo ser desmarcado "
                f"o verificado como desmarcado dentro de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_desmarcar", directorio)
            raise AssertionError(f"\n Checkbox no desmarcado/verificado (Timeout): {selector}") from e

        except Error as e: # Playwright-specific errors
            error_msg = (
                f"\n ❌ FALLO (Playwright Error): Problema al interactuar con el checkbox '{selector}'.\n"
                f"Posibles causas: Selector inválido, elemento no interactuable, DOM no estable.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_desmarcar", directorio)
            raise AssertionError(f"\n Error de Playwright con checkbox: {selector}") from e

        except Exception as e: # Catch-all for any other unexpected errors
            error_msg = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar desmarcar el checkbox '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_desmarcar", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    #20- Función para verificar el valor de un campo de texto
    def verificar_valor_campo(self, selector, valor_esperado: str, nombre_base, directorio, tiempo=0.5) -> bool:
        """
        Verifica que el valor de un campo de texto coincida con el valor esperado.

        Args:
            selector: El selector del campo de texto (puede ser un string o un Locator de Playwright).
            valor_esperado (str): El valor de texto que se espera encontrar en el campo.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la verificación.

        Returns:
            bool: True si el valor coincide, False en caso contrario.
        """
        self.logger.info(f"\n Verificando que el campo '{selector}' contiene el valor: '{valor_esperado}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector
            
            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_verificar_valor_campo", directorio)

            expect(locator).to_have_value(valor_esperado, timeout=tiempo * 1000)
            
            self.logger.info(f"\n ✔ ÉXITO: El campo '{selector}' contiene el valor esperado: '{valor_esperado}'.")
            self.tomar_captura(f"{nombre_base}_despues_verificar_valor_campo", directorio)
            return True

        except TimeoutError as e:
            actual_value = locator.input_value() if locator else "N/A"
            error_msg = (
                f"\n ❌ FALLO (Timeout): El campo '{selector}' no contiene el valor esperado '{valor_esperado}' "
                f"después de {tiempo} segundos. Valor actual: '{actual_value}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_valor_campo", directorio)
            return False

        except AssertionError as e:
            actual_value = locator.input_value() if locator else "N/A"
            error_msg = (
                f"\n ❌ FALLO (Aserción): El campo '{selector}' NO contiene el valor esperado '{valor_esperado}'. "
                f"Valor actual: '{actual_value}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_verificar_valor_campo", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar el valor del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_valor_campo", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el valor del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_valor_campo", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #21- Función para verificar el valor de un campo numérico (entero)
    def verificar_valor_campo_numerico_int(self, selector, valor_numerico_esperado: int, nombre_base, directorio, tiempo=0.5) -> bool:
        """
        Verifica que el valor de un campo numérico (entero) coincida con el valor esperado.

        Args:
            selector: El selector del campo de texto (puede ser un string o un Locator de Playwright).
            valor_numerico_esperado (int): El valor entero numérico que se espera encontrar en el campo.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la verificación.

        Returns:
            bool: True si el valor numérico (entero) coincide, False en caso contrario.
        """
        self.logger.info(f"\n Verificando que el campo '{selector}' contiene el valor numérico entero: '{valor_numerico_esperado}'.")

        if not isinstance(valor_numerico_esperado, int):
            error_msg = f"\n ❌ ERROR: 'valor_numerico_esperado' debe ser un entero, se recibió: {type(valor_numerico_esperado).__name__}."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_tipo_valor_int", directorio)
            raise TypeError(error_msg)

        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector
            
            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_verificar_valor_int", directorio)

            # Esperar a que el campo tenga algún valor y luego obtenerlo
            expect(locator).to_have_value(str(valor_numerico_esperado), timeout=tiempo * 1000)
            
            self.logger.info(f"\n ✔ ÉXITO: El campo '{selector}' contiene el valor numérico entero esperado: '{valor_numerico_esperado}'.")
            self.tomar_captura(f"{nombre_base}_despues_verificar_valor_int", directorio)
            return True

        except TimeoutError as e:
            actual_value_str = locator.input_value() if locator else "N/A"
            error_msg = (
                f"\n ❌ FALLO (Timeout): El campo '{selector}' no contiene el valor entero esperado '{valor_numerico_esperado}' "
                f"después de {tiempo} segundos. Valor actual: '{actual_value_str}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_valor_int", directorio)
            return False

        except AssertionError as e:
            actual_value_str = locator.input_value() if locator else "N/A"
            try:
                actual_value = int(actual_value_str)
                comparison_msg = f"\n (Actual: {actual_value}, Esperado: {valor_numerico_esperado})"
            except ValueError:
                comparison_msg = f"\n (Valor actual no numérico: '{actual_value_str}', Esperado: {valor_numerico_esperado})"

            error_msg = (
                f"\n ❌ FALLO (Aserción): El campo '{selector}' NO contiene el valor numérico entero esperado. "
                f"{comparison_msg}. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_verificar_valor_int", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar el valor numérico entero del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_valor_int", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el valor numérico entero del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_valor_int", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #22- Función para verificar el valor de un campo numérico (flotante)
    def verificar_valor_campo_numerico_float(self, selector, valor_numerico_esperado: float, nombre_base, directorio, tiempo=0.5, tolerancia: float = 1e-6) -> bool:
        """
        Verifica que el valor de un campo numérico (flotante) coincida con el valor esperado.

        Args:
            selector: El selector del campo de texto (puede ser un string o un Locator de Playwright).
            valor_numerico_esperado (float): El valor flotante numérico que se espera encontrar en el campo.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la verificación.
            tolerancia (float): Tolerancia para la comparación de números flotantes.

        Returns:
            bool: True si el valor numérico (flotante) coincide, False en caso contrario.
        """
        self.logger.info(f"\n Verificando que el campo '{selector}' contiene el valor numérico flotante: '{valor_numerico_esperado}'.")

        if not isinstance(valor_numerico_esperado, float):
            error_msg = f"\n ❌ ERROR: 'valor_numerico_esperado' debe ser un flotante, se recibió: {type(valor_numerico_esperado).__name__}."
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_tipo_valor_float", directorio)
            raise TypeError(error_msg)

        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector
            
            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_verificar_valor_float", directorio)

            # Esperar a que el campo tenga algún valor y luego obtenerlo
            # No podemos usar to_have_value directamente con float por la representación de string
            expect(locator).to_be_visible(timeout=tiempo * 1000) # Primero asegurar que está visible
            actual_value_str = locator.input_value()

            try:
                actual_value_float = float(actual_value_str)
                if abs(actual_value_float - valor_numerico_esperado) < tolerancia:
                    self.logger.info(f"\n ✔ ÉXITO: El campo '{selector}' contiene el valor numérico flotante esperado: '{valor_numerico_esperado}'. Valor actual: '{actual_value_float}'.")
                    self.tomar_captura(f"{nombre_base}_despues_verificar_valor_float", directorio)
                    return True
                else:
                    error_msg = (
                        f"\n ❌ FALLO (Inexactitud): El campo '{selector}' NO contiene el valor numérico flotante esperado. "
                        f"Actual: {actual_value_float}, Esperado: {valor_numerico_esperado}, Diferencia: {abs(actual_value_float - valor_numerico_esperado)} (Tolerancia: {tolerancia})."
                    )
                    self.logger.warning(error_msg)
                    self.tomar_captura(f"{nombre_base}_fallo_inexactitud_float", directorio)
                    return False

            except ValueError:
                error_msg = (
                    f"\n ❌ FALLO (Valor no numérico): El valor actual del campo '{selector}' ('{actual_value_str}') "
                    f"no pudo ser convertido a flotante para comparación. Se esperaba '{valor_numerico_esperado}'."
                )
                self.logger.warning(error_msg)
                self.tomar_captura(f"{nombre_base}_fallo_valor_no_float", directorio)
                return False

        except TimeoutError as e:
            actual_value_str = locator.input_value() if locator else "N/A"
            error_msg = (
                f"\n ❌ FALLO (Timeout): El campo '{selector}' no se hizo visible o no se pudo obtener su valor "
                f"después de {tiempo} segundos para verificar el flotante '{valor_numerico_esperado}'. "
                f"Valor actual (si disponible): '{actual_value_str}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_valor_float", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar el valor numérico flotante del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_valor_float", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el valor numérico flotante del campo '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_valor_float", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)

    #23- Función para verificar el texto 'alt' de una imagen
    def verificar_alt_imagen(self, selector, texto_alt_esperado: str, nombre_base, directorio, tiempo=0.5) -> bool:
        """
        Verifica que el texto 'alt' de una imagen coincida con el valor esperado.

        Args:
            selector: El selector de la imagen (puede ser un string o un Locator de Playwright).
            texto_alt_esperado (str): El texto 'alt' que se espera encontrar en la imagen.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la visibilidad y verificación.

        Returns:
            bool: True si el texto 'alt' coincide, False en caso contrario.
        """
        self.logger.info(f"\n Verificando el texto 'alt' para la imagen con selector: '{selector}' con valor esperado: '{texto_alt_esperado}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_verificar_alt_imagen", directorio)

            # Esperar a que la imagen sea visible
            expect(locator).to_be_visible(timeout=tiempo * 1000)
            self.logger.debug(f"\n La imagen con selector '{selector}' es visible.")

            # Obtener el atributo 'alt' de la imagen
            alt_text_actual = locator.get_attribute("alt", timeout=tiempo * 1000)

            # Validar que el atributo 'alt' no sea None y coincida con el texto esperado
            if alt_text_actual == texto_alt_esperado:
                self.logger.info(f"\n ✔ ÉXITO: El texto 'alt' de la imagen es '{alt_text_actual}' y coincide con el esperado ('{texto_alt_esperado}').")
                self.tomar_captura(f"{nombre_base}_alt_ok", directorio)
                return True
            else:
                error_msg = (
                    f"\n ❌ FALLO (No Coincide): El texto 'alt' actual es '{alt_text_actual}', "
                    f"pero se esperaba '{texto_alt_esperado}'. Imagen: '{selector}'."
                )
                self.logger.warning(error_msg)
                self.tomar_captura(f"{nombre_base}_alt_error", directorio)
                return False

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): La imagen con selector '{selector}' no se hizo visible "
                f"o no se pudo obtener su atributo 'alt' después de {tiempo} segundos. Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_alt_imagen", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar el texto 'alt' de la imagen '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_alt_imagen", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el texto 'alt' de la imagen '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_alt_imagen", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
                
    #24- Función para Verifica que una imagen se cargue exitosamente (sin enlaces rotos).
    def verificar_carga_exitosa_imagen(self, selector, nombre_base, directorio, tiempo_espera_red= 10, tiempo= 1) -> bool:
        image_url = None
        self.logger.info(f"\n Iniciando verificación de carga exitosa para la imagen con selector: '{selector}'")
        try:
            # 1. Resaltar el elemento (útil para depuración visual)
            selector.highlight()
            self.logger.debug(f"\n Elemento con selector '{selector}' resaltado.")

            # 2. Esperar a que la imagen sea visible en el DOM
            expect(selector).to_be_visible(timeout=tiempo_espera_red * 1000)
            self.logger.info(f"\n La imagen con selector '{selector}' es visible en el DOM.")

            # 3. Obtener la URL de la imagen
            image_url = selector.get_attribute("src")
            if not image_url:
                self.logger.error(f"\n Error: El atributo 'src' de la imagen con selector '{selector}' está vacío o no existe.")
                self.tomar_captura(f"{nombre_base}_src_vacio", directorio)
                return False

            self.logger.info(f"\n URL de la imagen a verificar: {image_url}")

            # 4. Monitorear la carga de la imagen en la red
            # Usamos page.wait_for_response para esperar la respuesta HTTP de la imagen
            # Esto es más robusto que solo verificar la visibilidad, ya que asegura que el recurso fue descargado
            self.logger.debug(f"\n Esperando respuesta de red para la imagen con URL: {image_url} (timeout: {tiempo_espera_red}s).")
            response = self.page.wait_for_response(
                lambda response: response.url == image_url and response.request.resource_type == "image",
                timeout=tiempo_espera_red * 1000 # Playwright espera milisegundos
            )

            # 5. Verificar el código de estado de la respuesta HTTP
            if 200 <= response.status <= 299:
                self.logger.info(f"\n ✔ ÉXITO: La imagen con URL '{image_url}' cargó exitosamente con estado {response.status}.")
                self.tomar_captura(f"{nombre_base}_carga_ok", directorio)
                self.esperar_fijo(tiempo) # Uso consistente de la función de espera fija
                return True
            else:
                self.logger.error(f"\n ❌ FALLO: La imagen con URL '{image_url}' cargó con un estado de error: {response.status}.")
                self.tomar_captura(f"{nombre_base}_carga_fallida_status_{response.status}", directorio)
                return False

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo verificar la carga de la imagen con selector '{selector}' "
                f"y URL '{image_url if image_url else 'N/A'}'.\n"
                f"El elemento no apareció a tiempo o la respuesta de red excedió el tiempo de espera.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_timeout_verificacion", directorio)
            return False
        except Error as e: # Captura errores específicos de Playwright
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar la carga de la imagen con selector '{selector}' "
                f"y URL '{image_url if image_url else 'N/A'}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            return False
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la carga de la imagen con selector '{selector}' "
                f"y URL '{image_url if image_url else 'N/A'}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            return False
    
    #25- Función para cargar archivo(s)
    def cargar_archivo(self, selector, nombre_base, directorio, base_dir, file_names: str | list[str], tiempo= 3):
        # Normalizar file_names a una lista si se pasa una sola cadena
        if isinstance(file_names, str):
            file_names_list = [file_names]
        else:
            file_names_list = file_names

        self.logger.info(f"\n Intentando cargar archivo(s) '{file_names_list}' en el selector: '{selector}'.")

        # Construir las rutas completas de los archivos
        full_file_paths = []
        for name in file_names_list:
            full_path = os.path.join(base_dir, name)
            full_file_paths.append(full_path)
            self.logger.debug(f"\n Construida ruta completa para archivo: '{full_path}'")

        # Verificar que todos los archivos existan antes de intentar la carga
        for path in full_file_paths:
            if not os.path.exists(path):
                error_msg = f"\n ❌ Error: El archivo no existe en la ruta especificada: '{path}'."
                self.logger.error(error_msg, exc_info=True)
                self.tomar_captura(f"{nombre_base}_archivo_no_encontrado", directorio)
                raise FileNotFoundError(error_msg)

        try:
            # Usar expect para asegurar que el elemento esté visible y habilitado
            self.logger.debug(f"\n Esperando que el selector '{selector}' esté visible y habilitado.")
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            self.logger.info(f"\n El selector '{selector}' está visible y habilitado.")

            # Opcional: Resaltar el elemento para depuración
            selector.highlight()
            self.logger.debug(f"\n Elemento con selector '{selector}' resaltado.")

            # Usar set_input_files para cargar el archivo(s)
            # Playwright espera una lista de rutas completas
            self.logger.info(f"\n Adjuntando archivo(s) {file_names_list} al selector '{selector}'.")
            selector.set_input_files(full_file_paths)

            # Construir mensaje de éxito basado en si es uno o varios archivos
            if len(file_names_list) == 1:
                success_msg = f"\n ✅ Archivo '{file_names_list[0]}' cargado exitosamente desde '{base_dir}' en el selector '{selector}'."
                self.logger.info(success_msg)
            else:
                success_msg = f"\n ✅ Archivos {file_names_list} cargados exitosamente desde '{base_dir}' en el selector '{selector}'."
                self.logger.info(success_msg)
            
            self.tomar_captura(f"{nombre_base}_archivos_cargados", directorio) # Mantener la captura con nombre consistente
            self.esperar_fijo(tiempo) # Uso consistente de la función de espera fija

        except Exception as e:
            # Capturar los nombres de archivo para el mensaje de error
            error_files_info = file_names_list[0] if len(file_names_list) == 1 else file_names_list
            error_msg = (
                f"\n❌ FALLO: Error al cargar el archivo(s) '{error_files_info}' desde '{base_dir}' "
                f"en el selector '{selector}': {e}"
            )
            self.logger.error(error_msg, exc_info=True) # Incluir exc_info=True para el traceback
            self.tomar_captura(f"{nombre_base}_Error_cargar_archivo", directorio) # Nombre de captura corregido si es necesario
            raise # Re-lanza la excepción para que el test falle si ocurre un error
        
    #26- Función para remover carga de archivo(s)
    def remover_carga_de_archivo(self, selector, nombre_base, directorio, tiempo= 3):
        self.logger.info(f"\n Intentando remover la carga de archivo para el selector: '{selector}'.")
        try:
            # Usar expect para asegurar que el elemento esté visible y habilitado
            self.logger.debug(f"\n Esperando que el selector '{selector}' esté visible y habilitado para remover la carga.")
            expect(selector).to_be_visible()
            expect(selector).to_be_enabled()
            self.logger.info(f"\n El selector '{selector}' está visible y habilitado.")

            # Resaltar el elemento para depuración
            selector.highlight()
            self.logger.debug(f"\n Elemento con selector '{selector}' resaltado.")

            # Usar set_input_files con una lista vacía para remover el archivo
            self.logger.info(f"\n Estableciendo input files a vacío para el selector '{selector}'.")
            selector.set_input_files([])

            self.logger.info(f"\n ✅ Carga de archivo removida exitosamente para el selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_remocion_completa", directorio)
            
            self.esperar_fijo(tiempo) # Uso consistente de la función de espera fija

        except Exception as e:
            error_msg = f"\n ❌ FALLO: Error al remover la carga del archivo para el selector '{selector}': {e}"
            self.logger.error(error_msg, exc_info=True) # Log con información de la excepción
            self.tomar_captura(f"{nombre_base}_error_en_remocion_completa", directorio)
            raise # Re-lanza la excepción para que el test falle si ocurre un error
        
    #27- Función para contar filas y columnas
    def obtener_dimensiones_tabla(self, selector, nombre_base, directorio, tiempo= 1) -> tuple[int, int]:
        # Intentar obtener información útil del selector para los logs
        selector_info = selector.get_attribute('id') or selector.get_attribute('name')
        if not selector_info:
            try:
                # Si no hay id/name, intentar obtener el HTML externo de la etiqueta principal
                selector_info = selector.evaluate("el => el.outerHTML.split('>')[0] + '>'")
            except Exception:
                selector_info = "\n Tabla sin identificador/texto visible"

        self.logger.info(f"\n Obteniendo dimensiones de la tabla con selector: '{selector_info}'")

        try:
            selector.highlight()
            self.logger.debug(f"\n Tabla con selector '{selector_info}' resaltada.")

            # Contar el número de filas de datos (dentro de tbody)
            filas_datos = selector.locator("tbody tr")
            num_filas = filas_datos.count()
            self.logger.debug(f"\n Filas de datos encontradas (tbody tr): {num_filas}.")

            # Contar el número de columnas
            num_columnas = 0
            headers = selector.locator("th")
            if headers.count() > 0:
                num_columnas = headers.count()
                self.logger.debug(f"\n Columnas contadas desde encabezados (th): {num_columnas}.")
            else:
                # Si no hay thead/th, intenta contar td's de la primera fila de datos
                first_row_tds = selector.locator("tr").first.locator("td")
                if first_row_tds.count() > 0:
                    num_columnas = first_row_tds.count()
                    self.logger.debug(f"\n Columnas contadas desde celdas de la primera fila (td): {num_columnas}.")
                else:
                    self.logger.warning(f"\n ADVERTENCIA: No se pudieron encontrar encabezados (th) ni celdas (td) en la primera fila "
                                        f"para la tabla con selector '{selector_info}'. Asumiendo 0 columnas.")
                    # En este caso, num_columnas seguirá siendo 0.

            self.tomar_captura(f"{nombre_base}_dimensiones_obtenidas", directorio)
            self.logger.info(f"\n ✅ ÉXITO: Dimensiones de la tabla '{selector_info}' obtenidas.")
            self.logger.info(f"--> Filas encontradas: {num_filas}")
            self.logger.info(f"--> Columnas encontradas: {num_columnas}")
            self.esperar_fijo(tiempo) # Uso consistente de la función de espera fija

            return (num_filas, num_columnas)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo obtener las dimensiones de la tabla con selector '{selector_info}'.\n"
                f"La tabla o sus elementos internos no estuvieron disponibles a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dimensiones_timeout", directorio)
            return (-1, -1)

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al intentar obtener las dimensiones de la tabla con selector '{selector_info}'.\n"
                f"Posibles causas: Selector de tabla inválido, estructura de tabla inesperada.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dimensiones_error_playwright", directorio)
            raise # Relanzar porque es un error de Playwright

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al obtener las dimensiones de la tabla con selector '{selector_info}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_dimensiones_error_inesperado", directorio)
            raise # Relanzar por ser un error inesperado
        
    #28- Función para buscar datos parcial e imprimir la fila
    def busqueda_coincidencia_e_imprimir_fila(self, table_selector, texto_buscado, nombre_base, directorio, tiempo= 1) -> bool:
        self.logger.info(f"\n Iniciando búsqueda de coincidencia parcial para '{texto_buscado}' en la tabla con selector: '{table_selector}'.")
        encontrado = False
        try:
            expect(table_selector).to_be_visible(timeout=tiempo * 1000)
            self.logger.debug(f"\n Tabla con selector '{table_selector}' está visible.")
            table_selector.highlight()

            filas = table_selector.locator("tbody tr")
            num_filas = filas.count()
            self.logger.debug(f"\n Número de filas de datos encontradas en la tabla: {num_filas}.")

            for i in range(num_filas):
                fila = filas.nth(i)
                fila_texto = fila.text_content()
                self.logger.debug(f"\n Analizando fila {i+1}: '{fila_texto}'.")

                if texto_buscado.lower() in fila_texto.lower():
                    self.logger.info(f"\n ✅ ÉXITO: Texto '{texto_buscado}' encontrado (coincidencia parcial) en la fila {i+1}.")
                    self.logger.info(f"Contenido completo de la fila: '{fila_texto}'")
                    fila.highlight()
                    self.tomar_captura(f"{nombre_base}_coincidencia_parcial_encontrada_fila_{i+1}", directorio)
                    encontrado = True
                    # Si solo se necesita encontrar la primera coincidencia y terminar, descomentar el 'break'
                    # break 
            
            if not encontrado:
                self.logger.info(f"\n ℹ️ Texto '{texto_buscado}' (coincidencia parcial) NO encontrado en ninguna fila de la tabla.")
                self.tomar_captura(f"{nombre_base}_coincidencia_parcial_no_encontrada", directorio)

            self.esperar_fijo(tiempo)
            return encontrado

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla con selector '{table_selector}' "
                f"o sus filas no estuvieron disponibles a tiempo durante la búsqueda de '{texto_buscado}'.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_coincidencia_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al buscar coincidencia para '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Posibles causas: Selector de tabla inválido, estructura de tabla inesperada, o problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_coincidencia_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al buscar coincidencia para '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_coincidencia_error_inesperado", directorio)
            raise
        
    #29- Función para buscar datos exacto e imprimir la fila
    def busqueda_estricta_imprimir_fila(self, table_selector, texto_buscado, nombre_base, directorio, tiempo= 1) -> bool:
        self.logger.info(f"\n Iniciando búsqueda estricta para '{texto_buscado}' en la tabla con selector: '{table_selector}'.")
        encontrado = False
        try:
            expect(table_selector).to_be_visible(timeout=tiempo * 1000)
            self.logger.debug(f"\n Tabla con selector '{table_selector}' está visible.")
            table_selector.highlight()

            filas = table_selector.locator("tbody tr")
            num_filas = filas.count()
            self.logger.debug(f"\n Número de filas de datos encontradas en la tabla: {num_filas}.")

            for i in range(num_filas):
                fila = filas.nth(i)
                celdas = fila.locator("td") # Asumiendo celdas de datos son 'td'
                num_celdas = celdas.count()
                fila_texto_completo = ""
                self.logger.debug(f"\n Analizando fila {i+1} para búsqueda estricta.")

                for j in range(num_celdas):
                    celda = celdas.nth(j)
                    celda_texto = celda.text_content().strip() # Eliminar espacios en blanco alrededor
                    fila_texto_completo += celda_texto + " | " # Concatenar para imprimir la fila completa

                    if celda_texto == texto_buscado: # Coincidencia estricta
                        self.logger.info(f"\n ✅ ÉXITO: Texto '{texto_buscado}' encontrado (coincidencia estricta) en la celda {j+1} de la fila {i+1}.")
                        self.logger.info(f"Contenido completo de la fila: '{fila_texto_completo.strip(' | ')}'")
                        celda.highlight() # Resaltar la celda donde se encontró la coincidencia
                        fila.highlight() # También resaltar la fila para mejor visibilidad
                        self.tomar_captura(f"{nombre_base}_coincidencia_estricta_encontrada_fila_{i+1}_celda_{j+1}", directorio)
                        encontrado = True
                        # Si solo se necesita encontrar la primera coincidencia y terminar, descomentar el 'break'
                        # break # Rompe el bucle de celdas
                
                if encontrado:
                    # break # Rompe el bucle de filas si se encontró una coincidencia y se desea parar
                    pass # Si se desea seguir buscando en otras filas

            if not encontrado:
                self.logger.info(f"\n ℹ️ Texto '{texto_buscado}' (coincidencia estricta) NO encontrado en ninguna celda de la tabla.")
                self.tomar_captura(f"{nombre_base}_coincidencia_estricta_no_encontrada", directorio)

            self.esperar_fijo(tiempo)
            return encontrado

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla con selector '{table_selector}' "
                f"o sus elementos internos no estuvieron disponibles a tiempo durante la búsqueda estricta de '{texto_buscado}'.\n"
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_estricta_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al buscar estrictamente '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Posibles causas: Selector de tabla inválido, estructura de tabla inesperada, o problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_estricta_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al buscar estrictamente '{texto_buscado}' "
                f"en la tabla con selector '{table_selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_busqueda_estricta_error_inesperado", directorio)
            raise
        
    #30- Función para valida que todos los valores en una columna específica de una tabla sean numéricos
    def verificar_precios_son_numeros(self, tabla_selector, columna_nombre, nombre_base, directorio, tiempo_espera_celda: int = 5000,tiempo = 1) -> bool:
        self.logger.info(f"⚙️ Verificando que todos los precios en la columna '{columna_nombre}' son números.")
        try:
            print(f"\n ⚙️ Verificando que todos los precios en la columna '{columna_nombre}' son números.")

            # No necesitamos volver a llamar a table_locator.to_be_visible() si ya se hizo antes de pasarla
            # Si se pasa como un selector string, entonces sí, el primer to_be_visible() es necesario.
            # Asumiendo que 'tabla_locator' ya es un Locator visible cuando se pasa.
            tabla_selector.highlight()
            self.logger.debug(f"\n Tabla resaltada para verificación: {tabla_selector}")

            # --- PASO CRÍTICO: Esperar a que el tbody exista y tenga contenido ---
            # Esperar a que al menos el tbody y la primera fila dentro de él sean visibles.
            # Esto maneja casos donde el tbody se carga dinámicamente o sus filas tardan en aparecer.
            tbody_locator = tabla_selector.locator("tbody")
            expect(tbody_locator).to_be_visible(timeout=15000)
            self.logger.info("\n ✅ El tbody de la tabla es visible.")
            expect(tbody_locator.locator("tr").first).to_be_visible(timeout=10000)
            self.logger.info("\n ✅ Al menos la primera fila de datos en la tabla es visible.")


            # 2. Encontrar el índice de la columna "Price"
            # Nos aseguramos de que los headers también estén visibles si no lo estaban ya
            headers = tabla_selector.locator("th")
            expect(headers.first).to_be_visible(timeout=5000)

            col_index = -1
            header_texts = []
            for i in range(headers.count()):
                header_text = headers.nth(i).text_content().strip()
                header_texts.append(header_text)
                if header_text == columna_nombre:
                    col_index = i
            
            self.logger.info(f"\n 🔍 Cabeceras encontradas: {header_texts}")

            if col_index == -1:
                self.logger.error(f"\n ❌ Error: No se encontró la columna '{columna_nombre}' en la tabla. Cabeceras disponibles: {header_texts}")
                self.tomar_captura(f"{nombre_base}_columna_no_encontrada", directorio)
                return False

            self.logger.info(f"\n 🔍 Columna '{columna_nombre}' encontrada en el índice: {col_index}")

            # 3. Obtener todas las filas de la tabla (excluyendo la cabecera)
            rows = tbody_locator.locator("tr")
            num_rows = rows.count()
            if num_rows == 0:
                self.logger.warning("\n ⚠️ Advertencia: La tabla no contiene filas de datos.")
                self.tomar_captura(f"{nombre_base}_tabla_vacia", directorio)
                return True

            self.logger.info(f"\n 🔍 Se encontraron {num_rows} filas de datos.")

            all_prices_are_numbers = True
            for i in range(num_rows):
                # Localizar la celda de precio en la fila actual
                row_locator = rows.nth(i)
                # Aquí también puedes añadir expect(row_locator).to_be_visible() si las filas individuales se "desvanecen"
                
                price_cell = row_locator.locator(f"td").nth(col_index)
                
                # --- AGREGAR UN EXPECT.TO_BE_VISIBLE() PARA LA CELDA ESPECÍFICA ---
                # Esto es crucial si las celdas se cargan de forma diferida.
                # El error indica que esta aserción falla porque el elemento no se encuentra.
                # Lo reubicamos y mejoramos la captura de error.
                expect(price_cell).to_be_visible(timeout=tiempo_espera_celda) 

                price_text = price_cell.text_content().strip()
                price_cell.highlight() # Resaltar la celda actual

                self.logger.debug(f"\n  Procesando fila {i+1}, precio: '{price_text}'")

                try:
                    float(price_text)
                    self.logger.debug(f"\n  ✅ '{price_text}' es un número válido.")
                except ValueError:
                    self.logger.error(f"\n  ❌ Error: '{price_text}' en la fila {i+1} no es un número válido.")
                    self.tomar_captura(f"{nombre_base}_precio_invalido_fila_{i+1}", directorio)
                    all_prices_are_numbers = False
                    # Continuamos para encontrar todos los errores, no solo el primero.

            if all_prices_are_numbers:
                self.logger.info(f"\n ✅ Todos los precios en la columna '{columna_nombre}' son números válidos.")
                self.tomar_captura(f"{nombre_base}_precios_ok", directorio)
                time.sleep(tiempo)
                return True
            else:
                self.logger.error(f"\n ❌ Se encontraron precios no numéricos en la columna '{columna_nombre}'.")
                # No lanzamos la excepción aquí, solo retornamos False. Si el llamador quiere que falle, debe verificar el valor de retorno.
                return False

        except TimeoutError as e:
            self.logger.error(f"\n ❌ FALLO (Timeout): La tabla o sus elementos no se volvieron visibles a tiempo. Error: {e}")
            self.tomar_captura(f"{nombre_base}_timeout_verificacion_precios", directorio)
            raise AssertionError(f"\n Elementos de la tabla no disponibles a tiempo para verificación de precios: {tabla_selector}") from e
        
        except Error as e:
            self.logger.error(f"\n ❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al verificar los precios. Detalles: {e}")
            self.tomar_captura(f"{nombre_base}_playwright_error_verificacion_precios", directorio)
            raise AssertionError(f"\n Error de Playwright al verificar precios: {tabla_selector}") from e
        
        except Exception as e:
            self.logger.error(f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al verificar los precios en la tabla. Error: {type(e).__name__}: {e}")
            self.tomar_captura(f"{nombre_base}_excepcion_inesperada", directorio)
            raise AssertionError(f"\n Error inesperado al verificar precios: {tabla_selector}") from e
        
    #31- Función para extrae y retorna el valor de un elemento dado su selector.
    def obtener_valor_elemento(self, selector, nombre_base, directorio, tiempo= 0.5) -> str | None:
        self.logger.info(f"\n ⚙️ Extrayendo valor del elemento con selector: '{selector}'")
        valor_extraido = None
        
        try:
            selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_extraccion_valor", directorio)

            # Usamos expect para asegurar que el elemento es visible y habilitado antes de intentar extraer
            expect(selector).to_be_visible(timeout=5000)
            self.logger.debug(f"\n Elemento '{selector}' es visible.")
            expect(selector).to_be_enabled(timeout=5000)
            self.logger.debug(f"\n Elemento '{selector}' es habilitado.")

            # Priorizamos input_value para campos de formulario (incluyendo <select>)
            try:
                # Playwright's input_value() es lo que necesitas para <select> 'value'
                valor_extraido = selector.input_value(timeout=1000)
                self.logger.debug(f"\n Valor extraído (input_value) de '{selector}': '{valor_extraido}'")
            except Error: # Usar Error para errores específicos de Playwright (e.g., no es un elemento de entrada)
                self.logger.debug(f"\n input_value no aplicable o falló para '{selector}'. Intentando text_content/inner_text.")
                
                # Si falla input_value, intentamos con text_content o inner_text para otros elementos
                try:
                    valor_extraido = selector.text_content(timeout=1000)
                    if valor_extraido is not None:
                        # Si text_content devuelve solo espacios en blanco o es vacío,
                        # intentamos inner_text (que a veces es más preciso para texto visible)
                        if valor_extraido.strip() == "":
                            valor_extraido = selector.inner_text(timeout=1000)
                            self.logger.debug(f"\n Valor extraído (inner_text) de '{selector}': '{valor_extraido}'")
                        else:
                            self.logger.debug(f"\n Valor extraído (text_content) de '{selector}': '{valor_extraido}'")
                    else:
                        valor_extraido = selector.inner_text(timeout=1000)
                        self.logger.debug(f"\n Valor extraído (inner_text) de '{selector}': '{valor_extraido}'")
                except Error:
                    self.logger.warning(f"\n No se pudo extraer input_value, text_content ni inner_text de '{selector}'.")
                    valor_extraido = None # Asegurarse de que sea None si todo falla

            if valor_extraido is not None:
                # Stripping whitespace for cleaner results if it's a string
                valor_final = valor_extraido.strip() if isinstance(valor_extraido, str) else valor_extraido
                self.logger.info(f"\n ✅ Valor final obtenido del elemento '{selector}': '{valor_final}'")
                self.tomar_captura(f"{nombre_base}_valor_extraido_exito", directorio)
                return valor_final
            else:
                self.logger.warning(f"\n ❌ No se pudo extraer ningún valor significativo del elemento '{selector}'.")
                self.tomar_captura(f"{nombre_base}_fallo_extraccion_valor_no_encontrado", directorio)
                return None

        except TimeoutError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Timeout): El elemento '{selector}' "
                f"no se volvió visible/habilitado a tiempo para extraer su valor. Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_extraccion_valor", directorio)
            raise AssertionError(f"\n Elemento no disponible para extracción de valor: {selector}") from e

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al intentar extraer el valor de '{selector}'. Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_extraccion_valor", directorio)
            raise AssertionError(f"\n Error de Playwright al extraer valor: {selector}") from e

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar extraer el valor de '{selector}'. Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_extraccion_valor", directorio)
            raise AssertionError(f"\n Error inesperado al extraer valor: {selector}") from e
        
    #32- Función para verificar que los encabezados de las columnas de una tabla sean correctos y estén presentes.
    def verificar_encabezados_tabla(self, selector, encabezados_esperados: list[str], nombre_base, directorio, tiempo= 1) -> bool:
        self.logger.info(f"\n Verificando encabezados de la tabla con selector '{selector}'...")
        
        try:
            # 1. Verificar la presencia de la tabla misma
            table_locator = selector
            # Esperar a que la tabla esté visible. Esto es crucial para evitar errores prematuros.
            #expect(table_locator).to_be_visible(timeout=5000) # Ajusta el timeout según necesidad

            # 2. Verificar la presencia del elemento thead dentro de la tabla
            # Usamos .locator().count() en vez de .to_have_count(0) para poder manejar el caso de no existencia
            # sin lanzar un error de expect, ya que queremos manejarlo explícitamente.
            thead_locator = table_locator.locator("thead") # Correcto: encadenamiento de Locators
            num_theads_actuales = thead_locator.count()

            if num_theads_actuales == 0:
                self.logger.error(f"\n ❌ --> FALLO: La tabla con selector '{selector}' no contiene un elemento '<thead>' (cabecera).")
                self.tomar_captura(f"{nombre_base}_no_thead_encontrado", directorio)
                return False

            # Si llegamos aquí, significa que thead existe. Ahora buscamos los th dentro de él.
            encabezados_actuales_locators = thead_locator.locator("th")
            encabezados_actuales_locators.highlight() # Resaltar la fila encontrada   
            num_encabezados_actuales = encabezados_actuales_locators.count()
            num_encabezados_esperados = len(encabezados_esperados)

            if num_encabezados_actuales == 0:
                # Este caso cubre que el thead existe, pero está vacío de th, o que los th no se encontraron
                # dentro del thead. Es una advertencia, pero se convierte en fallo si no coincide con los esperados.
                self.logger.warning(f"\n ⚠️  --> ADVERTENCIA: Se encontró el '<thead>', pero no se encontraron elementos '<th>' dentro con el selector '{selector} thead th'.")
                self.tomar_captura(f"{nombre_base}_no_encabezados_th_encontrados", directorio)
                if num_encabezados_esperados > 0: # Si se esperaban encabezados pero no se encontraron th
                    self.logger.error(f"\n ❌ --> FALLO: Se esperaban {num_encabezados_esperados} encabezados, pero no se encontraron '<th>' dentro de la cabecera.")
                    return False
                else: # Si no se esperaban encabezados (lista vacía) y no hay th, entonces es un éxito.
                    self.logger.info("\n ✅ ÉXITO: No se esperaban encabezados y no se encontraron '<th>' dentro de la cabecera.")
                    self.tomar_captura(f"{nombre_base}_no_encabezados_esperados_y_no_th", directorio)
                    return True

            if num_encabezados_actuales != num_encabezados_esperados:
                self.logger.error(f"\n ❌ --> FALLO: El número de encabezados '<th>' encontrados ({num_encabezados_actuales}) "
                      f"no coincide con el número de encabezados esperados ({num_encabezados_esperados}).")
                self.tomar_captura(f"{nombre_base}_cantidad_encabezados_incorrecta", directorio)
                return False

            todos_correctos = True
            for i in range(num_encabezados_esperados):
                encabezado_locator = encabezados_actuales_locators.nth(i)
                # Usamos all_text_contents() si quieremos comparar sin espacios/newlines y en orden
                # O text_content().strip() si es uno a uno como lo tienes.
                texto_encabezado_actual = encabezado_locator.text_content().strip()
                encabezado_esperado = encabezados_esperados[i]

                if texto_encabezado_actual == encabezado_esperado:
                    print(f"\n  ✅ Encabezado {i+1}: '{texto_encabezado_actual}' coincide con el esperado '{encabezado_esperado}'.")
                    # encabezado_locator.highlight() # Opcional: resaltar el encabezado
                    time.sleep(tiempo) # Pausa para ver el resaltado
                else:
                    self.logger.error(f"\n  ❌ FALLO: Encabezado {i+1} esperado era '{encabezado_esperado}', pero se encontró '{texto_encabezado_actual}'.")
                    encabezado_locator.highlight()
                    self.tomar_captura(f"{nombre_base}_encabezado_incorrecto_{i+1}", directorio)
                    todos_correctos = False
                    time.sleep(tiempo) # Pausa para ver el resaltado

            if todos_correctos:
                self.logger.info("\n ✅ ÉXITO: Todos los encabezados de columna son correctos y están en el orden esperado.")
                self.tomar_captura(f"{nombre_base}_encabezados_verificados_ok", directorio)
            else:
                self.logger.error("\n ❌ FALLO: Uno o más encabezados de columna son incorrectos o no están en el orden esperado.")
                self.tomar_captura(f"{nombre_base}_encabezados_verificados_fallo", directorio)

            return todos_correctos

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla o sus encabezados con el selector '{selector}'.\n"
                f"Posiblemente la tabla no estuvo disponible a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_verificar_encabezados_timeout", directorio)
            return False

        except Error as e: # Catch Playwright-specific errors
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al intentar verificar la tabla o sus encabezados con el selector '{selector}'.\n"
                f"Posibles causas: Selector inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_verificar_encabezados_error_playwright", directorio)
            raise # Relanzar por ser un error de Playwright que podría indicar un problema mayor.

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar los encabezados de la tabla con el selector '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_verificar_encabezados_error_inesperado", directorio)
            raise # Relanzar por ser un error inesperado.
        
    #33- Función para verificar los datos de las filas de una tabla
    def verificar_datos_filas_tabla(self, selector, datos_filas_esperados: list[dict], nombre_base, directorio, tiempo: int = 1) -> bool:
        self.logger.info(f"\n --- Iniciando verificación de datos de las filas de la tabla con locator '{selector}' ---")
        self.tomar_captura(f"{nombre_base}_inicio_verificacion_datos_filas", directorio) # Renombrado para claridad

        try:
            # Asegurarse de que la tabla está visible y disponible
            expect(selector).to_be_visible(timeout=10000) # Ajusta el timeout si es necesario
            self.logger.info("\n Tabla visible. Procediendo a verificar los datos.")

            # Obtener los encabezados para mapear los índices de las columnas
            header_locators = selector.locator("thead th")
            headers = [h.text_content().strip() for h in header_locators.all()]
            
            if not headers:
                self.logger.error(f"\n ❌ --> FALLO: No se encontraron encabezados en la tabla con locator '{selector}'. No se pueden verificar los datos de las filas.")
                self.tomar_captura(f"{nombre_base}_no_headers_para_datos_filas", directorio)
                return False

            # Obtener todas las filas del cuerpo de la tabla (excluyendo thead)
            row_locators = selector.locator("tbody tr")
            num_filas_actuales = row_locators.count()
            num_filas_esperadas = len(datos_filas_esperados)

            if num_filas_actuales == 0 and num_filas_esperadas == 0:
                self.logger.info("\n ✅ ÉXITO: No se esperaban filas y no se encontraron filas en la tabla.")
                self.tomar_captura(f"{nombre_base}_no_rows_expected_and_found", directorio)
                return True
            
            if num_filas_actuales != num_filas_esperadas:
                self.logger.error(f"\n ❌ --> FALLO: El número de filas encontradas ({num_filas_actuales}) "
                      f"no coincide con el número de filas esperadas ({num_filas_esperadas}).")
                self.tomar_captura(f"{nombre_base}_cantidad_filas_incorrecta", directorio)
                return False

            # --- Variable principal para el retorno ---
            todos_los_datos_correctos = True 

            for i in range(num_filas_esperadas):
                fila_actual_locator = row_locators.nth(i)
                datos_fila_esperada = datos_filas_esperados[i]
                self.logger.info(f"\n   Verificando Fila {i+1} (ID esperado: {datos_fila_esperada.get('ID', 'N/A')})...")
                fila_actual_locator.highlight() # Resaltar la fila actual en la captura para debug.

                # Bandera para saber si la fila actual tiene algún fallo
                fila_actual_correcta = True 

                for col_name, expected_value in datos_fila_esperada.items():
                    try:
                        # Encontrar el índice de la columna por su nombre
                        if col_name not in headers:
                            self.logger.error(f"\n   ❌ FALLO: Columna '{col_name}' esperada no encontrada en los encabezados de la tabla. Encabezados actuales: {headers}")
                            self.tomar_captura(f"{nombre_base}_columna_no_encontrada", directorio)
                            todos_los_datos_correctos = False # Falla general
                            fila_actual_correcta = False # Falla en esta fila
                            continue # Pasa a la siguiente columna esperada o fila

                        col_index = headers.index(col_name)
                        
                        # Localizar la celda específica (td) dentro de la fila por su índice
                        celda_locator = fila_actual_locator.locator("td").nth(col_index)
                        
                        if col_name == "Select": # Lógica específica para el checkbox
                            checkbox_locator = celda_locator.locator("input[type='checkbox']")
                            if checkbox_locator.count() == 0:
                                self.logger.error(f"\n   ❌ FALLO: Checkbox no encontrado en la columna '{col_name}' de la Fila {i+1}.")
                                checkbox_locator.highlight()
                                self.tomar_captura(f"{nombre_base}_fila_{i+1}_no_checkbox", directorio)
                                todos_los_datos_correctos = False
                                fila_actual_correcta = False
                            elif isinstance(expected_value, bool): # Si se espera un estado específico (True/False)
                                if checkbox_locator.is_checked() != expected_value:
                                    self.logger.error(f"\n   ❌ FALLO: El checkbox de la Fila {i+1}, Columna '{col_name}' estaba "
                                          f"{'marcado' if checkbox_locator.is_checked() else 'desmarcado'}, se esperaba {'marcado' if expected_value else 'desmarcado'}.")
                                    checkbox_locator.highlight()
                                    self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_estado_incorrecto", directorio)
                                    todos_los_datos_correctos = False
                                    fila_actual_correcta = False
                                else:
                                    self.logger.info(f"\n   ✅ Fila {i+1}, Columna '{col_name}': Checkbox presente y estado correcto ({'marcado' if expected_value else 'desmarcado'}).")
                            else: # Si solo se espera que el checkbox exista, pero no se especificó un estado booleano
                                self.logger.info(f"\n   ✅ Fila {i+1}, Columna '{col_name}': Checkbox presente (estado no verificado explícitamente).")
                        else: # Para otras columnas de texto
                            actual_value = celda_locator.text_content().strip()
                            # Aseguramos que expected_value también sea una cadena para la comparación
                            if actual_value != str(expected_value).strip(): 
                                self.logger.error(f"\n   ❌ FALLO: Fila {i+1}, Columna '{col_name}'. Se esperaba '{expected_value}', se encontró '{actual_value}'.")
                                celda_locator.highlight()
                                self.tomar_captura(f"{nombre_base}_fila_{i+1}_col_{col_name}_incorrecta", directorio)
                                todos_los_datos_correctos = False
                                fila_actual_correcta = False
                            else:
                                self.logger.info(f"\n   ✅ Fila {i+1}, Columna '{col_name}': '{actual_value}' coincide con lo esperado.")
                        
                    except Exception as col_e:
                        self.logger.error(f"\n   ❌ ERROR INESPERADO al verificar la columna '{col_name}' de la Fila {i+1}: {col_e}")
                        self.tomar_captura(f"{nombre_base}_error_columna_inesperado", directorio)
                        todos_los_datos_correctos = False
                        fila_actual_correcta = False
                        # Podrías decidir si quieres continuar con el resto de las columnas/filas
                        # o si este error debe detener la verificación.

                # Pausa solo si la fila actual tuvo algún fallo para que la captura sea más útil
                if not fila_actual_correcta:
                    time.sleep(tiempo) 

            # --- Retorno final basado en el estado acumulado ---
            if todos_los_datos_correctos:
                self.logger.info("\n ✅ ÉXITO: Todos los datos de las filas y checkboxes son correctos y están presentes.")
                self.tomar_captura(f"{nombre_base}_datos_filas_verificados_ok", directorio)
                return True
            else:
                self.logger.error("\n ❌ FALLO: Uno o más datos de las filas o checkboxes son incorrectos o faltan.")
                self.tomar_captura(f"{nombre_base}_datos_filas_verificados_fallo", directorio)
                return False

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla o sus filas con el locator '{selector}'.\n"
                f"Posiblemente la tabla no estuvo disponible a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_verificar_datos_filas_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al intentar verificar las filas con el locator '{selector}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_verificar_datos_filas_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar los datos de las filas con el locator '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_verificar_datos_filas_error_inesperado", directorio)
            raise
    
    #34- Función para seleccionar y verificar el estado de checkboxes de filas aleatorias.
    def seleccionar_y_verificar_checkboxes_aleatorios(self, selector_tabla, num_checkboxes_a_interactuar: int, nombre_base, directorio, tiempo: int = 1) -> bool:
        self.logger.info(f"\n Iniciando selección y verificación de {num_checkboxes_a_interactuar} checkbox(es) aleatorio(s) en la tabla con locator '{selector_tabla}'...")
        self.tomar_captura(f"{nombre_base}_inicio_seleccion_checkbox", directorio)

        try:
            # Asegurarse de que la tabla está visible
            expect(selector_tabla).to_be_visible(timeout=10000)
            
            # Obtener todos los locators de los checkboxes en las celdas de la tabla
            all_checkbox_locators = selector_tabla.locator("tbody tr td input[type='checkbox']")
            
            num_checkboxes_disponibles = all_checkbox_locators.count()

            if num_checkboxes_disponibles == 0:
                self.logger.error(f"\n ❌ --> FALLO: No se encontraron checkboxes en la tabla con locator '{selector_tabla.locator('tbody tr td input[type=\"checkbox\"]')}'.")
                self.tomar_captura(f"{nombre_base}_no_checkboxes_encontrados", directorio)
                return False
            
            if num_checkboxes_a_interactuar <= 0:
                self.logger.warning("\n ⚠️  ADVERTENCIA: El número de checkboxes a interactuar es 0 o negativo. No se realizará ninguna acción.")
                return True

            if num_checkboxes_a_interactuar > num_checkboxes_disponibles:
                self.logger.error(f"\n ❌ --> FALLO: Se solicitaron {num_checkboxes_a_interactuar} checkboxes, pero solo hay {num_checkboxes_disponibles} disponibles.")
                self.tomar_captura(f"{nombre_base}_no_suficientes_checkboxes", directorio)
                return False

            self.logger.info(f"\n Se encontraron {num_checkboxes_disponibles} checkboxes. Seleccionando {num_checkboxes_a_interactuar} aleatoriamente...")

            # Seleccionar N índices de checkboxes aleatorios y únicos
            random_indices = random.sample(range(num_checkboxes_disponibles), num_checkboxes_a_interactuar)
            
            todos_correctos = True

            for i, idx in enumerate(random_indices):
                checkbox_to_interact = all_checkbox_locators.nth(idx)
                
                # Resaltar el checkbox actual para la captura/visualización
                checkbox_to_interact.highlight()
                self.tomar_captura(f"{nombre_base}_checkbox_{i+1}_seleccionado_idx_{idx}", directorio)
                time.sleep(tiempo)

                # Obtener el ID del producto asociado a esta fila (asumiendo ID en la primera columna)
                try:
                    row_locator = selector_tabla.locator("tbody tr").nth(idx)
                    product_id = row_locator.locator("td").nth(0).text_content().strip()
                except Exception:
                    product_id = "Desconocido"
                
                initial_state = checkbox_to_interact.is_checked()
                self.logger.info(f"\n  Checkbox del Producto ID: {product_id} (Fila: {idx+1}): Estado inicial {'MARCADO' if initial_state else 'DESMARCADO'}.")

                # --- Lógica para asegurar que el click lo deje en estado 'seleccionado' ---
                if initial_state: # Si ya está marcado, lo desmarcamos primero con un clic
                    self.logger.info(f"\n  El checkbox del Producto ID: {product_id} ya está MARCADO. Haciendo clic para desmarcar antes de seleccionar.")
                    checkbox_to_interact.uncheck()
                    time.sleep(0.5) # Pausa para que el DOM se actualice
                    if checkbox_to_interact.is_checked():
                        self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no se desmarcó correctamente para la interacción.")
                        checkbox_to_interact.highlight()
                        self.tomar_captura(f"{nombre_base}_fila_{idx+1}_no_se_desmarco", directorio)
                        todos_correctos = False
                        continue # Pasa al siguiente checkbox aleatorio
                
                # Ahora el checkbox debería estar DESMARCADO (o siempre lo estuvo)
                self.logger.info(f"\n  Haciendo clic en el checkbox del Producto ID: {product_id} para MARCARLO...")
                checkbox_to_interact.check()
                time.sleep(0.5) # Pausa para que el DOM se actualice

                final_state = checkbox_to_interact.is_checked()
                if not final_state: # Si no está marcado (seleccionado) después del clic
                    self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no cambió a MARCADO después del clic. Sigue DESMARCADO.")
                    checkbox_to_interact.highlight()
                    self.tomar_captura(f"{nombre_base}_fila_{idx+1}_no_se_marco", directorio)
                    todos_correctos = False
                else:
                    self.logger.info(f"\n  ✅ ÉXITO: El checkbox del Producto ID: {product_id} ahora está MARCADO (seleccionado).")
                    self.tomar_captura(f"{nombre_base}_fila_{idx+1}_marcado_ok", directorio)
                
                if not todos_correctos: # Si hubo un fallo en este checkbox, pausa antes del siguiente
                    time.sleep(tiempo)

            if todos_correctos:
                self.logger.info(f"\n ✅ ÉXITO: Todos los {num_checkboxes_a_interactuar} checkbox(es) aleatorio(s) fueron seleccionados y verificados correctamente.")
                self.tomar_captura(f"{nombre_base}_todos_seleccionados_ok", directorio)
            else:
                self.logger.error(f"\n ❌ FALLO: Uno o más checkbox(es) aleatorio(s) no pudieron ser seleccionados o verificados.")
                self.tomar_captura(f"{nombre_base}_fallo_general_seleccion", directorio)

            return todos_correctos

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla o los checkboxes con el locator '{selector_tabla}'.\n"
                f"Posiblemente los elementos no estuvieron disponibles a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_seleccion_checkbox_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al seleccionar y verificar checkboxes en la tabla '{selector_tabla}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_seleccion_checkbox_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al seleccionar y verificar checkboxes aleatorios.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_seleccion_checkbox_error_inesperado", directorio)
            raise
    
    #35- Función para seleccionar y verificar el estado de checkboxes de filas CONSECUTIVAS.
    def seleccionar_y_verificar_checkboxes_consecutivos(self, selector_tabla, start_index: int, num_checkboxes_a_interactuar: int, nombre_base, directorio, tiempo= 1) -> bool:
        self.logger.info(f"\n Iniciando selección y verificación de {num_checkboxes_a_interactuar} checkbox(es) consecutivo(s) "
              f"a partir del índice {start_index} en la tabla con locator '{selector_tabla}'...")
        self.tomar_captura(f"{nombre_base}_inicio_seleccion_consecutiva_checkbox", directorio)

        try:
            # Asegurarse de que la tabla está visible
            expect(selector_tabla).to_be_visible(timeout=10000)
            
            # Obtener todos los locators de los checkboxes en las celdas de la tabla
            all_checkbox_locators = selector_tabla.locator("tbody tr td input[type='checkbox']")
            
            num_checkboxes_disponibles = all_checkbox_locators.count()

            if num_checkboxes_disponibles == 0:
                self.logger.error(f"\n ❌ --> FALLO: No se encontraron checkboxes en la tabla con locator '{selector_tabla.locator('tbody tr td input[type=\"checkbox\"]')}'.")
                self.tomar_captura(f"{nombre_base}_no_checkboxes_encontrados_consec", directorio)
                return False
            
            if num_checkboxes_a_interactuar <= 0:
                self.logger.warning("\n ⚠️  ADVERTENCIA: El número de checkboxes a interactuar es 0 o negativo. No se realizará ninguna acción.")
                return True # Consideramos éxito si no hay nada que hacer

            if start_index < 0 or start_index >= num_checkboxes_disponibles:
                self.logger.error(f"\n ❌ --> FALLO: El 'posición de inicio' ({start_index}) está fuera del rango válido de checkboxes disponibles (0 a {num_checkboxes_disponibles - 1}).")
                self.tomar_captura(f"{nombre_base}_start_index_invalido_consec", directorio)
                return False
            
            if (start_index + num_checkboxes_a_interactuar) > num_checkboxes_disponibles:
                self.logger.error(f"\n ❌ --> FALLO: Se solicitaron {num_checkboxes_a_interactuar} checkboxes a partir del índice {start_index}, "
                      f"pero solo hay {num_checkboxes_disponibles} disponibles. El rango excede los límites de la tabla.")
                self.tomar_captura(f"{nombre_base}_rango_excedido_consec", directorio)
                return False

            self.logger.info(f"\n Interactuando con {num_checkboxes_a_interactuar} checkbox(es) consecutivo(s) "
                  f"desde el índice {start_index} hasta el {start_index + num_checkboxes_a_interactuar - 1}...")
            
            todos_correctos = True

            for i in range(num_checkboxes_a_interactuar):
                current_idx = start_index + i
                checkbox_to_interact = all_checkbox_locators.nth(current_idx)
                
                # Resaltar el checkbox actual para la captura/visualización
                checkbox_to_interact.highlight()
                self.tomar_captura(f"{nombre_base}_checkbox_consecutivo_{i+1}_idx_{current_idx}", directorio)
                time.sleep(tiempo)

                # Obtener el ID del producto asociado a esta fila (asumiendo ID en la primera columna)
                try:
                    row_locator = selector_tabla.locator("tbody tr").nth(current_idx)
                    product_id = row_locator.locator("td").nth(0).text_content().strip()
                except Exception:
                    product_id = "Desconocido"
                
                initial_state = checkbox_to_interact.is_checked()
                self.logger.info(f"\n  Checkbox del Producto ID: {product_id} (Fila: {current_idx+1}): Estado inicial {'MARCADO' if initial_state else 'DESMARCADO'}.")

                # --- Lógica para asegurar que el click lo deje en estado 'seleccionado' ---
                if initial_state: # Si ya está marcado, lo desmarcamos primero con un clic
                    self.logger.info(f"\n  El checkbox del Producto ID: {product_id} ya está MARCADO. Haciendo clic para desmarcar antes de seleccionar.")
                    checkbox_to_interact.uncheck()
                    time.sleep(0.5) # Pausa para que el DOM se actualice
                    if checkbox_to_interact.is_checked():
                        self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no se desmarcó correctamente para la interacción.")
                        checkbox_to_interact.highlight()
                        self.tomar_captura(f"{nombre_base}_fila_{current_idx+1}_no_se_desmarco_consec", directorio)
                        todos_correctos = False
                        continue # Pasa al siguiente checkbox consecutivo
                
                # Ahora el checkbox debería estar DESMARCADO (o siempre lo estuvo)
                self.logger.info(f"\n  Haciendo clic en el checkbox del Producto ID: {product_id} para MARCARLO...")
                checkbox_to_interact.check()
                time.sleep(0.5) # Pausa para que el DOM se actualice

                final_state = checkbox_to_interact.is_checked()
                if not final_state: # Si no está marcado (seleccionado) después del clic
                    self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no cambió a MARCADO después del clic. Sigue DESMARCADO.")
                    checkbox_to_interact.highlight()
                    self.tomar_captura(f"{nombre_base}_fila_{current_idx+1}_no_se_marco_consec", directorio)
                    todos_correctos = False
                else:
                    self.logger.info(f"\n  ✅ ÉXITO: El checkbox del Producto ID: {product_id} ahora está MARCADO (seleccionado).")
                    self.tomar_captura(f"{nombre_base}_fila_{current_idx+1}_marcado_ok_consec", directorio)
                
                if not todos_correctos: # Si hubo un fallo en este checkbox, pausa antes del siguiente
                    time.sleep(tiempo)

            if todos_correctos:
                self.logger.info(f"\n ✅ ÉXITO: Todos los {num_checkboxes_a_interactuar} checkbox(es) consecutivo(s) fueron seleccionados y verificados correctamente.")
                self.tomar_captura(f"{nombre_base}_todos_seleccionados_ok_consec", directorio)
            else:
                self.logger.error(f"\n ❌ FALLO: Uno o más checkbox(es) consecutivo(s) no pudieron ser seleccionados o verificados.")
                self.tomar_captura(f"{nombre_base}_fallo_general_seleccion_consec", directorio)

            return todos_correctos

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla o los checkboxes con el locator '{selector_tabla}'.\n"
                f"Posiblemente los elementos no estuvieron disponibles a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_seleccion_consec_checkbox_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al seleccionar y verificar checkboxes consecutivos en la tabla '{selector_tabla}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_seleccion_consec_checkbox_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al seleccionar y verificar checkboxes consecutivos.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_seleccion_consec_checkbox_error_inesperado", directorio)
            raise
        
    #36- Función para deseleccionar un checkbox ya marcado y verificar su estado.
    def deseleccionar_y_verificar_checkbox_marcado_aleatorio(self, selector_tabla, nombre_base, directorio, tiempo= 1) -> bool:
        self.logger.info(f"\n Iniciando deselección y verificación de TODOS los checkboxes marcados "
              f"en la tabla con locator '{selector_tabla}'...")
        self.tomar_captura(f"{nombre_base}_inicio_deseleccion_todos_marcados", directorio)

        try:
            # Asegurarse de que la tabla está visible
            expect(selector_tabla).to_be_visible(timeout=10000)
            
            # Obtener todos los locators de los checkboxes en las celdas de la tabla
            all_checkbox_locators = selector_tabla.locator("tbody tr td input[type='checkbox']")
            
            num_checkboxes_disponibles = all_checkbox_locators.count()

            if num_checkboxes_disponibles == 0:
                self.logger.error(f"\n ❌ --> FALLO: No se encontraron checkboxes en la tabla con locator '{selector_tabla.locator('tbody tr td input[type=\"checkbox\"]')}'.")
                self.tomar_captura(f"{nombre_base}_no_checkboxes_encontrados_todos", directorio)
                return False
            
            # Recolectar todos los checkboxes que están actualmente marcados para deseleccionar
            checkboxes_to_deselect = []
            for i in range(num_checkboxes_disponibles):
                checkbox = all_checkbox_locators.nth(i)
                if checkbox.is_checked():
                    checkboxes_to_deselect.append({"locator": checkbox, "original_index": i})
            
            if not checkboxes_to_deselect:
                self.logger.warning("\n ⚠️  ADVERTENCIA: No se encontró ningún checkbox actualmente MARCADO en la tabla para deseleccionar. Prueba completada sin acciones.")
                self.tomar_captura(f"{nombre_base}_no_marcados_para_deseleccionar", directorio)
                return True # Consideramos éxito si no hay nada que deseleccionar

            self.logger.info(f"\n Se encontraron {len(checkboxes_to_deselect)} checkbox(es) marcado(s) para deseleccionar.")

            todos_deseleccionados_correctamente = True

            for i, checkbox_info in enumerate(checkboxes_to_deselect):
                checkbox_to_interact = checkbox_info["locator"]
                original_idx = checkbox_info["original_index"]
                
                # Resaltar el checkbox actual
                checkbox_to_interact.highlight()
                self.tomar_captura(f"{nombre_base}_deseleccion_actual_{i+1}_idx_{original_idx}", directorio)
                time.sleep(tiempo)

                # Obtener el ID del producto asociado a esta fila (asumiendo ID en la primera columna)
                try:
                    row_locator = selector_tabla.locator("tbody tr").nth(original_idx)
                    product_id = row_locator.locator("td").nth(0).text_content().strip()
                except Exception:
                    product_id = "Desconocido"
                
                self.logger.info(f"\n  Procesando checkbox del Producto ID: {product_id} (Fila: {original_idx+1}). Estado inicial: MARCADO (esperado).")

                # --- Interacción: Clic para deseleccionar ---
                self.logger.info(f"\n  Haciendo clic en el checkbox del Producto ID: {product_id} para DESMARCARLO...")
                checkbox_to_interact.click()
                time.sleep(0.5) # Pausa para que el DOM se actualice

                final_state = checkbox_to_interact.is_checked()
                if final_state: # Si sigue marcado después del clic
                    self.logger.error(f"\n  ❌ FALLO: El checkbox del Producto ID: {product_id} no cambió a DESMARCADO después del clic. Sigue MARCADO.")
                    checkbox_to_interact.highlight()
                    self.tomar_captura(f"{nombre_base}_fila_{original_idx+1}_no_desmarcado", directorio)
                    todos_deseleccionados_correctamente = False
                else:
                    self.logger.info(f"\n  ✅ ÉXITO: El checkbox del Producto ID: {product_id} ahora está DESMARCADO (deseleccionado).")
                    self.tomar_captura(f"{nombre_base}_fila_{original_idx+1}_desmarcado_ok", directorio)
                
                if not todos_deseleccionados_correctamente:
                    time.sleep(tiempo) # Pausa si hubo un fallo para visualización

            if todos_deseleccionados_correctamente:
                self.logger.info(f"\n ✅ ÉXITO: Todos los {len(checkboxes_to_deselect)} checkbox(es) marcados fueron deseleccionados y verificados correctamente.")
                self.tomar_captura(f"{nombre_base}_todos_deseleccionados_ok", directorio)
            else:
                self.logger.error(f"\n ❌ FALLO: Uno o más checkbox(es) marcados no pudieron ser deseleccionados o verificados.")
                self.tomar_captura(f"{nombre_base}_fallo_general_deseleccion_todos", directorio)

            return todos_deseleccionados_correctamente

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se pudo encontrar la tabla o los checkboxes con el locator '{selector_tabla}'.\n"
                f"Posiblemente los elementos no estuvieron disponibles a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_deseleccion_todos_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al deseleccionar y verificar todos los checkboxes marcados en la tabla '{selector_tabla}'.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_deseleccion_todos_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al deseleccionar y verificar todos los checkboxes marcados.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_deseleccion_todos_error_inesperado", directorio)
            raise
    
    #37- Función para busca un 'texto_a_buscar' en todas las celdas de una tabla (tbody) y, si lo encuentra, intenta marcar el checkbox asociado en la misma fila..   
    def seleccionar_checkbox_por_contenido_celda(self, selector_tabla, texto_a_buscar: str, nombre_base, directorio, case_sensitive: bool = False, timeout: int = 10000, tiempo= 0.5) -> bool:
        self.logger.info(f"\n--- Iniciando búsqueda de '{texto_a_buscar}' en la tabla {selector_tabla} para marcar checkboxes ---")
        self.tomar_captura(f"{nombre_base}_inicio_busqueda_celdas", directorio)

        try:
            # Asegurarse de que la tabla está visible y cargada
            expect(selector_tabla).to_be_visible(timeout=timeout)
            self.logger.info("T\n abla visible. Comenzando a iterar por filas y celdas.")

            # Obtener todas las filas del cuerpo de la tabla
            filas = selector_tabla.locator("tbody tr")
            num_filas = filas.count()

            if num_filas == 0:
                self.logger.error(f"\n ❌ --> FALLO: No se encontraron filas en el 'tbody' de la tabla con locator '{selector_tabla}'.")
                self.tomar_captura(f"{nombre_base}_no_filas_encontradas", directorio)
                return False

            self.logger.info(f"\n Se encontraron {num_filas} filas en la tabla.")
            
            checkboxes_marcados_exitosamente = 0
            
            # Normalizar el texto de búsqueda si no es sensible a mayúsculas/minúsculas
            search_text_normalized = texto_a_buscar if case_sensitive else texto_a_buscar.lower()

            for i in range(num_filas):
                fila_actual = filas.nth(i)
                # Obtener todas las celdas (td) de la fila actual, excluyendo posibles celdas de encabezado (th) si las hubiera en tbody
                celdas = fila_actual.locator("td")
                num_celdas = celdas.count()

                if num_celdas == 0:
                    self.logger.warning(f"\n ADVERTENCIA: La fila {i+1} no contiene celdas (td). Saltando.")
                    continue

                celda_encontrada = False
                for j in range(num_celdas):
                    celda_actual = celdas.nth(j)
                    celda_texto = celda_actual.text_content().strip()
                    
                    # Normalizar el texto de la celda para la comparación
                    celda_texto_normalized = celda_texto if case_sensitive else celda_texto.lower()

                    if search_text_normalized in celda_texto_normalized:
                        self.logger.info(f"\n ✅ Coincidencia encontrada en Fila {i+1}, Celda {j+1}: '{celda_texto}' contiene '{texto_a_buscar}'.")
                        celda_encontrada = True
                        
                        # Buscar el checkbox dentro de la misma fila
                        checkbox_locator = fila_actual.locator("input[type='checkbox']")
                        
                        if checkbox_locator.count() > 0:
                            checkbox = checkbox_locator.first
                            checkbox.highlight()
                            self.tomar_captura(f"{nombre_base}_fila_{i+1}_coincidencia", directorio)
                            time.sleep(tiempo)

                            if not checkbox.is_checked():
                                self.logger.info(f"\n  --> Marcando checkbox en Fila {i+1}...")
                                checkbox.check()
                                time.sleep(tiempo)
                                if checkbox.is_checked():
                                    self.logger.info(f"\n  ✅ Checkbox en Fila {i+1} marcado correctamente.")
                                    checkboxes_marcados_exitosamente += 1
                                    self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_marcado", directorio)
                                else:
                                    self.logger.error(f"\n  ❌ FALLO: No se pudo marcar el checkbox en Fila {i+1}.")
                                    self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_no_marcado", directorio)
                            else:
                                self.logger.warning(f"\n  ⚠️ Checkbox en Fila {i+1} ya estaba marcado. No se requiere acción.")
                                self.tomar_captura(f"{nombre_base}_fila_{i+1}_checkbox_ya_marcado", directorio)
                        else:
                            self.logger.warning(f"\n  ⚠️ ADVERTENCIA: No se encontró un checkbox en la Fila {i+1} a pesar de la coincidencia.")
                        break # Salir del bucle de celdas una vez encontrada la coincidencia en la fila

                if not celda_encontrada:
                    self.logger.info(f"\n  No se encontró '{texto_a_buscar}' en la Fila {i+1}.")

            if checkboxes_marcados_exitosamente > 0:
                self.logger.info(f"\n ✅ ÉXITO: Se marcaron {checkboxes_marcados_exitosamente} checkbox(es) basados en la búsqueda de '{texto_a_buscar}'.")
                self.tomar_captura(f"{nombre_base}_busqueda_finalizada_exito", directorio)
                return True
            else:
                self.logger.warning(f"\n ⚠️ ADVERTENCIA: No se encontraron coincidencias para '{texto_a_buscar}' o no se pudo marcar ningún checkbox.")
                self.tomar_captura(f"{nombre_base}_busqueda_finalizada_sin_marcados", directorio)
                return False

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): La tabla con el locator '{selector_tabla}' no estuvo visible a tiempo (timeout de {timeout}ms).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_timeout_tabla", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error al interactuar con la tabla o los checkboxes.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado durante la búsqueda y marcado de checkboxes.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
        
    #38- Función para verifica que la página inicial esperada esté seleccionada y resaltada en un componente de paginación.
    def verificar_pagina_inicial_seleccionada(self, selector_paginado, texto_pagina_inicial: str, nombre_base, directorio, clase_resaltado: str = "active", timeout: int = 10000) -> bool:
        self.logger.info(f"\n--- Iniciando verificación del estado inicial de la paginación ---")
        self.logger.info(f"Contenedor de paginación locator: '{selector_paginado}'")
        self.logger.info(f"Página inicial esperada: '{texto_pagina_inicial}'")
        self.tomar_captura(f"{nombre_base}_inicio_verificacion_paginacion", directorio)

        try:
            # Asegurarse de que el contenedor de paginación está visible
            expect(selector_paginado).to_be_visible(timeout=timeout)
            self.logger.info("\n Contenedor de paginación visible. Procediendo a verificar la página inicial.")

            # Intentar encontrar el elemento de la página inicial por su texto dentro del contenedor
            # Se usa text= para una coincidencia exacta del texto visible del número de página
            # Es importante asegurarse que este selector apunte al elemento clickable de la página (ej. un <a> o <span> dentro de un <li>)
            # Podría ser necesario ajustar el selector para ser más específico si el texto '1' aparece en otros lugares.
            # Por ejemplo: selector_contenedor_paginacion.locator(f"li a:has-text('{texto_pagina_inicial}')")
            pagina_inicial_locator = selector_paginado.locator(f"text='{texto_pagina_inicial}'").first

            # Esperar a que el elemento de la página inicial esté visible y sea interactuable
            expect(pagina_inicial_locator).to_be_visible(timeout=timeout)
            self.logger.info(f"\n Elemento para la página '{texto_pagina_inicial}' encontrado.")

            # 1. Verificar que la página inicial esperada esté seleccionada (marcada con la clase de resaltado)
            self.logger.info(f"\n Verificando si la página '{texto_pagina_inicial}' tiene la clase de resaltado '{clase_resaltado}'...")
            pagina_inicial_locator.highlight() # Resaltar el elemento para la captura
            self.tomar_captura(f"{nombre_base}_pagina_inicial_encontrada", directorio)

            # --- CAMBIO CLAVE AQUÍ: Usar get_attribute("class") y verificar la subclase ---
            current_classes = pagina_inicial_locator.get_attribute("class")
            
            if current_classes and clase_resaltado in current_classes.split():
                self.logger.info(f"\n  ✅ ÉXITO: La página '{texto_pagina_inicial}' está seleccionada y resaltada con la clase '{clase_resaltado}'.")
                self.tomar_captura(f"{nombre_base}_pagina_inicial_seleccionada_ok", directorio)
                return True
            else:
                self.logger.error(f"\n  ❌ FALLO: La página '{texto_pagina_inicial}' no tiene la clase de resaltado esperada '{clase_resaltado}'.")
                self.logger.info(f"  Clases actuales del elemento: '{current_classes}'")
                self.tomar_captura(f"{nombre_base}_pagina_inicial_no_resaltada", directorio)
                return False

        except TimeoutError as e:
            error_msg = (
                f"\n❌ FALLO (Timeout): El contenedor de paginación '{selector_paginado}' "
                f"o la página inicial '{texto_pagina_inicial}' no estuvieron visibles a tiempo "
                f"(timeout de {timeout}ms).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_timeout_paginacion", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n❌ FALLO (Playwright): Error al interactuar con el componente de paginación.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise # Relanzar la excepción para que el framework de pruebas la maneje

        except Exception as e:
            error_msg = (
                f"\n❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la paginación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise # Relanzar la excepción
        
    #39- Función para navega a un número de página específico en un componente de paginación
    def navegar_y_verificar_pagina(self, selector_paginado, numero_pagina_a_navegar: str, nombre_base, directorio, clase_resaltado: str = "active", timeout: int = 10000, tiempo= 0.5) -> bool:
        self.logger.info(f"\n--- Iniciando navegación a la página '{numero_pagina_a_navegar}' y verificación de resaltado ---")
        self.logger.info(f"Contenedor de paginación locator: '{selector_paginado}'")
        self.tomar_captura(f"{nombre_base}_inicio_navegacion_pagina_{numero_pagina_a_navegar}", directorio)

        try:
            # Asegurarse de que el contenedor de paginación está visible
            expect(selector_paginado).to_be_visible(timeout=timeout)
            self.logger.info("\n Contenedor de paginación visible.")

            # --- 1. Obtener el número total de páginas y la página actual ---
            # Asumimos que los elementos de paginación son 'li' dentro del contenedor y que el último 'li'
            # visible (que no sea "Siguiente" o "Última") contiene el número de la última página.
            # Podrías necesitar ajustar este selector si tu HTML es diferente.
            todos_los_botones_pagina = selector_paginado.locator("li")
            num_botones = todos_los_botones_pagina.count()
            
            # Este locator debería apuntar al elemento que realmente tiene la clase 'active'
            # Basado en tu HTML/JS, la clase 'active' está en el <a>
            pagina_actual_locator = selector_paginado.locator(f"a.{clase_resaltado}").first
            # O si el <li> es el que está marcado, pero necesitas obtener el texto del <a> dentro
            # pagina_actual_locator = selector_contenedor_paginacion.locator(f"li.{clase_resaltado} a").first
            pagina_actual_texto = pagina_actual_locator.text_content().strip() if pagina_actual_locator.is_visible() else "Desconocida"
            self.logger.info(f"\n Página actualmente seleccionada: {pagina_actual_texto}")

            # Calcular el número total de páginas (puede requerir un ajuste si hay botones "Siguiente", "Anterior", etc.)
            total_paginas = 0
            if num_botones > 0:
                # Filtrar elementos que son solo números o el último elemento numérico
                for i in range(num_botones - 1, -1, -1): # Iterar al revés para encontrar el último número
                    btn = todos_los_botones_pagina.nth(i)
                    btn_text = btn.text_content().strip()
                    if btn_text.isdigit(): # Si el texto es un número
                        total_paginas = int(btn_text)
                        break
            
            self.logger.info(f"\n Número total de páginas detectadas: {total_paginas}")

            # --- Condicional 1: Página a ir es mayor que el total de páginas ---
            if total_paginas > 0 and int(numero_pagina_a_navegar) > total_paginas:
                self.logger.warning(f"\n ⚠️ ADVERTENCIA: La página de destino '{numero_pagina_a_navegar}' es mayor que el número total de páginas disponibles '{total_paginas}'.")
                self.tomar_captura(f"{nombre_base}_pagina_destino_fuera_rango", directorio)
                return False # Considerar esto como un fallo o una advertencia, dependiendo del caso de prueba.

            # --- Condicional 2: La página a ir es la misma en la que ya está ubicado ---
            if pagina_actual_texto == numero_pagina_a_navegar:
                self.logger.warning(f"\n ⚠️ ADVERTENCIA: Ya estás en la página '{numero_pagina_a_navegar}'. No se requiere navegación.")
                # Opcional: Podrías verificar que siga resaltada.
                self.tomar_captura(f"{nombre_base}_pagina_destino_actual", directorio)
                return True # Considerar esto un éxito, ya que el estado es el esperado.

            # 1. Encontrar y hacer clic en el botón de la página deseada
            pagina_destino_locator = selector_paginado.locator(
                f"li:has-text('{numero_pagina_a_navegar}') a" # Ajusta este selector si es necesario
            ).first

            expect(pagina_destino_locator).to_be_visible(timeout=timeout)
            expect(pagina_destino_locator).to_be_enabled(timeout=timeout)
            self.logger.info(f"\n Elemento de la página '{numero_pagina_a_navegar}' encontrado y habilitado.")

            pagina_destino_locator.highlight()
            self.tomar_captura(f"{nombre_base}_pagina_a_navegar_encontrada", directorio)
            
            self.logger.info(f"\n Haciendo clic en la página '{numero_pagina_a_navegar}'...")
            pagina_destino_locator.click()
            time.sleep(tiempo) # Pausa para permitir la carga de la página y la aplicación de estilos
            
            self.tomar_captura(f"{nombre_base}_pagina_{numero_pagina_a_navegar}_clic", directorio)

            # 2. Verificar que la página de destino se resalte
            self.logger.info(f"\n Verificando si la página '{numero_pagina_a_navegar}' tiene la clase de resaltado '{clase_resaltado}'...")
            
            pagina_destino_locator.highlight() # Resaltar el elemento para la captura final

            current_classes = pagina_destino_locator.get_attribute("class")
            
            if current_classes and clase_resaltado in current_classes.split():
                self.logger.info(f"\n  ✅ ÉXITO: La página '{numero_pagina_a_navegar}' está seleccionada y resaltada con la clase '{clase_resaltado}'.")
                self.tomar_captura(f"{nombre_base}_pagina_{numero_pagina_a_navegar}_seleccionada_ok", directorio)
                return True
            else:
                self.logger.error(f"\n  ❌ FALLO: La página '{numero_pagina_a_navegar}' no tiene la clase de resaltado esperada '{clase_resaltado}'.")
                self.logger.info(f"  Clases actuales del elemento: '{current_classes}'")
                self.tomar_captura(f"{nombre_base}_pagina_{numero_pagina_a_navegar}_no_resaltada", directorio)
                return False

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): El contenedor de paginación '{selector_paginado}' "
                f"o la página '{numero_pagina_a_navegar}' no estuvieron visibles/interactuables a tiempo "
                f"(timeout de {timeout}ms).\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_timeout_navegacion", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error al interactuar con el componente de paginación durante la navegación.\n"
                f"Posibles causas: Locator inválido, problemas de interacción con el DOM, elemento no clickable.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al navegar y verificar la paginación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
        
    #40- Función para verifica una alerta simple utilizando page.expect_event().
    def verificar_alerta_simple_con_expect_event(self, selector, mensaje_esperado: str, nombre_base, directorio, tiempo_espera= 5) -> bool:
        self.logger.info(f"\n--- Ejecutando con expect_event (Alerta Simple): {nombre_base} ---")
        self.logger.info(f"Verificando alerta al hacer clic en '{selector}'")
        self.logger.info(f"  --> Mensaje de alerta esperado: '{mensaje_esperado}'")

        try:
            self.logger.debug(f"  --> Validando visibilidad y habilitación del botón '{selector}'...")
            expect(selector).to_be_visible(timeout=tiempo_espera * 1000)
            expect(selector).to_be_enabled(timeout=tiempo_espera * 1000)
            selector.highlight()
            time.sleep(0.5)

            self.logger.debug("  --> Preparando expect_event para la alerta y haciendo clic...")
            with self.page.expect_event("dialog", timeout=(tiempo_espera + 5) * 1000) as info_dialogo:
                self.logger.debug(f"  --> Haciendo clic en el botón '{selector}'...")
                selector.click(timeout=tiempo_espera * 1000) # Añadir timeout explícito al click

            dialogo = info_dialogo.value
            self.logger.info(f"\n  --> Alerta detectada. Tipo: '{dialogo.type}', Mensaje: '{dialogo.message}'")

            if dialogo.type != "alert":
                dialogo.accept() # Aceptar para no bloquear si es un tipo inesperado
                self.logger.error(f"\n ⚠️Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'alert'.")
                raise ValueError(f"\n ⚠️Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'alert'.")

            if mensaje_esperado not in dialogo.message:
                self.tomar_captura(f"{nombre_base}_alerta_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n ❌ FALLO: Mensaje de alerta incorrecto.\n"
                    f"  --> Esperado: '{mensaje_esperado}'\n"
                    f"  --> Obtenido: '{dialogo.message}'"
                )
                self.logger.error(error_msg)
                dialogo.accept() # Aceptar para no bloquear antes de fallar
                return False

            dialogo.accept()
            self.logger.info("\n  ✅  --> Alerta ACEPTADA.")

            # Opcional: Verificar el resultado en la página después de la interacción
            # Si tu página muestra un mensaje después de aceptar la alerta, verifícalo aquí.
            # Por ejemplo: expect(self.page.locator("#some_result_element")).to_have_text("Alerta cerrada", timeout=2000)

            self.tomar_captura(f"{nombre_base}_alerta_exitosa", directorio)
            self.logger.info(f"\n ✅  --> ÉXITO: La alerta se mostró, mensaje verificado y aceptada correctamente.")
            time.sleep(0.5)
            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Tiempo de espera excedido): La alerta no apareció después de {tiempo_espera} segundos "
                f"al hacer clic en '{selector}', o la verificación del resultado falló.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_alerta_NO_aparece_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except ValueError as e:
            error_msg = (
                f"\n ❌ FALLO (Validación): Error en la validación de la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_alerta", directorio)
            return False

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
    
    #41- Función para verifica una alerta simple utilizando page.on("dialog") con page.once().
    def verificar_alerta_simple_con_on(self, selector, mensaje_alerta_esperado: str, nombre_base, directorio, tiempo_espera= 5) -> bool:
        self.logger.info(f"\n--- Ejecutando con page.on (Alerta Simple): {nombre_base} ---")
        self.logger.info(f"Verificando alerta simple al hacer clic en el botón '{selector}'")
        self.logger.info(f"  --> Mensaje de alerta esperado: '{mensaje_alerta_esperado}'")

        # Resetear el estado para cada ejecución del test
        self._alerta_detectada = False
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""

        try:
            # Validar que el botón es visible y habilitado antes de hacer clic
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}'...")
            expect(selector).to_be_visible(timeout=tiempo_espera * 1000)
            expect(selector).to_be_enabled(timeout=tiempo_espera * 1000)
            selector.highlight()
            time.sleep(0.5)

            # === Registrar el listener ANTES de la acción ===
            self.logger.debug("\n  --> Registrando listener para la alerta con page.once()...")
            # Usa page.once para que el listener se desregistre automáticamente después de una vez.
            self.page.once("dialog", self._get_simple_alert_handler_for_on())

            # Hacer clic en el botón que dispara la alerta
            self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}'...")
            # Añadir un timeout explícito para el click
            selector.click(timeout=tiempo_espera * 1000)


            # Esperar a que el listener haya detectado y manejado la alerta
            self.logger.debug("\n  --> Esperando a que la alerta sea detectada y manejada por el listener...")
            start_time = time.time()
            while not self._alerta_detectada and (time.time() - start_time) * 1000 < (tiempo_espera * 1000 + 2000):
                time.sleep(0.1)

            if not self._alerta_detectada:
                self.logger.error(f"\n La alerta no fue detectada por el listener después de {tiempo_espera} segundos.")
                raise TimeoutError(f"\n La alerta no fue detectada por el listener después de {tiempo_espera} segundos.")

            # Validaciones después de que el listener ha actuado
            if self._alerta_tipo_capturado != "alert":
                self.logger.error(f"\n ⚠️Tipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'alert'.")
                raise ValueError(f"\n ⚠️Tipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'alert'.")

            if mensaje_alerta_esperado not in self._alerta_mensaje_capturado:
                self.tomar_captura(f"{nombre_base}_alerta_mensaje_incorrecto", directorio)
                self.logger.error(f"\n ❌ FALLO: Mensaje de alerta incorrecto.\n  --> Esperado: '{mensaje_alerta_esperado}'\n  --> Obtenido: '{self._alerta_mensaje_capturado}'")
                return False

            # Opcional: Verificar que el elemento afectado por la alerta ha cambiado o desaparecido
            # En el caso de una alerta simple, a menudo no hay un cambio visible directo en la página,
            # pero si tu aplicación lo hace, deberías añadir una verificación aquí.
            # Por ejemplo, si un div muestra "Alerta cerrada":
            # expect(self.page.locator("#resultado_alerta")).to_have_text("Alerta cerrada", timeout=2000)

            self.tomar_captura(f"{nombre_base}_alerta_exitosa", directorio)
            self.logger.info(f"\n ✅  --> ÉXITO: La alerta se mostró, mensaje verificado y aceptada correctamente.")
            time.sleep(0.5)
            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Tiempo de espera excedido): La alerta no apareció o no fue manejada después de {tiempo_espera} segundos "
                f"al hacer clic en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_alerta_NO_aparece_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except ValueError as e:
            error_msg = (
                f"\n ❌ FALLO (Validación): Error en la validación de la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_alerta", directorio)
            return False

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la alerta.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
        
    #42- Función para verifica una alerta de confirmación utilizando page.expect_event(). Este método maneja el diálogo exclusivamente con expect_event.
    def verificar_confirmacion_expect_event(self, selector, mensaje_esperado: str, accion_confirmacion: str, nombre_base, directorio, tiempo_espera= 5) -> bool:
        self.logger.info(f"\n--- Ejecutando con expect_event (Confirmación): {nombre_base} ---")
        self.logger.info(f"Verificando confirmación al hacer clic en '{selector}' para '{accion_confirmacion}'")
        self.logger.info(f"  --> Mensaje de confirmación esperado: '{mensaje_esperado}'")

        try:
            # Validar que el botón es visible y habilitado antes de hacer clic
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}'...")
            expect(selector).to_be_visible(timeout=tiempo_espera * 1000)
            expect(selector).to_be_enabled(timeout=tiempo_espera * 1000)
            selector.highlight()
            time.sleep(0.5)

            # Usar 'with page.expect_event' para sincronizar el clic con la aparición del diálogo
            self.logger.debug("\n  --> Preparando expect_event para la confirmación y haciendo clic...")
            # Aumentar el timeout aquí para el evento de diálogo y el click
            with self.page.expect_event("dialog", timeout=(tiempo_espera + 10) * 1000) as info_dialogo:
                self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}'...")
                selector.click(timeout=(tiempo_espera + 5) * 1000) # Añadir timeout explícito al click (10s si tiempo_espera es 5)

            # Obtener el objeto Dialog una vez que el evento ocurre
            dialogo = info_dialogo.value
            self.logger.info(f"\n  --> Confirmación detectada. Tipo: '{dialogo.type}', Mensaje: '{dialogo.message}'")

            # Verificar el tipo de diálogo
            if dialogo.type != "confirm":
                if accion_confirmacion == 'accept':
                    dialogo.accept()
                else:
                    dialogo.dismiss()
                self.logger.error(f"\n ⚠️Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'confirm'.")
                raise ValueError(f"\n ⚠️Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'confirm'.")

            # Verificar el mensaje de la alerta
            if mensaje_esperado not in dialogo.message:
                self.tomar_captura(f"{nombre_base}_confirmacion_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n ❌ FALLO: Mensaje de confirmación incorrecto.\n"
                    f"  --> Esperado: '{mensaje_esperado}'\n"
                    f"  --> Obtenido: '{dialogo.message}'"
                )
                self.logger.error(error_msg)
                if accion_confirmacion == 'accept':
                    dialogo.accept()
                else:
                    dialogo.dismiss()
                return False

            # Realizar la acción solicitada (Aceptar o Cancelar)
            if accion_confirmacion == 'accept':
                dialogo.accept()
                self.logger.info("\n  ✅  --> Confirmación ACEPTADA.")
            elif accion_confirmacion == 'dismiss':
                dialogo.dismiss()
                self.logger.info("\n  ✅  --> Confirmación CANCELADA.")
            else:
                self.logger.error(f"\n Acción de confirmación no válida: '{accion_confirmacion}'. Use 'accept' o 'dismiss'.")
                raise ValueError(f"\n Acción de confirmación no válida: '{accion_confirmacion}'. Use 'accept' o 'dismiss'.")

            # Opcional: Verificar el resultado en la página después de la interacción
            if accion_confirmacion == 'accept':
                expect(self.page.locator("#demo")).to_have_text("You pressed OK!", timeout=5000)
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed OK!' verificado.")
            elif accion_confirmacion == 'dismiss':
                expect(self.page.locator("#demo")).to_have_text("You pressed Cancel!", timeout=5000)
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed Cancel!' verificado.")

            self.tomar_captura(f"{nombre_base}_confirmacion_exitosa_{accion_confirmacion}", directorio)
            self.logger.info(f"\n ✅  --> ÉXITO: La confirmación se mostró, mensaje verificado y '{accion_confirmacion}' correctamente.")
            time.sleep(0.5)
            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Tiempo de espera excedido): La confirmación no apareció después de {tiempo_espera} segundos "
                f"al hacer clic en '{selector}', o la verificación del resultado falló.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_confirmacion_NO_aparece_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except ValueError as e:
            error_msg = (
                f"\n ❌ FALLO (Validación): Error en la validación de la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_confirmacion", directorio)
            return False

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
        
    #43- Función para verifica una alerta de confirmación utilizando page.on("dialog") con page.once().
    def verificar_confirmacion_on_dialog(self, selector, mensaje_esperado: str, accion_confirmacion: str, nombre_base, directorio, tiempo_espera= 5) -> bool:
        self.logger.info(f"\n--- Ejecutando con page.on (Confirmación): {nombre_base} ---")
        self.logger.info(f"Verificando confirmación al hacer clic en '{selector}' para '{accion_confirmacion}'")
        self.logger.info(f"  --> Mensaje de confirmación esperado: '{mensaje_esperado}'")

        # Resetear el estado para cada ejecución del test
        self._alerta_detectada = False
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""

        try:
            # Validar que el botón es visible y habilitado antes de hacer clic
            self.logger.debug(f"\n  --> Validando visibilidad y habilitación del botón '{selector}'...")
            expect(selector).to_be_visible(timeout=tiempo_espera * 1000)
            expect(selector).to_be_enabled(timeout=tiempo_espera * 1000)
            selector.highlight()
            time.sleep(0.5)

            # Registrar el listener para la confirmación.
            self.logger.debug("\n  --> Registrando listener para la confirmación con page.once()...")
            self.page.once("dialog", self._get_confirmation_dialog_handler_for_on(accion_confirmacion))

            # Hacer clic en el botón
            self.logger.debug(f"\n  --> Haciendo clic en el botón '{selector}'...")
            selector.click(timeout=(tiempo_espera + 5) * 1000)

            # Esperar a que el listener haya detectado y manejado la confirmación
            self.logger.debug("\n  --> Esperando a que la confirmación sea detectada y manejada por el listener...")
            start_time = time.time()
            # Aumentar el timeout de espera del loop para asegurar que el handler se dispara
            while not self._alerta_detectada and (time.time() - start_time) * 1000 < (tiempo_espera * 1000 + 5000):
                time.sleep(0.1)

            if not self._alerta_detectada:
                self.logger.error(f"\n La confirmación no fue detectada por el listener después de {tiempo_espera} segundos.")
                raise TimeoutError(f"\n La confirmación no fue detectada por el listener después de {tiempo_espera} segundos.")

            # Validaciones después de que el listener ha actuado
            if self._alerta_tipo_capturado != "confirm":
                self.logger.error(f"\n ⚠️Tipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'confirm'.")
                raise ValueError(f"\n ⚠️Tipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'confirm'.")

            if mensaje_esperado not in self._alerta_mensaje_capturado:
                self.tomar_captura(f"{nombre_base}_confirmacion_mensaje_incorrecto", directorio)
                self.logger.error(f"\n ❌ FALLO: Mensaje de confirmación incorrecto.\n  --> Esperado: '{mensaje_esperado}'\n  --> Obtenido: '{self._alerta_mensaje_capturado}'")
                return False

            # Opcional: Verificar el resultado en la página después de la interacción
            if accion_confirmacion == 'accept':
                expect(self.page.locator("#demo")).to_have_text("You pressed OK!", timeout=5000)
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed OK!' verificado.")
            elif accion_confirmacion == 'dismiss':
                expect(self.page.locator("#demo")).to_have_text("You pressed Cancel!", timeout=5000)
                self.logger.info("\n  ✅  --> Resultado en página: 'You pressed Cancel!' verificado.")

            self.tomar_captura(f"{nombre_base}_confirmacion_exitosa_{accion_confirmacion}", directorio)
            self.logger.info(f"\n ✅  --> ÉXITO: La confirmación se mostró, mensaje verificado y '{accion_confirmacion}' correctamente.")
            time.sleep(0.5)
            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Tiempo de espera excedido): La confirmación no apareció después de {tiempo_espera} segundos "
                f"al hacer clic en '{selector}', o la verificación del resultado falló.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_confirmacion_NO_aparece_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except ValueError as e:
            error_msg = (
                f"\n ❌ FALLO (Validación): Error en la validación de la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_confirmacion", directorio)
            return False

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar la confirmación.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
    
    #44- Funció para verificar_prompt_expect_event (Implementación para Prompt Alert con expect_event)
    def verificar_prompt_expect_event(self, selector, mensaje_prompt_esperado: str, input_text: str, accion_prompt: str, nombre_base, directorio, tiempo_espera= 5) -> bool:
        self.logger.info(f"\n--- Ejecutando con expect_event (Prompt Alert): {nombre_base} ---")
        self.logger.info(f"Verificando prompt al hacer clic en '{selector}' para '{accion_prompt}'")
        self.logger.info(f"  --> Mensaje del prompt esperado: '{mensaje_prompt_esperado}'")
        if accion_prompt == 'accept':
            self.logger.info(f"\n  --> Texto a introducir: '{input_text}'")

        try:
            self.logger.info(f"\n  --> Validando visibilidad y habilitación del botón '{selector}'...")
            expect(selector).to_be_visible(timeout=tiempo_espera * 1000)
            expect(selector).to_be_enabled(timeout=tiempo_espera * 1000)
            selector.highlight()
            time.sleep(0.5)

            self.logger.info("\n  --> Preparando expect_event para el prompt y haciendo clic...")
            with self.page.expect_event("dialog", timeout=(tiempo_espera + 10) * 1000) as info_dialogo:
                self.logger.info(f"\n  --> Haciendo clic en el botón '{selector}'...")
                selector.click(timeout=(tiempo_espera + 5) * 1000)

            dialogo = info_dialogo.value
            self.logger.info(f"\n  --> Prompt detectado. Tipo: '{dialogo.type}', Mensaje: '{dialogo.message}'")

            # Verificar el tipo de diálogo
            if dialogo.type != "prompt":
                # Si el tipo es inesperado, intenta cerrarlo para no bloquear el test
                if accion_prompt == 'accept':
                    dialogo.accept()
                else:
                    dialogo.dismiss()
                raise ValueError(f"\n ⚠️Tipo de diálogo inesperado: '{dialogo.type}'. Se esperaba 'prompt'.")

            # Verificar el mensaje del prompt
            if mensaje_prompt_esperado not in dialogo.message:
                self.tomar_captura(f"{nombre_base}_prompt_mensaje_incorrecto", directorio)
                error_msg = (
                    f"\n ❌ FALLO: Mensaje del prompt incorrecto.\n"
                    f"  --> Esperado: '{mensaje_prompt_esperado}'\n"
                    f"  --> Obtenido: '{dialogo.message}'"
                )
                self.logger.error(error_msg)
                # Intenta cerrar el diálogo antes de fallar
                if accion_prompt == 'accept':
                    dialogo.accept()
                else:
                    dialogo.dismiss()
                return False

            # Realizar la acción solicitada (Introducir texto y Aceptar, o Cancelar)
            if accion_prompt == 'accept':
                # --- CORRECCIÓN AQUÍ ---
                dialogo.accept(input_text)
                self.logger.info(f"\n  --> Texto '{input_text}' introducido en el prompt y aceptado.")
            elif accion_prompt == 'dismiss':
                dialogo.dismiss()
                self.logger.info("\n  ✅  --> Prompt CANCELADO.")
            else:
                raise ValueError(f"Acción de prompt no válida: '{accion_prompt}'. Use 'accept' o 'dismiss'.")

            return True # Se agregó retorno explícito True para la ruta de éxito
        
        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Tiempo de espera excedido): El prompt no apareció después de {tiempo_espera} segundos "
                f"al hacer clic en '{selector}', o la verificación del resultado falló.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_prompt_NO_aparece_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except ValueError as e:
            error_msg = (
                f"\n ❌ FALLO (Validación): Error en la validación del prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_prompt", directorio)
            return False

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
        
    #45- Función para verifica una alerta de tipo 'prompt' utilizando page.on("dialog") con page.once().
    def verificar_prompt_on_dialog(self, selector, mensaje_prompt_esperado: str, input_text: str, accion_prompt: str, nombre_base, directorio, tiempo_espera= 5) -> bool:
        self.logger.info(f"\n--- Ejecutando con page.on (Prompt Alert): {nombre_base} ---")
        self.logger.info(f"Verificando prompt al hacer clic en '{selector}' para '{accion_prompt}'")
        self.logger.info(f"  --> Mensaje del prompt esperado: '{mensaje_prompt_esperado}'")
        if accion_prompt == 'accept':
            self.logger.info(f"\n  --> Texto a introducir: '{input_text}'")

        # Resetear el estado para cada ejecución del test
        self._alerta_detectada = False
        self._alerta_mensaje_capturado = ""
        self._alerta_tipo_capturado = ""
        self._alerta_input_capturado = "" # Resetear también el input capturado

        try:
            self.logger.info(f"\n  --> Validando visibilidad y habilitación del botón '{selector}'...")
            expect(selector).to_be_visible(timeout=tiempo_espera * 1000)
            expect(selector).to_be_enabled(timeout=tiempo_espera * 1000)
            selector.highlight()
            time.sleep(0.5)

            # === Registrar el listener ANTES de la acción ===
            self.logger.info("\n  --> Registrando listener para el prompt con page.once()...")
            # Usamos el handler corregido que pasa el input_text a dialog.accept()
            self.page.once("dialog", self._get_prompt_dialog_handler_for_on(input_text, accion_prompt))

            # Hacer clic en el botón que dispara el prompt
            self.logger.info(f"\n  --> Haciendo clic en el botón '{selector}'...")
            selector.click(timeout=(tiempo_espera + 5) * 1000)

            # Esperar a que el listener haya detectado y manejado el prompt
            self.logger.info("\n  --> Esperando a que el prompt sea detectado y manejado por el listener...")
            start_time = time.time()
            while not self._alerta_detectada and (time.time() - start_time) * 1000 < (tiempo_espera * 1000 + 5000):
                time.sleep(0.1)

            if not self._alerta_detectada:
                raise TimeoutError(f"\n El prompt no fue detectado por el listener después de {tiempo_espera} segundos.")

            # Validaciones después de que el listener ha actuado
            if self._alerta_tipo_capturado != "prompt":
                raise ValueError(f"\n ⚠️Tipo de diálogo inesperado: '{self._alerta_tipo_capturado}'. Se esperaba 'prompt'.")

            if mensaje_prompt_esperado not in self._alerta_mensaje_capturado:
                self.tomar_captura(f"{nombre_base}_prompt_mensaje_incorrecto", directorio)
                self.logger.error(f"\n ❌ FALLO: Mensaje del prompt incorrecto.\n  --> Esperado: '{mensaje_prompt_esperado}'\n  --> Obtenido: '{self._alerta_mensaje_capturado}'")
                return False

            # Verificar que el texto introducido (si es el caso) se ha guardado correctamente
            if accion_prompt == 'accept' and self._alerta_input_capturado != input_text:
                self.tomar_captura(f"{nombre_base}_prompt_input_incorrecto", directorio)
                self.logger.error(f"\n ❌ FALLO: Texto introducido en el prompt incorrecto.\n  --> Esperado: '{input_text}'\n  --> Obtenido (capturado): '{self._alerta_input_capturado}'")
                return False

            # >>> Lógica para verificar el resultado en la página ELIMINADA de esta función <<<

            self.tomar_captura(f"{nombre_base}_prompt_exitosa_{accion_prompt}", directorio)
            self.logger.info(f"\n ✅  --> ÉXITO: El prompt se mostró, mensaje verificado y acción '{accion_prompt}' completada.")
            time.sleep(0.5)
            return True

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Tiempo de espera excedido): El prompt no apareció o no fue manejado después de {tiempo_espera} segundos "
                f"al hacer clic en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_prompt_NO_aparece_timeout", directorio)
            return False

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al interactuar con el botón o el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_playwright", directorio)
            raise

        except ValueError as e:
            error_msg = (
                f"\n ❌ FALLO (Validación): Error en la validación del prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_validacion_prompt", directorio)
            return False

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el prompt.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado", directorio)
            raise
        
    #46- Función para espera por una nueva pestaña/página (popup) que se haya abierto y cambia el foco de la instancia 'page' actual a esa nueva pestaña.
    def abrir_y_cambiar_a_nueva_pestana(self, selector_boton_apertura, nombre_base, directorio, tiempo_espera=15) -> Page | None:
        self.logger.info(f"\n 🔄 Preparando para hacer clic y esperar nueva pestaña/popup. Esperando hasta {tiempo_espera} segundos...")

        nueva_pagina = None
        try:
            # Usar page.context.expect_event("page") para esperar la nueva página
            # y realizar la acción de click DENTRO de este contexto.
            # Esto asegura que la página capturada es la que se abre DESPUÉS del click.
            with self.page.context.expect_event("page", timeout=tiempo_espera * 1000) as event_info:
                # Realizar el clic en el botón que abre la nueva pestaña
                self.hacer_click_en_elemento(selector_boton_apertura, f"{nombre_base}_click_para_nueva_pestana", directorio, None)
            
            nueva_pagina = event_info.value # El objeto 'Page' de la nueva pestaña
            
            # Esperar a que la nueva página cargue completamente y el body sea visible
            nueva_pagina.wait_for_load_state()
            nueva_pagina.wait_for_selector("body", timeout=tiempo_espera * 1000)

            self.logger.info(f"\n ✅ Nueva pestaña abierta y detectada: URL = {nueva_pagina.url}, Título = {nueva_pagina.title}")
            
            # Actualizar self.page para que las subsiguientes operaciones usen la nueva página
            self.page = nueva_pagina 
            self.tomar_captura(f"{nombre_base}_nueva_pestana_abierta", directorio)
            
            return nueva_pagina

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ FALLO (Timeout): No se detectó ninguna nueva pestaña/página después de {tiempo_espera} segundos "
                f"al intentar hacer clic en el botón de apertura. Asegúrate de que el clic abre una nueva pestaña.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_no_se_detecto_popup_timeout", directorio)
            return None
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al intentar abrir y cambiar a la nueva pestaña.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado_abrir_pestana", directorio)
            raise

    #47- Función que cierra la pestaña actual y, si hay otras pestañas abiertas en el mismo contexto, cambia el foco a la primera pestaña disponible.
    def cerrar_pestana_actual(self, nombre_base, directorio, tiempo= 1):
        self.logger.info(f"\n 🚪 Cerrando la pestaña actual: URL = {self.page.url}")
        try:
            current_page_url = self.page.url # Guardar la URL antes de cerrar para el log
            
            # ¡IMPORTANTE! Tomar la captura *antes* de cerrar la página.
            # Cambié el sufijo para indicar que es antes del cierre.
            self.tomar_captura(f"{nombre_base}_antes_de_cerrar", directorio) 
            
            self.page.close()
            self.logger.info(f"\n ✅ Pestaña con URL '{current_page_url}' cerrada exitosamente.")
            
            time.sleep(tiempo) # Pequeña espera después de cerrar

            # Si hay otras páginas abiertas en el contexto, intenta cambiar a la primera disponible
            if self.page.context.pages:
                self.page = self.page.context.pages[0]
                self.logger.info(f"\n 🔄 Foco cambiado automáticamente a la primera pestaña disponible: URL = {self.page.url}")
                # Opcional: Podrías tomar otra captura aquí si quieres mostrar el estado de la nueva pestaña activa.
                # self.tomar_captura(f"{nombre_base}_foco_cambiado", directorio)
            else:
                self.logger.warning("\n ⚠️ No hay más pestañas abiertas en el contexto del navegador. La instancia 'page' no apunta a ninguna página activa.")
                self.page = None # No hay página activa

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error al intentar cerrar la pestaña actual.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            # NOTA: Si el error ya es 'Target page, context or browser has been closed',
            # intentar tomar otra captura con self.page.screenshot() aquí también fallará.
            # Por lo tanto, se recomienda NO intentar tomar una captura en el bloque de error
            # si el problema es que la página ya está cerrada.
            # self.tomar_captura(f"{nombre_base}_error_cerrar_pestana", directorio) # Eliminar o comentar esta línea si está aquí
            raise # Re-lanzar la excepción para que el test falle correctamente
        
    #48- Función para hacer clic en un selector y espera que se abran nuevas ventanas/pestañas.
    # Retorna una lista de objetos Page para las nuevas ventanas.
    def hacer_clic_y_abrir_nueva_ventana(self, selector, nombre_base, directorio, nombre_paso="", timeout=30000) -> List[Page]:
        self.logger.info(f"\n {nombre_paso}: Haciendo clic en '{selector}' para abrir nuevas ventanas.")
        self.tomar_captura(f"{nombre_base}_antes_clic_nueva_ventana", directorio)
        
        # Limpiar la lista de páginas antes de la interacción para evitar acumulación
        self._all_new_pages_opened_by_click = []

        try:
            # Crea una tarea que espera por la nueva página antes de hacer el click.
            # Esto es crucial para Playwright: el "listener" debe estar activo ANTES de la acción.
            with self.page.context.expect_event("page", timeout=timeout) as page_info:
                selector.click() # Realiza el click que debería abrir una nueva ventana
            
            # La nueva página se añadió a _all_new_pages_opened_by_click por el _on_new_page handler.
            # Esperar a que la nueva(s) página(s) cargue(n) completamente
            for new_page in self._all_new_pages_opened_by_click:
                new_page.wait_for_load_state("load")
                new_page.wait_for_load_state("domcontentloaded")
                new_page.wait_for_load_state("networkidle")
                self.logger.info(f"\n  --> Nueva página cargada: URL = {new_page.url}")
            
            self.tomar_captura(f"{nombre_base}_despues_clic_nueva_ventana", directorio)
            self.logger.info(f"\n  ✅ Se han detectado y cargado {len(self._all_new_pages_opened_by_click)} nueva(s) ventana(s).")
            return self._all_new_pages_opened_by_click

        except TimeoutError:
            self.logger.error(f"\n ❌ FALLO: No se detectó ninguna nueva ventana después de hacer clic en '{selector}' dentro del tiempo de espera de {timeout/1000} segundos.")
            self.tomar_captura(f"{nombre_base}_no_nueva_ventana", directorio)
            return []

        except Exception as e:
            error_msg = f"\n ❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error al intentar abrir nueva ventana.\nDetalles: {e}"
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_abrir_nueva_ventana", directorio)
            raise # Re-lanzar la excepción para que el test falle correctamente

    #49- Función para cambia el foco del navegador a una ventana/pestaña específica,
    #ya sea por su índice (int) o por una parte de su URL o título (str).
    def cambiar_foco_entre_ventanas(self, nombre_base, directorio, opcion_ventana: Union[int, str], nombre_paso=""):
        self.logger.info(f"\n {nombre_paso}: Intentando cambiar el foco a la ventana/pestaña: '{opcion_ventana}'")
        
        target_page_to_focus: Page = None
        
        try:
            # Obtener todas las páginas actuales en el contexto del navegador
            all_pages_in_context = self.page.context.pages
            self.logger.info(f"\n Ventanas/pestañas abiertas actualmente: {len(all_pages_in_context)}")
            for i, p in enumerate(all_pages_in_context):
                self.logger.info(f"\n  [{i}] URL: {p.url} | Título: {p.title()}")

            if isinstance(opcion_ventana, int):
                if 0 <= opcion_ventana < len(all_pages_in_context):
                    target_page_to_focus = all_pages_in_context[opcion_ventana]
                    self.logger.info(f"\n  --> Seleccionada por índice: {opcion_ventana}")
                else:
                    error_msg = f"\n ❌ FALLO: El índice '{opcion_ventana}' está fuera del rango de pestañas abiertas (0-{len(all_pages_in_context)-1})."
                    self.logger.error(error_msg)
                    self.tomar_captura(f"{nombre_base}_error_indice_invalido", directorio)
                    raise IndexError(error_msg)
            elif isinstance(opcion_ventana, str):
                # Intentar encontrar por URL o título
                for p in all_pages_in_context:
                    if opcion_ventana in p.url or opcion_ventana in p.title():
                        target_page_to_focus = p
                        self.logger.info(f"\n  --> Seleccionada por coincidencia de URL/Título: '{opcion_ventana}'")
                        break
                if not target_page_to_focus:
                    error_msg = f"\n ❌ FALLO: No se encontró ninguna pestaña con la URL o título que contenga '{opcion_ventana}'."
                    self.logger.error(error_msg)
                    self.tomar_captura(f"{nombre_base}_error_no_coincidencia_foco", directorio)
                    raise ValueError(error_msg)
            else:
                error_msg = f"\n ❌ FALLO: El tipo de 'opcion_ventana' no es válido. Debe ser int o str."
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_error_tipo_opcion_foco", directorio)
                raise TypeError(error_msg)

            # Si la página objetivo ya es la página actual, no es necesario cambiar
            if target_page_to_focus == self.page:
                self.logger.info(f"\n ✅ El foco ya está en la ventana seleccionada. No es necesario cambiar.")
            else:
                self.page = target_page_to_focus
                self.logger.info(f"\n ✅ Foco cambiado exitosamente a la ventana/pestaña seleccionada.")
            
            self.logger.info(f"\n   URL de la pestaña actual: {self.page.url}")
            self.logger.info(f"\n   Título de la pestaña actual: {self.page.title()}")
            self.tomar_captura(f"{nombre_base}_foco_cambiado", directorio)

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error al intentar cambiar el foco de ventana.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_cambiar_foco_ventana", directorio)
            raise

    #50- Función que cierra una Page object específica.
    #Intenta cambiar el foco a la primera página disponible si la página cerrada era la actual.
    def cerrar_pestana_especifica(self, page_to_close: Page, nombre_base, directorio, nombre_paso=""):
        self.logger.info(f"\n {nombre_paso}: Intentando cerrar la pestaña con URL: {page_to_close.url}")
        try:
            if not page_to_close.is_closed():
                is_current_page = (self.page == page_to_close)
                closed_url = page_to_close.url
                page_to_close.close()
                self.logger.info(f"\n ✅ Pestaña '{closed_url}' cerrada exitosamente.")
                self.tomar_captura(f"{nombre_base}_pestana_cerrada", directorio)
                
                # Si la página cerrada era la página actual (self.page), cambiar el foco
                if is_current_page:
                    self.logger.info("\n Detectado: La pestaña cerrada era la pestaña activa.")
                    # Buscar la primera página disponible en el contexto
                    if self.page.context.pages:
                        self.page = self.page.context.pages[0]
                        self.logger.info(f"\n 🔄 Foco cambiado automáticamente a la primera pestaña disponible: URL = {self.page.url}")
                        self.tomar_captura(f"{nombre_base}_foco_cambiado_despues_cerrar", directorio)
                    else:
                        self.logger.warning("\n ⚠️ No hay más pestañas abiertas en el contexto del navegador. La instancia 'page' no apunta a ninguna página activa.")
                        self.page = None # No hay página activa
            else:
                self.logger.info(f"ℹ️ La pestaña con URL '{page_to_close.url}' ya estaba cerrada.")

        except Error as e: # Playwright-specific error
            # Esto puede ocurrir si la página ya se cerró por alguna razón externa.
            if "Target page, context or browser has been closed" in str(e):
                self.logger.warning(f"\n ⚠️ Advertencia: La pestaña ya estaba cerrada o el contexto ya no es válido. Detalles: {e}")
            else:
                error_msg = (
                    f"\n ❌ FALLO (Playwright Error) - {nombre_paso}: Ocurrió un error de Playwright al intentar cerrar la pestaña.\n"
                    f"Detalles: {e}"
                )
                self.logger.error(error_msg)
                self.tomar_captura(f"{nombre_base}_error_cerrar_pestana_playwright", directorio)
                raise
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error al intentar cerrar la pestaña.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_cerrar_pestana", directorio)
            raise
        
    #51- Función para realizar una operación de "Drag and Drop" de un elemento a otro.
    def realizar_drag_and_drop(self, elemento_origen, elemento_destino, nombre_base, directorio, nombre_paso: str = "", tiempo_espera_manual: float = 1.0, timeout_ms: int = 1000):
        """
        Función para realizar una operación de "Drag and Drop" de un elemento a otro.
        Intenta primero con el método estándar de Playwright. Si falla o no está disponible,
        recurre a un método manual de acciones de ratón.

        Args:
            elemento_origen (Locator): El localizador del elemento a arrastrar.
            elemento_destino (Locator): El localizador del área donde soltar el elemento.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde guardar las capturas de pantalla.
            nombre_paso (str, opcional): Descripción del paso para logs y capturas. Por defecto "".
            tiempo_espera_manual (float, opcional): Tiempo en segundos para las pausas en el modo manual. Por defecto 1.0.
            timeout_ms (int, opcional): Tiempo máximo en milisegundos para esperar la operación de Drag and Drop. Por defecto 15000ms (15 segundos).
        """
        self.logger.info(f"\n {nombre_paso}: Intentando realizar 'Drag and Drop' de '{elemento_origen}' a '{elemento_destino}'")
        
        # 1. Verificar que ambos elementos estén visibles y habilitados antes de interactuar (pre-verificación)
        try:
            # --- Mejoras en la pre-validación con expect.to_be_enabled() ---
            self.logger.info(f"\n 🔍 Validando que el elemento de origen '{elemento_origen}' esté habilitado y listo para interactuar.")
            expect(elemento_origen).to_be_enabled(timeout=timeout_ms)
            self.logger.info(f"\n ✅ El elemento de origen '{elemento_origen}' está habilitado.")

            self.logger.info(f"\n 🔍 Validando que el elemento de destino '{elemento_destino}' esté habilitado y listo para interactuar.")
            expect(elemento_destino).to_be_enabled(timeout=timeout_ms)
            self.logger.info(f"\n ✅ El elemento de destino '{elemento_destino}' está habilitado.")

        except Error as e: # Captura TimeoutError y otros errores de Playwright en la validación
            error_msg = (
                f"\n ❌ FALLO (Pre-validación Playwright) - {nombre_paso}: Ocurrió un error de Playwright durante la validación inicial de los elementos.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_pre_validacion_playwright_drag_and_drop", directorio)
            raise # Re-lanza la excepción si falla la validación inicial
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Pre-validación Inesperado) - {nombre_paso}: Ocurrió un error inesperado durante la validación inicial de los elementos.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_pre_validacion_inesperado_drag_and_drop", directorio)
            raise

        # --- Intento 1: Usar el método .drag_to() del Locator ---
        try:
            self.logger.info(f"\n 🔄 Intentando 'Drag and Drop' con el método estándar de Playwright (locator.drag_to())...")
            elemento_origen.drag_to(elemento_destino, timeout=timeout_ms) 

            self.logger.info(f"\n ✅ 'Drag and Drop' realizado exitosamente con el método estándar.")
            self.tomar_captura(f"{nombre_base}_drag_and_drop_exitoso", directorio)
            return # Si funciona, salimos de la función

        except Error as e:
            # Captura errores de Playwright, incluyendo TimeoutError
            self.logger.warning(f"\n ⚠️ Advertencia: El método directo 'locator.drag_to()' falló con error de Playwright: {type(e).__name__}: {e}")
            self.logger.info("\n 🔄 Intentando 'Drag and Drop' con método manual de Playwright (mouse.hover, mouse.down, mouse.up)...")
            self.tomar_captura(f"{nombre_base}_fallo_directo_intentando_manual", directorio)
            # Pasamos el tiempo_espera_manual y el timeout_ms al método manual
            self._realizar_drag_and_drop_manual(elemento_origen, elemento_destino, nombre_base, directorio, nombre_paso, tiempo=tiempo_espera_manual, timeout_ms=timeout_ms)
            self.logger.info(f"\n ✅ 'Drag and Drop' realizado exitosamente con el método manual.")
        
        except Exception as e: # Cualquier otro error inesperado en el primer intento
            error_msg = (
                f"\n ❌ FALLO (Inesperado - Intento 1) - {nombre_paso}: Ocurrió un error inesperado al intentar realizar 'Drag and Drop' con el método estándar.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado_intento1_drag_and_drop", directorio)
            raise # Si falla aquí, y no es un error de Playwright, es un error más serio, lo re-lanzamos.
        
    #52- Función para mover sliders de rango (con dos pulgares)
    def mover_slider_rango(self, pulgar_izquierdo_locator, pulgar_derecho_locator, barra_slider_locator,
                           porcentaje_destino_izquierdo: float, porcentaje_destino_derecho: float,
                           nombre_base, directorio, nombre_paso= ""):
        self.logger.info(f"\n {nombre_paso}: Intentando mover el slider de rango. Pulgar Izquierdo a {porcentaje_destino_izquierdo*100:.0f}%, Pulgar Derecho a {porcentaje_destino_derecho*100:.0f}%")

        # Margen de error para la comparación de posiciones, en píxeles.
        # Un valor pequeño como 2 o 3 píxeles es razonable.
        TOLERANCIA_PIXELES = 3

        # 1. Validaciones iniciales de porcentajes
        if not (0.0 <= porcentaje_destino_izquierdo <= 1.0) or not (0.0 <= porcentaje_destino_derecho <= 1.0):
            raise ValueError("\n ❌ Los porcentajes de destino para ambos pulgares deben ser valores flotantes entre 0.0 (0%) y 1.0 (100%).")
        
        # Validación de negocio: el porcentaje izquierdo no puede ser mayor que el derecho
        if porcentaje_destino_izquierdo > porcentaje_destino_derecho:
            raise ValueError("\n ❌ El porcentaje del pulgar izquierdo no puede ser mayor que el del pulgar derecho.")
        
        localizadores = {
            "pulgar izquierdo": pulgar_izquierdo_locator,
            "pulgar derecho": pulgar_derecho_locator,
            "barra del slider": barra_slider_locator
        }

        try:
            barra_slider_locator.highlight() # Esto es para visualización en el navegador durante la ejecución
            
            # 2. Verificar visibilidad y habilitación de todos los elementos
            for nombre_elemento, localizador_elemento in localizadores.items():
                if not localizador_elemento.is_visible():
                    raise ValueError(f"\n ❌ El elemento '{nombre_elemento}' ('{localizador_elemento.selector}') no está visible.")
                if not localizador_elemento.is_enabled():
                    raise ValueError(f"\n ❌El elemento '{nombre_elemento}' ('{localizador_elemento.selector}') no está habilitado.")
            
            # Obtener el bounding box de la barra del slider (esencial para el cálculo)
            caja_barra = barra_slider_locator.bounding_box()
            if not caja_barra:
                raise RuntimeError(f"\n ❌ No se pudo obtener el bounding box de la barra del slider '{barra_slider_locator.selector}'.")

            inicio_x_barra = caja_barra['x']
            ancho_barra = caja_barra['width']
            
            # --- Mover Pulgar Izquierdo (Mínimo) ---
            caja_pulgar_izquierdo = pulgar_izquierdo_locator.bounding_box()
            if not caja_pulgar_izquierdo:
                raise RuntimeError(f"\n ❌ No se pudo obtener el bounding box del pulgar izquierdo '{pulgar_izquierdo_locator.selector}'.")

            posicion_x_destino_izquierdo = inicio_x_barra + (ancho_barra * porcentaje_destino_izquierdo)
            posicion_y_destino = caja_pulgar_izquierdo['y'] + (caja_pulgar_izquierdo['height'] / 2) # Y central del pulgar

            # Calcular la posición X central actual del pulgar izquierdo
            posicion_x_actual_izquierdo = caja_pulgar_izquierdo['x'] + (caja_pulgar_izquierdo['width'] / 2)

            # Verificar si el pulgar izquierdo ya está en la posición deseada
            if abs(posicion_x_actual_izquierdo - posicion_x_destino_izquierdo) < TOLERANCIA_PIXELES:
                self.logger.info(f"\n  > Pulgar izquierdo ya se encuentra en la posición deseada ({porcentaje_destino_izquierdo*100:.0f}%). No se requiere movimiento.")
            else:
                self.logger.info(f"\n  > Moviendo pulgar izquierdo de X={posicion_x_actual_izquierdo:.0f} a X={posicion_x_destino_izquierdo:.0f} ({porcentaje_destino_izquierdo*100:.0f}%)...")
                self.page.mouse.move(posicion_x_actual_izquierdo, posicion_y_destino) # Mover al centro del pulgar actual
                self.page.mouse.down() # Presionar
                time.sleep(0.2)
                self.page.mouse.move(posicion_x_destino_izquierdo, posicion_y_destino, steps=10) # Arrastrar
                time.sleep(0.2)
                self.page.mouse.up() # Soltar
                self.logger.info(f"\n  > Pulgar izquierdo movido a X={posicion_x_destino_izquierdo:.0f}.")
            time.sleep(0.5) # Pausa adicional después de procesar el primer pulgar

            # --- Mover Pulgar Derecho (Máximo) ---
            caja_pulgar_derecho = pulgar_derecho_locator.bounding_box()
            if not caja_pulgar_derecho:
                raise RuntimeError(f"\n ❌ No se pudo obtener el bounding box del pulgar derecho '{pulgar_derecho_locator.selector}'.")

            posicion_x_destino_derecho = inicio_x_barra + (ancho_barra * porcentaje_destino_derecho)
            posicion_y_destino_derecho = caja_pulgar_derecho['y'] + (caja_pulgar_derecho['height'] / 2) # Y central del pulgar

            # Calcular la posición X central actual del pulgar derecho
            posicion_x_actual_derecho = caja_pulgar_derecho['x'] + (caja_pulgar_derecho['width'] / 2)

            # Verificar si el pulgar derecho ya está en la posición deseada
            if abs(posicion_x_actual_derecho - posicion_x_destino_derecho) < TOLERANCIA_PIXELES:
                self.logger.info(f"\n  > Pulgar derecho ya se encuentra en la posición deseada ({porcentaje_destino_derecho*100:.0f}%). No se requiere movimiento.")
            else:
                self.logger.info(f"\n  > Moviendo pulgar derecho de X={posicion_x_actual_derecho:.0f} a X={posicion_x_destino_derecho:.0f} ({porcentaje_destino_derecho*100:.0f}%)...")
                # Siempre movemos el ratón a la posición actual del pulgar antes de "down" para asegurar el arrastre correcto
                self.page.mouse.move(posicion_x_actual_derecho, posicion_y_destino_derecho)
                self.page.mouse.down() # Presionar
                time.sleep(0.2)
                self.page.mouse.move(posicion_x_destino_derecho, posicion_y_destino_derecho, steps=10) # Arrastrar
                time.sleep(0.2)
                self.page.mouse.up() # Soltar
                self.logger.info(f"\n  > Pulgar derecho movido a X={posicion_x_destino_derecho:.0f}.")

            self.logger.info(f"\n ✅ Slider de rango procesado exitosamente. Izquierdo a {porcentaje_destino_izquierdo*100:.0f}%, Derecho a {porcentaje_destino_derecho*100:.0f}%.")
            self.tomar_captura(f"{nombre_base}_slider_rango_procesado_{porcentaje_destino_izquierdo*100:.0f}_{porcentaje_destino_derecho*100:.0f}pc", directorio)

        except ValueError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Validación) - {nombre_paso}: {e}\n"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_error_validacion_slider_rango", directorio)
            raise

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright) - {nombre_paso}: Ocurrió un error de Playwright al intentar mover el slider de rango.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_error_playwright_slider_rango", directorio)
            raise

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Inesperado) - {nombre_paso}: Ocurrió un error inesperado al intentar mover el slider de rango.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_error_inesperado_slider_rango", directorio)
            raise
    
    #53- Función para seleccionar una opción en un ComboBox (elemento <select>) por su atributo 'value'.
    def seleccionar_opcion_por_valor(self, combobox_locator, valor_a_seleccionar, nombre_base, directorio):
        self.logger.info(f"\n Iniciando selección de '{valor_a_seleccionar}' en ComboBox por valor: '{combobox_locator}'")

        try:
            # 1. Asegurarse de que el ComboBox esté visible y habilitado
            self.logger.info(f"\n Esperando que el ComboBox '{combobox_locator}' sea visible y habilitado.")
            expect(combobox_locator).to_be_visible(timeout=5000) # Espera 5 segundos
            combobox_locator.highlight() # Para visualización durante la ejecución
            expect(combobox_locator).to_be_enabled(timeout=5000) # Espera 5 segundos
            self.logger.info(f"ComboBox '{combobox_locator}' es visible y habilitado.")
            
            # 2. Tomar captura antes de la selección
            self.tomar_captura(f"{nombre_base}_antes_de_seleccionar_combo", directorio)

            # 3. Seleccionar la opción por su valor
            combobox_locator.select_option(valor_a_seleccionar)
            self.logger.info(f"\n ✅ Opción '{valor_a_seleccionar}' seleccionada exitosamente en '{combobox_locator}'.")

            # 4. Verificar que la opción fue seleccionada correctamente
            expect(combobox_locator).to_have_value(valor_a_seleccionar, timeout=5000)
            self.logger.info(f"\n ✅ ComboBox '{combobox_locator}' verificado con valor '{valor_a_seleccionar}'.")

            # 5. Tomar captura después de la selección exitosa
            self.tomar_captura(f"{nombre_base}_despues_de_seleccionar_combo_exito", directorio)

        except TimeoutError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Timeout): El ComboBox '{combobox_locator}' "
                f"no se volvió visible/habilitado o la opción '{valor_a_seleccionar}' no se pudo seleccionar a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_combo", directorio)
            raise AssertionError(f"\n ❌ No se pudo seleccionar opción '{valor_a_seleccionar}' en ComboBox: {combobox_locator}") from e

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright): Ocurrió un error al intentar seleccionar la opción '{valor_a_seleccionar}' en '{combobox_locator}'.\n"
                f"Posibles causas: Selector inválido, elemento no es un <select>, opción no existe.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_combo", directorio)
            raise AssertionError(f"\n ❌ Error de Playwright al seleccionar ComboBox: {combobox_locator}") from e

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al manejar el ComboBox '{combobox_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_combo", directorio)
            raise AssertionError(f"\n ❌ Error inesperado con ComboBox: {combobox_locator}") from e
        
    #54- Función para seleccionar una opción en un ComboBox (elemento <select>) por su texto visible (label).
    def seleccionar_opcion_por_label(self, combobox_locator, label_a_seleccionar, nombre_base, directorio, value_esperado: str = None, tiempo=1):
        # Si nombre_paso no se proporciona, usa el selector del locator para mayor claridad en el log
        self.logger.info(f"\n Iniciando selección de '{label_a_seleccionar}' en ComboBox por label: '{combobox_locator}'")

        try:
            # 1. Asegurarse de que el ComboBox esté visible y habilitado
            self.logger.info(f"\n Esperando que el ComboBox '{combobox_locator}' sea visible y habilitado.")
            expect(combobox_locator).to_be_visible(timeout=5000) # Espera 5 segundos
            combobox_locator.highlight() # Para visualización durante la ejecución
            expect(combobox_locator).to_be_enabled(timeout=5000) # Espera 5 segundos
            self.logger.info(f"\n ComboBox '{combobox_locator}' es visible y habilitado.")
            
            # 2. Tomar captura antes de la selección
            self.tomar_captura(f"{nombre_base}_antes_de_seleccionar_combo_label", directorio)

            # 3. Seleccionar la opción por su texto visible (label)
            # El método select_option() espera automáticamente a que el elemento
            # sea visible, habilitado y con la opción disponible.
            combobox_locator.select_option(label_a_seleccionar)
            self.logger.info(f"✅ Opción '{label_a_seleccionar}' seleccionada exitosamente en '{combobox_locator}' por label.")

            # 4. Verificar que la opción fue seleccionada correctamente
            # Usamos to_have_text() para asegurar que el texto visible del select cambió al esperado.
            # Ojo: to_have_text() verifica el texto del elemento <select> en sí, que a menudo
            # es el texto de la opción seleccionada. Si el ComboBox es complejo (no un <select> nativo),
            # podrías necesitar verificar el texto de un elemento adyacente que muestra la selección.
            valor_para_comparar = value_esperado if value_esperado is not None else label_a_seleccionar

            expect(combobox_locator).to_have_value(valor_para_comparar, timeout=5000)
            self.logger.info(f"✅ ComboBox '{combobox_locator}' verificado con valor seleccionado '{valor_para_comparar}'.")

            # 5. Tomar captura después de la selección exitosa
            self.tomar_captura(f"{nombre_base}_despues_de_seleccionar_combo_label_exito", directorio)

        except TimeoutError as e:
            mensaje_error = (
                f"❌ FALLO (Timeout): El ComboBox '{combobox_locator}' "
                f"no se volvió visible/habilitado o la opción '{label_a_seleccionar}' no se pudo seleccionar a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_combo_label", directorio)
            raise AssertionError(f"❌ No se pudo seleccionar opción '{label_a_seleccionar}' en ComboBox por label: {combobox_locator}") from e

        except Error as e:
            mensaje_error = (
                f"❌ FALLO (Error de Playwright): Ocurrió un error al intentar seleccionar la opción '{label_a_seleccionar}' en '{combobox_locator}'.\n"
                f"Posibles causas: Selector inválido, elemento no es un <select>, opción con ese label no existe.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_combo_label", directorio)
            raise AssertionError(f"❌ Error de Playwright al seleccionar ComboBox por label: {combobox_locator}") from e

        except Exception as e:
            mensaje_error = (
                f"❌ FALLO (Error Inesperado): Ocurrió un error desconocido al manejar el ComboBox '{combobox_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_combo_label", directorio)
            raise AssertionError(f"❌ Error inesperado con ComboBox por label: {combobox_locator}") from e
    
    #55- Función para presionar la tecla TAB en el teclado
    def Tab_Pess(self, tiempo= 0.5):
        self.logger.info(f"\n Presionando la tecla TAB y esperando {tiempo} segundos.")
        self.page.keyboard.press("Tab")
        time.sleep(tiempo)
        self.logger.info("\n Tecla TAB presionada exitosamente.")
        
    #56- Función optimizada para seleccionar múltiples opciones en un ComboBox múltiple.
    def seleccionar_multiples_opciones_combo(self, combobox_multiple_locator, valores_a_seleccionar: list[str], nombre_base, directorio):
        self.logger.info(f"\n Iniciando selección de múltiples opciones {valores_a_seleccionar} en ComboBox: '{combobox_multiple_locator}'")

        try:
            # 1. Asegurarse de que el ComboBox esté visible y habilitado
            self.logger.info(f"\n Esperando que el ComboBox múltiple '{combobox_multiple_locator}' sea visible y habilitado.")
            expect(combobox_multiple_locator).to_be_visible(timeout=5000) # Espera 5 segundos
            combobox_multiple_locator.highlight() # Para visualización durante la ejecución
            expect(combobox_multiple_locator).to_be_enabled(timeout=5000) # Espera 5 segundos
            self.logger.info(f"\n ComboBox múltiple '{combobox_multiple_locator}' es visible y habilitado.")
            
            # Verificar que sea un select múltiple, si es crítico
            # Opcional: Esto puede no ser necesario si tu locator ya apunta específicamente a un 'select[multiple]'
            # expect(combobox_multiple_locator).to_have_attribute("multiple", "") 
            # print("  > ComboBox verificado como select múltiple.")

            # 2. Tomar captura antes de la selección
            self.tomar_captura(f"{nombre_base}_antes_de_seleccionar_multi_combo", directorio)

            # 3. Seleccionar las opciones
            # Playwright's select_option() para listas maneja tanto valores como labels.
            # Pasando una lista de strings seleccionará las opciones correspondientes.
            combobox_multiple_locator.select_option(valores_a_seleccionar)
            self.logger.info(f"\n ✅ Opciones '{valores_a_seleccionar}' seleccionadas exitosamente en '{combobox_multiple_locator}'.")

            # 4. Verificar que las opciones fueron seleccionadas correctamente
            # input_value() para un select múltiple retorna una lista de los valores seleccionados.
            expect(combobox_multiple_locator).to_have_values(valores_a_seleccionar, timeout=5000)
            self.logger.info(f"\n ✅ ComboBox múltiple '{combobox_multiple_locator}' verificado con valores seleccionados: {valores_a_seleccionar}.")

            # 5. Tomar captura después de la selección exitosa
            self.tomar_captura(f"{nombre_base}_despues_de_seleccionar_multi_combo_exito", directorio)

        except TimeoutError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Timeout): El ComboBox múltiple '{combobox_multiple_locator}' "
                f"no se volvió visible/habilitado o las opciones '{valores_a_seleccionar}' no se pudieron seleccionar a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_multi_combo", directorio)
            raise AssertionError(f"\n ❌ No se pudieron seleccionar opciones '{valores_a_seleccionar}' en ComboBox múltiple: {combobox_multiple_locator}") from e

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright): Ocurrió un error al intentar seleccionar las opciones '{valores_a_seleccionar}' en '{combobox_multiple_locator}'.\n"
                f"Posibles causas: Selector inválido, elemento no es un <select multiple>, alguna opción no existe.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_multi_combo", directorio)
            raise AssertionError(f"\n ❌ Error de Playwright al seleccionar ComboBox múltiple: {combobox_multiple_locator}") from e

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al manejar el ComboBox múltiple '{combobox_multiple_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_multi_combo", directorio)
            raise AssertionError(f"\n ❌ Error inesperado con ComboBox múltiple: {combobox_multiple_locator}") from e
        
    #57 Función que obtiene y imprime los valores y el texto de todas las opciones en un dropdown list.
    def obtener_valores_dropdown(self, selector_dropdown, nombre_base, directorio, timeout_ms: int = 5000) -> List[Dict[str, str]] | None:
        self.logger.info(f"\n--- Extrayendo valores del dropdown '{selector_dropdown}' ---")
        valores_opciones: List[Dict[str, str]] = []

        try:
            # 1. Asegurar que el dropdown es visible y habilitado
            selector_dropdown.highlight()
            self.tomar_captura(f"{nombre_base}_dropdown_antes_extraccion", directorio)

            self.logger.info(f"\n ⏳ Esperando que el dropdown '{selector_dropdown}' sea visible y habilitado...")
            expect(selector_dropdown).to_be_visible(timeout=timeout_ms)
            expect(selector_dropdown).to_be_enabled(timeout=timeout_ms)
            self.logger.info(f"\n ✅ Dropdown '{selector_dropdown}' es visible y habilitado.")

            # 2. Obtener todas las opciones dentro del dropdown
            # Usamos `all()` para obtener una lista de locators para cada opción
            option_locators = selector_dropdown.locator("option").all()

            if not option_locators:
                self.logger.warning(f"\n ⚠️ No se encontraron opciones dentro del dropdown '{selector_dropdown}'.")
                self.tomar_captura(f"{nombre_base}_dropdown_sin_opciones", directorio)
                return None

            self.logger.info(f"\n Encontradas {len(option_locators)} opciones para '{selector_dropdown}':")

            # 3. Iterar sobre cada opción y extraer su 'value' y 'text_content'
            for i, option_locator in enumerate(option_locators):
                value = option_locator.get_attribute("value")
                text = option_locator.text_content()

                # Limpieza de espacios en blanco
                clean_value = value.strip() if value else ""
                clean_text = text.strip() if text else ""

                valores_opciones.append({'value': clean_value, 'text': clean_text})
                self.logger.info(f"\n  Opción {i+1}: Value='{clean_value}', Text='{clean_text}'")

            self.logger.info(f"\n ✅ Valores obtenidos exitosamente del dropdown '{selector_dropdown}'.")
            self.tomar_captura(f"{nombre_base}_dropdown_valores_extraidos", directorio)
            return valores_opciones

        except TimeoutError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Timeout): El dropdown '{selector_dropdown}' "
                f"no se volvió visible/habilitado o sus opciones no cargaron a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_timeout", directorio)
            raise AssertionError(f"\n Dropdown no disponible: {selector_dropdown}") from e

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al intentar obtener los valores del dropdown '{selector_dropdown}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_playwright_error", directorio)
            raise AssertionError(f"\n Error de Playwright al extraer valores del dropdown: {selector_dropdown}") from e

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar obtener los valores del dropdown '{selector_dropdown}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_inesperado", directorio)
            raise AssertionError(f"\n Error inesperado al extraer valores del dropdown: {selector_dropdown}") from e
        
    #58- Función que obtiene y imprime los valores y el texto de todas las opciones en un dropdown list.
    #Opcionalmente, compara las opciones obtenidas con una lista de opciones esperadas.
    def obtener_y_comparar_valores_dropdown(self, dropdown_locator, nombre_base, directorio, expected_options: List[Union[str, Dict[str, str]]] = None, compare_by_text: bool = True, compare_by_value: bool = False, timeout_ms: int = 5000) -> List[Dict[str, str]] | None:
        """
        Args:
            dropdown_locator (Locator): El objeto Locator de Playwright para el elemento <select> del dropdown.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            expected_options (List[Union[str, Dict[str, str]]], optional):
                Lista de opciones esperadas para la comparación. Puede ser:
                - List[str]: Si solo quieres comparar por el texto visible de las opciones.
                - List[Dict[str, str]]: Si quieres comparar por 'value' y 'text'.
                  Ej: [{'value': 'usa', 'text': 'Estados Unidos'}].
                Por defecto es None (no se realiza comparación).
            compare_by_text (bool): Si es True, compara el texto visible de las opciones.
                                    Usado si expected_options es List[str] o List[Dict].
            compare_by_value (bool): Si es True, compara el atributo 'value' de las opciones.
                                     Usado si expected_options es List[Dict].
            timeout_ms (int): Tiempo máximo de espera en milisegundos.

        Returns:
            List[Dict[str, str]] | None: Una lista de diccionarios con las opciones reales.
            Retorna None si ocurre un error o no se encuentran opciones.
            La función generará una AssertionError si la comparación falla.
        """
        self.logger.info(f"\n --- Extrayendo y comparando valores del dropdown '{dropdown_locator}' ---")
        valores_opciones_reales: List[Dict[str, str]] = []

        try:
            # 1. Asegurar que el dropdown es visible y habilitado
            dropdown_locator.highlight()
            self.tomar_captura(f"{nombre_base}_dropdown_antes_extraccion_y_comparacion", directorio)

            self.logger.info(f"\n ⏳ Esperando que el dropdown '{dropdown_locator}' sea visible y habilitado...")
            expect(dropdown_locator).to_be_visible(timeout=timeout_ms)
            expect(dropdown_locator).to_be_enabled(timeout=timeout_ms)
            self.logger.info(f"\n ✅ Dropdown '{dropdown_locator}' es visible y habilitado.")

            # 2. Obtener todas las opciones dentro del dropdown
            option_locators = dropdown_locator.locator("option").all()

            if not option_locators:
                self.logger.warning(f"\n ⚠️ No se encontraron opciones dentro del dropdown '{dropdown_locator}'.")
                self.tomar_captura(f"{nombre_base}_dropdown_sin_opciones", directorio)
                # Si se esperaban opciones y no hay ninguna, esto es un fallo
                if expected_options:
                    raise AssertionError(f"\n ❌ FALLO: No se encontraron opciones en el dropdown '{dropdown_locator}', pero se esperaban {len(expected_options)}.")
                return None

            self.logger.info(f"\n Encontradas {len(option_locators)} opciones reales para '{dropdown_locator}':")

            # 3. Iterar sobre cada opción y extraer su 'value' y 'text_content'
            for i, option_locator in enumerate(option_locators):
                value = option_locator.get_attribute("value")
                text = option_locator.text_content()

                clean_value = value.strip() if value else ""
                clean_text = text.strip() if text else ""

                valores_opciones_reales.append({'value': clean_value, 'text': clean_text})
                self.logger.info(f"\n   Opción Real {i+1}: Value='{clean_value}', Text='{clean_text}'")

            self.logger.info(f"\n ✅ Valores obtenidos exitosamente del dropdown '{dropdown_locator}'.")
            self.tomar_captura(f"{nombre_base}_dropdown_valores_extraidos", directorio)

            # 4. Comparar con las opciones esperadas (si se proporcionan)
            if expected_options is not None:
                self.logger.info("\n--- Realizando comparación de opciones ---")
                try:
                    expected_set = set()
                    real_set = set()

                    # Preparar los conjuntos para la comparación
                    for opt in expected_options:
                        if isinstance(opt, str):
                            if compare_by_text:
                                expected_set.add(opt.strip().lower()) # Normalizar para comparación
                        elif isinstance(opt, dict):
                            if compare_by_text and 'text' in opt:
                                expected_set.add(opt['text'].strip().lower())
                            if compare_by_value and 'value' in opt:
                                expected_set.add(opt['value'].strip().lower())
                        else:
                            self.logger.warning(f"\n ⚠️ Advertencia: Formato de opción esperada no reconocido: {opt}. Ignorando.")

                    for opt_real in valores_opciones_reales:
                        if compare_by_text:
                            real_set.add(opt_real['text'].strip().lower())
                        if compare_by_value:
                            real_set.add(opt_real['value'].strip().lower())

                    # Comprobar si los conjuntos son idénticos
                    if expected_set == real_set:
                        self.logger.info("\n ✅ ÉXITO: Las opciones del dropdown coinciden con las opciones esperadas.")
                        self.tomar_captura(f"{nombre_base}_dropdown_comparacion_exitosa", directorio)
                    else:
                        missing_in_real = list(expected_set - real_set)
                        missing_in_expected = list(real_set - expected_set)
                        error_msg = f"\n ❌ FALLO: Las opciones del dropdown NO coinciden con las esperadas.\n"
                        if missing_in_real:
                            error_msg += f"\n  - Opciones esperadas no encontradas en el dropdown: {missing_in_real}\n"
                        if missing_in_expected:
                            error_msg += f"\n  - Opciones encontradas en el dropdown que no estaban esperadas: {missing_in_expected}\n"
                        self.logger.error(error_msg)
                        self.tomar_captura(f"{nombre_base}_dropdown_comparacion_fallida", directorio)
                        raise AssertionError(f"\n Comparación de opciones del dropdown fallida para '{dropdown_locator}'. {error_msg.strip()}")

                except Exception as e:
                    self.logger.error(f"\n ❌ FALLO: Ocurrió un error durante la comparación de opciones: {e}")
                    self.tomar_captura(f"{nombre_base}_dropdown_error_comparacion", directorio)
                    raise AssertionError(f"\n Error al comparar opciones del dropdown '{dropdown_locator}': {e}") from e

            return valores_opciones_reales

        except TimeoutError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Timeout): El dropdown '{dropdown_locator}' " # Usar dropdown_locator directamente
                f"no se volvió visible/habilitado o sus opciones no cargaron a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_timeout", directorio)
            raise AssertionError(f"\n Dropdown no disponible: {dropdown_locator}") from e

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al intentar obtener los valores del dropdown '{dropdown_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_playwright_error", directorio)
            raise AssertionError(f"\n Error de Playwright al extraer valores del dropdown: {dropdown_locator}") from e

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar obtener los valores del dropdown '{dropdown_locator}'.\n"
                f"Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_dropdown_fallo_inesperado", directorio)
            raise AssertionError(f"\n Error inesperado al extraer valores del dropdown: {dropdown_locator}") from e
    
    #59- Función que detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo Excel.
    def num_Filas_excel(self, archivo_excel_path: str, hoja: str, has_header: bool = False) -> int:
        """
        Detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo Excel,
        opcionalmente descontando una fila de encabezado.

        Args:
            archivo_excel_path (str): La ruta completa al archivo Excel.
            hoja (str): El nombre de la hoja/pestaña dentro del archivo Excel.
            has_header (bool): Si es True, se descuenta una fila para el encabezado. Por defecto es False.

        Returns:
            int: El número de filas de datos. Retorna 0 si hay un error o la hoja está vacía.
        """
        try:
            self.logger.info(f"\n 🔎 Intentando obtener el número de filas para la hoja '{hoja}' en el archivo '{archivo_excel_path}' (tiene encabezado: {has_header}).")
            workbook = openpyxl.load_workbook(archivo_excel_path) # Carga el libro de trabajo Excel
            sheet = workbook[hoja] # Selecciona la hoja específica del libro
            num_physical_rows = sheet.max_row # Obtiene el número total de filas con contenido

            if has_header and num_physical_rows > 0: # Si tiene encabezado y hay al menos una fila
                num_data_rows = num_physical_rows - 1 # Resta 1 para no contar el encabezado
                self.logger.info(f"\n ✅ Se encontraron {num_data_rows} filas de datos (descontando encabezado) en la hoja '{hoja}'.")
                return num_data_rows # Ahora retorna el número de filas de DATOS
            else:
                self.logger.info(f"\n ✅ Se encontraron {num_physical_rows} filas ocupadas en la hoja '{hoja}'.")
                return num_physical_rows # Para hojas sin encabezado, es el número de filas físicas
        except FileNotFoundError:
            self.logger.error(f"\n ❌ Error: El archivo Excel no se encontró en la ruta: {archivo_excel_path}")
            return 0
        except KeyError:
            self.logger.error(f"\n ❌ Error: La hoja '{hoja}' no se encontró en el archivo Excel: {archivo_excel_path}")
            return 0
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el número de filas del Excel: {e}")
            return 0

    #60- Función que obtiene el valor de una celda específica de una hoja Excel, ajustando la fila si se indica que hay un encabezado.
    def dato_Columna_excel(self, archivo_excel_path: str, hoja: str, numero_fila_logica: int, nombre_o_indice_columna: Union[str, int], has_header_excel: bool = False) -> Union[str, int, float, None]:
        """
        Obtiene el valor de una celda específica de una hoja Excel, ajustando la fila
        si se indica que hay un encabezado.

        Args:
            archivo_excel_path (str): La ruta completa al archivo Excel.
            hoja (str): El nombre de la hoja/pestaña dentro del archivo Excel.
            fila (int): El número de fila de la celda (basado en 1). Si has_header es True,
                        esta es la fila lógica de datos (e.g., 1 para la primera fila después del encabezado).
            colum (int): El número de columna de la celda (basado en 1).
            has_header (bool): Si es True, la fila física en Excel será 'fila + 1' para compensar el encabezado.
                               Por defecto es False.

        Returns:
            Any: El valor de la celda. Retorna None si hay un error.
        """
        try:
            workbook = openpyxl.load_workbook(archivo_excel_path)
            sheet = workbook[hoja]

            # 1. Determinar el índice físico de la columna
            col_index = -1
            if isinstance(nombre_o_indice_columna, str):
                # Si se pasó un nombre de columna, buscar su índice en la primera fila (encabezado)
                header_found = False
                for col_idx, cell in enumerate(sheet[1], 1): # sheet[1] se refiere a la primera fila física del Excel
                    if cell.value is not None and str(cell.value).strip().lower() == nombre_o_indice_columna.strip().lower():
                        col_index = col_idx
                        header_found = True
                        break
                if not header_found:
                    self.logger.error(f"\n ❌ Error: La columna '{nombre_o_indice_columna}' no fue encontrada en el encabezado de la hoja '{hoja}'.")
                    return None
            elif isinstance(nombre_o_indice_columna, int):
                # Si se pasó un índice numérico de columna
                col_index = nombre_o_indice_columna
            else:
                self.logger.error(f"\n ❌ Error: El parámetro 'nombre_o_indice_columna' debe ser un string (nombre) o un entero (índice). Se recibió: '{nombre_o_indice_columna}' ({type(nombre_o_indice_columna).__name__}).")
                return None

            # Validar que el índice de columna sea válido
            if not (1 <= col_index <= sheet.max_column):
                self.logger.error(f"\n ❌ Error: Índice de columna '{col_index}' fuera de rango para la hoja '{hoja}' (máximo: {sheet.max_column}).")
                return None

            # 2. Determinar el índice físico de la fila
            # 'numero_fila_logica' es la fila de datos que el usuario piensa (1 para la primera fila de datos).
            # Si hay encabezado, la primera fila de datos (lógica 1) está en la fila física 2.
            # Por lo tanto, sumamos 1 si hay encabezado.
            actual_fila_fisica = numero_fila_logica + 1 if has_header_excel else numero_fila_logica

            # Validar que la fila física sea válida
            if not (1 <= actual_fila_fisica <= sheet.max_row):
                self.logger.warning(f"\n ⚠️ Advertencia: La fila física {actual_fila_fisica} (lógica: {numero_fila_logica}) está fuera del rango de filas de la hoja '{hoja}' (máximo: {sheet.max_row}). Retornando None.")
                return None
            
            self.logger.info(f"\n 🔎 Intentando obtener el dato de la celda (Fila lógica: {numero_fila_logica}, Fila física: {actual_fila_fisica}, Columna: {nombre_o_indice_columna}) de la hoja '{hoja}' en el archivo '{archivo_excel_path}' (tiene encabezado: {has_header_excel}).")
            
            # Obtiene el valor de la celda especificada
            cell_value = sheet.cell(row=actual_fila_fisica, column=col_index).value
            
            # Convertir a string para asegurar que 'rellenar_campo_de_texto' siempre reciba un str
            if cell_value is not None:
                valor_retorno = str(cell_value)
                self.logger.info(f"\n ✅ Dato obtenido de (Fila: {numero_fila_logica}, Columna: {nombre_o_indice_columna}) en '{hoja}': '{valor_retorno}'.")
                return valor_retorno
            else:
                self.logger.warning(f"\n ⚠️ La celda en Fila lógica: {numero_fila_logica}, Columna: {nombre_o_indice_columna} en '{hoja}' está vacía. Retornando None.")
                return None

        except FileNotFoundError:
            self.logger.error(f"\n ❌ Error: El archivo Excel no se encontró en la ruta: {archivo_excel_path}")
            return None
        except KeyError:
            self.logger.error(f"\n ❌ Error: La hoja '{hoja}' no se encontró en el archivo Excel: {archivo_excel_path}")
            return None
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el dato del Excel: {e}")
            return None
    
    #61- Función que detecta y devuelve el número total de filas ocupadas en una hoja específica de un archivo CSV.
    def num_Filas_csv(self, archivo_csv_path: str, delimiter: str = ',', has_header: bool = False) -> int:
        """
        Detecta y devuelve el número total de filas de datos en un archivo CSV,
        opcionalmente descontando una fila de encabezado.

        Args:
            archivo_csv_path (str): La ruta completa al archivo CSV.
            delimiter (str): El caracter utilizado como separador de datos en el CSV (e.g., ',', ';', '\t'). Por defecto es ','.
            has_header (bool): Si es True, se descuenta una fila para el encabezado. Por defecto es False.

        Returns:
            int: El número de filas de datos. Retorna 0 si hay un error o el archivo está vacío.
        """
        try:
            self.logger.info(f"\n 🔎 Intentando obtener el número de filas para el archivo CSV '{archivo_csv_path}' con delimitador '{delimiter}' (tiene encabezado: {has_header}).")
            with open(archivo_csv_path, 'r', newline='', encoding='utf-8') as csvfile: # Abre el archivo CSV en modo lectura, sin nuevas líneas automáticas y con encoding UTF-8
                csv_reader = csv.reader(csvfile, delimiter=delimiter) # Crea un objeto reader para iterar sobre las líneas del CSV, usando el delimitador especificado
                row_count = sum(1 for row in csv_reader) # Cuenta todas las filas en el CSV

            if has_header and row_count > 0: # Si tiene encabezado y el archivo no está vacío
                num_data_rows = row_count - 1 # Resta 1 para no contar el encabezado, obteniendo solo las filas de datos
                self.logger.info(f"\n ✅ Se encontraron {num_data_rows} filas de datos (descontando encabezado) en el archivo CSV '{archivo_csv_path}'.")
                return num_data_rows # Retorna el número de filas de datos
            else: # Si no tiene encabezado o el archivo está vacío
                self.logger.info(f"\n ✅ Se encontraron {row_count} filas ocupadas en el archivo CSV '{archivo_csv_path}'.")
                return row_count # Retorna el número total de filas (que son todas de datos en este caso)
        except FileNotFoundError:
            self.logger.error(f"\n ❌ Error: El archivo CSV no se encontró en la ruta: {archivo_csv_path}")
            return 0
        except csv.Error as e:
            self.logger.error(f"\n ❌ Error de CSV al leer '{archivo_csv_path}': {e}") 
            return 0
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el número de filas del CSV: {e}")
            return 0

    #62- Función que obtiene el valor de una "celda" específica de un archivo CSV, ajustando la fila si se indica que hay un encabezado y recibiendo el delimitador.
    def dato_Columna_csv(self, archivo_csv_path: str, fila: int, colum: int, delimiter: str = ',', has_header: bool = False) -> Union[str, None]:
        """
        Obtiene el valor de una "celda" específica de un archivo CSV, ajustando la fila
        si se indica que hay un encabezado y recibiendo el delimitador.

        Args:
            archivo_csv_path (str): La ruta completa al archivo CSV.
            fila (int): El número de fila de la celda (basado en 1, lógica). Si has_header es True,
                        esta es la fila lógica de datos (e.g., 1 para la primera fila después del encabezado).
            colum (int): El número de columna de la celda (basado en 1, lógica).
            delimiter (str): El caracter utilizado como separador de datos en el CSV. Por defecto es ','.
            has_header (bool): Si es True, la fila física en el CSV será 'fila + 1' para compensar el encabezado.
                               Por defecto es False.

        Returns:
            str | None: El valor de la celda como string. Retorna None si hay un error.
        """
        try:
            # Convierte el número de fila lógica (1-basada) a un índice 0-basado para Python
            actual_fila_0_indexed = fila - 1
            # Si hay encabezado, se suma 1 al índice 0-basado para saltar la fila de encabezado
            if has_header:
                actual_fila_0_indexed += 1

            # Convierte el número de columna lógica (1-basada) a un índice 0-basado para Python
            actual_col_0_indexed = colum - 1

            self.logger.info(f"\n 🔎 Intentando obtener el dato de la celda (Fila lógica: {fila}, Fila física 0-idx: {actual_fila_0_indexed}, Columna lógica: {colum}, Columna física 0-idx: {actual_col_0_indexed}) del archivo CSV '{archivo_csv_path}' con delimitador '{delimiter}' (tiene encabezado: {has_header}).")

            with open(archivo_csv_path, 'r', newline='', encoding='utf-8') as csvfile: # Abre el archivo CSV
                csv_reader = csv.reader(csvfile, delimiter=delimiter) # Crea el objeto reader
                rows = list(csv_reader) # Lee todas las filas del CSV en una lista de listas (cada sublista es una fila)

            if actual_fila_0_indexed < 0 or actual_fila_0_indexed >= len(rows): # Comprueba si el índice de fila está fuera de los límites del archivo
                self.logger.error(f"\n ❌ Error: La fila {fila} (física 0-indexed: {actual_fila_0_indexed}) está fuera de los límites del archivo CSV '{archivo_csv_path}'. Total filas: {len(rows)}.")
                return None # Retorna None en caso de error

            # Comprueba si el índice de columna está fuera de los límites de la fila específica
            if actual_col_0_indexed < 0 or actual_col_0_indexed >= len(rows[actual_fila_0_indexed]):
                self.logger.error(f"\n ❌ Error: La columna {colum} (física 0-indexed: {actual_col_0_indexed}) está fuera de los límites de la fila {fila} del archivo CSV '{archivo_csv_path}'. Total columnas en esa fila: {len(rows[actual_fila_0_indexed])}.")
                return None # Retorna None en caso de error

            cell_value = rows[actual_fila_0_indexed][actual_col_0_indexed] # Obtiene el valor de la celda usando los índices ajustados
            self.logger.info(f"\n ✅ Dato obtenido de (Fila lógica: {fila}, Columna lógica: {colum}) en '{archivo_csv_path}': '{cell_value}'.")
            return cell_value # Retorna el valor de la celda
        
        except FileNotFoundError:
            self.logger.error(f"\n ❌ Error: El archivo CSV no se encontró en la ruta: {archivo_csv_path}")
            return None
        except ValueError:
            self.logger.error(f"\n ❌ Error: Los parámetros 'fila' y 'colum' deben ser números enteros. Se recibieron: fila='{fila}', colum='{colum}'.")
            return None
        except csv.Error as e:
            self.logger.error(f"\n ❌ Error de CSV al leer '{archivo_csv_path}': {e}")
            return None
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el dato de la columna del CSV: {e}")
            return None
    
    #63- Función que lee y parsea un archivo JSON, devolviendo su contenido.
    def leer_json(self, json_file_path: str) -> Union[Dict, List, None]:
        """
        Lee y parsea un archivo JSON, devolviendo su contenido.

        Args:
            json_file_path (str): La ruta completa al archivo JSON.

        Returns:
            Union[Dict, List, None]: El contenido del JSON como un diccionario o lista, o None si hay un error.
        """
        try:
            self.logger.info(f"\n 🔎 Intentando leer el archivo JSON: '{json_file_path}'.")
            with open(json_file_path, 'r', encoding='utf-8') as file: # Abre el archivo JSON en modo lectura con encoding UTF-8
                data = json.load(file) # Carga (parsea) el contenido del archivo JSON
            self.logger.info(f"\n ✅ Archivo JSON '{json_file_path}' leído exitosamente.")
            return data # Retorna los datos parseados del JSON
        except FileNotFoundError: # Captura si el archivo JSON no se encuentra
            self.logger.error(f"\n ❌ Error: El archivo JSON no se encontró en la ruta: {json_file_path}")
            return None
        except json.JSONDecodeError as e:
            self.logger.error(f"\n ❌ Error de formato JSON al leer '{json_file_path}': {e}")
            return None
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el archivo JSON: {e}")
            return None
        
    #64- Función que lee el contenido completo de un archivo de texto plano.
    #Si se proporciona un delimitador, divide el contenido del archivo por el delimitador.
    def leer_texto(self, file_path: str, delimiter: str = None) -> Union[str, List[str], None]:
        """
        Lee el contenido completo de un archivo de texto plano.
        Si se proporciona un delimitador, divide el contenido del archivo por el delimitador.

        Args:
            file_path (str): La ruta completa al archivo de texto.
            delimiter (str, opcional): Si se proporciona, el contenido del archivo se dividirá por este delimitador
                                       y se devolverá como una lista de cadenas. Si es None, se devuelve el contenido
                                       completo como una sola cadena. Por defecto es None.

        Returns:
            Union[str, List[str], None]: El contenido del archivo como una cadena (si delimiter es None) o
                                         una lista de cadenas (si se usa delimiter), o None si hay un error.
        """
        try:
            self.logger.info(f"\n 🔎 Intentando leer el archivo de texto: '{file_path}' (Delimitador: {'Ninguno' if delimiter is None else f"'{delimiter}'"}).")
            with open(file_path, 'r', encoding='utf-8') as file: # Abre el archivo en modo lectura con encoding UTF-8
                content = file.read() # Lee todo el contenido del archivo

            if delimiter is not None:
                self.logger.info(f"\n ✅ Archivo de texto '{file_path}' leído y dividido por delimitador '{delimiter}' exitosamente.")
                return content.split(delimiter) # Divide el contenido por el delimitador y lo retorna como lista
            else:
                self.logger.info(f"\n ✅ Archivo de texto '{file_path}' leído exitosamente.")
                return content # Retorna el contenido completo como una cadena
            
        except FileNotFoundError:
            self.logger.error(f"\n ❌ Error: El archivo de texto no se encontró en la ruta: {file_path}")
            return None
        except IOError as e:
            self.logger.error(f"\n ❌ Error de E/S al leer el archivo de texto '{file_path}': {e}")
            return None
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el archivo de texto: {e}")
            return None

    #65- Función que escribe contenido en un archivo de texto plano.
    #Si el contenido es una lista de cadenas y se proporciona un delimitador,
    #las cadenas se unirán con el delimitador antes de escribirlas.
    def escribir_texto(self, file_path: str, content: Union[str, List[str]], append: bool = False, delimiter: str = None) -> bool:
        """
        Escribe contenido en un archivo de texto plano.
        Si el contenido es una lista de cadenas y se proporciona un delimitador,
        las cadenas se unirán con el delimitador antes de escribirlas.

        Args:
            file_path (str): La ruta completa al archivo de texto.
            content (Union[str, List[str]]): La cadena o lista de cadenas a escribir.
            append (bool): Si es True, el contenido se añadirá al final del archivo.
                           Si es False (por defecto), el archivo se sobrescribirá si existe.
            delimiter (str, opcional): Si se proporciona y 'content' es una lista de cadenas,
                                       las cadenas se unirán con este delimitador antes de la escritura.
                                       Si es None, las cadenas de una lista se escribirán directamente (puede que se concatenen sin separación).
                                       Por defecto es None.

        Returns:
            bool: True si la escritura fue exitosa, False en caso de error.
        """
        mode = 'a' if append else 'w' # Determina el modo de apertura: 'a' para añadir, 'w' para sobrescribir
        action = "añadir a" if append else "escribir en" # Descripción de la acción para el log
        
        text_to_write = ""
        if isinstance(content, list): # Si el contenido es una lista
            if delimiter is not None:
                text_to_write = delimiter.join(content) # Une las cadenas de la lista con el delimitador
                self.logger.info(f"El contenido de la lista será unido con el delimitador '{delimiter}' antes de escribir.")
            else:
                text_to_write = "".join(content) # Si no hay delimitador, solo las une (concatenación simple)
                self.logger.warning("\n Se proporcionó una lista para escribir_texto sin delimitador. Las cadenas se concatenarán sin separación explícita.")
        elif isinstance(content, str): # Si el contenido ya es una cadena
            text_to_write = content # Lo escribe tal cual
        else:
            self.logger.error(f"\n ❌ Error: El tipo de contenido proporcionado no es válido para escribir_texto. Se esperaba str o List[str], se recibió: {type(content)}")
            return False

        try:
            self.logger.info(f"\n ✍️ Intentando {action} el archivo de texto: '{file_path}' (Delimitador de escritura: {'Ninguno' if delimiter is None else f"'{delimiter}'"}).")
            # Asegurarse de que el directorio exista
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory)
                self.logger.info(f"☑️ Directorio creado para el archivo de texto: {directory}")

            with open(file_path, mode, encoding='utf-8') as file: # Abre el archivo en el modo especificado con encoding UTF-8
                file.write(text_to_write) # Escribe el contenido en el archivo
            self.logger.info(f"\n ✅ Contenido {action} exitosamente en '{file_path}'.") # Loguea el éxito de la escritura/añadido
            return True # Retorna True si la operación fue exitosa
        
        except IOError as e:
            self.logger.error(f"\n ❌ Error de E/S al {action} el archivo de texto '{file_path}': {e}")
            return False
        except Exception as e:
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al {action} el archivo de texto: {e}")
            return False
    
    #66- Función para leer archivos XML
    def leer_xml(self, xml_file_path: str) -> Union[ET.Element, None]:
        """
        Lee y parsea un archivo XML, devolviendo su elemento raíz.

        Args:
            xml_file_path (str): La ruta completa al archivo XML.

        Returns:
            xml.etree.ElementTree.Element | None: El elemento raíz del XML como un objeto Element,
                                                    o None si hay un error.
        """
        try:
            self.logger.info(f"\n 🔎 Intentando leer el archivo XML: '{xml_file_path}'.")
            # Abre y parsea el archivo XML
            tree = ET.parse(xml_file_path)
            # Obtiene el elemento raíz del XML
            root = tree.getroot()
            self.logger.info(f"\n ✅ Archivo XML '{xml_file_path}' leído exitosamente. Elemento raíz: '{root.tag}'.")
            return root # Retorna el elemento raíz
        except FileNotFoundError: # Captura si el archivo XML no se encuentra
            self.logger.error(f"\n ❌ Error: El archivo XML no se encontró en la ruta: {xml_file_path}")
            return None
        except ET.ParseError as e: # Captura errores específicos de formato XML
            self.logger.error(f"\n ❌ Error de formato XML al leer '{xml_file_path}': {e}")
            return None
        except Exception as e: # Captura cualquier otro error inesperado
            self.logger.error(f"\n ❌ Ocurrió un error inesperado al leer el archivo XML: {e}")
            return None
    
    #67- Función que realiza una acción de click derecho (context click) sobre un elemento.
    def hacer_click_derecho_en_elemento(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Realiza una acción de click derecho (context click) sobre un elemento.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después del click derecho.
        """
        self.logger.info(f"\n Intentando hacer click derecho sobre el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_click_derecho", directorio)

            locator.click(button="right", timeout=tiempo * 1000) # El atributo 'button="right"' es clave aquí
            self.logger.info(f"\n ✔ ÉXITO: Click derecho realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_click_derecho", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer click derecho en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_click_derecho", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer click derecho en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_click_derecho", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer click derecho en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_click_derecho", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
    
    #68- Función que realiza una acción de 'mouse down' (presionar el botón del ratón) sobre un elemento.
    def hacer_mouse_down_en_elemento(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Realiza una acción de 'mouse down' (presionar el botón del ratón) sobre un elemento.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después del mouse down.
        """
        self.logger.info(f"\n Intentando hacer 'mouse down' sobre el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_mouse_down", directorio)

            locator.click()
            self.logger.info(f"\n ✔ ÉXITO: 'Mouse down' realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_mouse_down", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer 'mouse down' en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_mouse_down", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer 'mouse down' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_mouse_down", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer 'mouse down' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_mouse_down", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
    
    #69- Función que realiza una acción de 'mouse up' (soltar el botón del ratón) sobre un elemento.
    def hacer_mouse_up_en_elemento(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Realiza una acción de 'mouse up' (soltar el botón del ratón) sobre un elemento.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después del mouse up.
        """
        self.logger.info(f"\n Intentando hacer 'mouse up' sobre el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_mouse_up", directorio)

            locator.click()
            self.logger.info(f"\n ✔ ÉXITO: 'Mouse up' realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_mouse_up", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer 'mouse up' en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_mouse_up", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer 'mouse up' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_mouse_up", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer 'mouse up' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_mouse_up", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
    
    #70- Función que realiza una acción de 'focus' (enfocar) sobre un elemento.
    def hacer_focus_en_elemento(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Realiza una acción de 'focus' (enfocar) sobre un elemento.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después de hacer focus.
        """
        self.logger.info(f"\n Intentando hacer 'focus' sobre el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_focus", directorio)

            # El método focus() de Playwright establece el foco en el elemento
            locator.focus(timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: 'Focus' realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_focus", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer 'focus' en '{selector}'.\n"
                f"Posibles causas: El elemento no apareció, no fue visible/habilitado a tiempo.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_focus", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer 'focus' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_focus", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer 'focus' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_focus", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
    
    #71- Función que realiza una acción de 'blur' (desenfocar) sobre un elemento.
    def hacer_blur_en_elemento(self, selector, nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5):
        """
        Realiza una acción de 'blur' (desenfocar) sobre un elemento.

        Args:
            selector: El selector del elemento (puede ser un string o un Locator de Playwright).
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo de espera después de hacer blur.
        """
        self.logger.info(f"\n Intentando hacer 'blur' sobre el elemento con selector: '{selector}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector

            # Opcional: Podrías querer resaltar el elemento ANTES de desenfocarlo
            # locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_blur", directorio)

            # El método blur() de Playwright quita el foco del elemento
            locator.blur(timeout=tiempo * 1000)
            self.logger.info(f"\n ✔ ÉXITO: 'Blur' realizado exitosamente en el elemento con selector '{selector}'.")
            self.tomar_captura(f"{nombre_base}_despues_blur", directorio)

        except TimeoutError as e:
            error_msg = (
                f"\n ❌ ERROR (Timeout): El tiempo de espera se agotó al hacer 'blur' en '{selector}'.\n"
                f"Posibles causas: El elemento no estaba enfocado o no fue visible/habilitado a tiempo para la acción.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_timeout_blur", directorio)
            raise Error(error_msg) from e

        except Error as e:
            error_msg = (
                f"\n ❌ ERROR (Playwright): Ocurrió un problema de Playwright al hacer 'blur' en '{selector}'.\n"
                f"Verifica la validez del selector y el estado del elemento en el DOM.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_blur", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ ERROR (Inesperado): Se produjo un error desconocido al intentar hacer 'blur' en '{selector}'.\n"
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_blur", directorio)
            raise
        finally:
            if tiempo > 0:
                self.esperar_fijo(tiempo)
    
    #72- Función para verifica el estado de un checkbox (marcado/desmarcado) o el valor de una opción seleccionada en un select.
    def verificar_estado_checkbox_o_select(self, selector, estado_esperado: Union[bool, str], nombre_base: str, directorio: str, tiempo: Union[int, float] = 0.5) -> bool:
        """
        Verifica el estado de un checkbox (marcado/desmarcado) o el valor de una opción seleccionada en un select.

        Args:
            selector: El selector del checkbox o del elemento <select>.
                      Puede ser un string o un Locator de Playwright.
            estado_esperado (Union[bool, str]):
                - Para checkbox: True si se espera que esté marcado, False si se espera que esté desmarcado.
                - Para select: El valor (value) de la opción que se espera que esté seleccionada.
            nombre_base (str): Nombre base para las capturas de pantalla.
            directorio (str): Directorio donde se guardarán las capturas de pantalla.
            tiempo (Union[int, float]): Tiempo máximo de espera para la verificación.

        Returns:
            bool: True si la verificación es exitosa, False en caso contrario.
        """
        self.logger.info(f"\n Verificando estado para el selector: '{selector}'. Estado esperado: '{estado_esperado}'.")
        try:
            if isinstance(selector, str):
                locator = self.page.locator(selector)
            else:
                locator = selector
            
            locator.highlight()
            self.tomar_captura(f"{nombre_base}_antes_verificar_estado", directorio)

            # --- Lógica de Verificación ---
            if isinstance(estado_esperado, bool): # Verificación para Checkbox
                if estado_esperado:
                    expect(locator).to_be_checked(timeout=tiempo * 1000)
                else:
                    expect(locator).not_to_be_checked(timeout=tiempo * 1000)
                
                tipo_elemento = "checkbox"
                valor_actual = locator.is_checked()
                mensaje_exito = f"El {tipo_elemento} '{selector}' está {'marcado' if estado_esperado else 'desmarcado'} como se esperaba."
                mensaje_fallo_esperado = f"se esperaba {'marcado' if estado_esperado else 'desmarcado'} pero está {'marcado' if valor_actual else 'desmarcado'}."
            
            elif isinstance(estado_esperado, str): # Verificación para Select (option)
                expect(locator).to_have_value(estado_esperado, timeout=tiempo * 1000)
                
                tipo_elemento = "select/option"
                valor_actual = locator.input_value() # o locator.evaluate("el => el.value")
                mensaje_exito = f"La opción '{estado_esperado}' está seleccionada en el {tipo_elemento} '{selector}' como se esperaba."
                mensaje_fallo_esperado = f"se esperaba la opción con valor '{estado_esperado}' pero la actual es '{valor_actual}'."
            
            else:
                raise ValueError("El 'estado_esperado' debe ser un booleano para checkbox o un string para select.")

            self.logger.info(f"\n ✔ ÉXITO: {mensaje_exito}")
            self.tomar_captura(f"{nombre_base}_despues_verificar_estado", directorio)
            return True

        except TimeoutError as e:
            actual_value_str = "N/A"
            if "tipo_elemento" in locals(): # Para evitar errores si el tipo_elemento no se definió
                if tipo_elemento == "checkbox":
                    actual_value_str = str(locator.is_checked())
                elif tipo_elemento == "select/option":
                    actual_value_str = locator.input_value()
            
            error_msg = (
                f"\n ❌ FALLO (Timeout): El {tipo_elemento if 'tipo_elemento' in locals() else 'elemento'} '{selector}' "
                f"no cumplió el estado esperado '{estado_esperado}' después de {tiempo} segundos. "
                f"Estado actual: '{actual_value_str}'. Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_verificar_estado", directorio)
            return False

        except AssertionError as e:
            actual_value_str = "N/A"
            if "tipo_elemento" in locals():
                if tipo_elemento == "checkbox":
                    actual_value_str = str(locator.is_checked())
                elif tipo_elemento == "select/option":
                    actual_value_str = locator.input_value()

            error_msg = (
                f"\n ❌ FALLO (Aserción): El {tipo_elemento if 'tipo_elemento' in locals() else 'elemento'} '{selector}' "
                f"NO cumple el estado esperado. {mensaje_fallo_esperado if 'mensaje_fallo_esperado' in locals() else ''} "
                f"Detalles: {e}"
            )
            self.logger.warning(error_msg)
            self.tomar_captura(f"{nombre_base}_fallo_verificar_estado", directorio)
            return False

        except ValueError as e:
            error_msg = (
                f"\n ❌ ERROR (Valor Inválido): Se proporcionó un 'estado_esperado' no válido. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_valor_invalido_verificar_estado", directorio)
            raise

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright): Error de Playwright al verificar el estado del elemento '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.error(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_playwright_verificar_estado", directorio)
            raise

        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado): Ocurrió un error inesperado al verificar el estado del elemento '{selector}'. "
                f"Detalles: {e}"
            )
            self.logger.critical(error_msg, exc_info=True)
            self.tomar_captura(f"{nombre_base}_error_inesperado_verificar_estado", directorio)
            raise
        finally:
            if tiempo is not None and tiempo > 0: # Ajuste para manejar None en tiempo
                self.esperar_fijo(tiempo)

    #73- Función para extrae y retorna el valor de un elemento deshabilitado dado su selector.
    def obtener_valor_elemento_enable(self, selector, nombre_base, directorio, tiempo= 0.5) -> str | None:
        self.logger.info(f"\n ⚙️ Extrayendo valor del elemento deshabilitado con selector: '{selector}'")
        valor_extraido = None
        
        try:
            selector.highlight()
            self.tomar_captura(f"{nombre_base}_antes_extraccion_valor", directorio)

            # Usamos expect para asegurar que el elemento es visible y habilitado antes de intentar extraer
            expect(selector).to_be_visible(timeout=5000)
            self.logger.debug(f"\n Elemento '{selector}' es visible.")

            # Priorizamos input_value para campos de formulario (incluyendo <select>)
            try:
                # Playwright's input_value() es lo que necesitas para <select> 'value'
                valor_extraido = selector.input_value(timeout=1000)
                self.logger.debug(f"\n Valor extraído (input_value) de '{selector}': '{valor_extraido}'")
            except Error: # Usar Error para errores específicos de Playwright (e.g., no es un elemento de entrada)
                self.logger.debug(f"\n input_value no aplicable o falló para '{selector}'. Intentando text_content/inner_text.")
                
                # Si falla input_value, intentamos con text_content o inner_text para otros elementos
                try:
                    valor_extraido = selector.text_content(timeout=1000)
                    if valor_extraido is not None:
                        # Si text_content devuelve solo espacios en blanco o es vacío,
                        # intentamos inner_text (que a veces es más preciso para texto visible)
                        if valor_extraido.strip() == "":
                            valor_extraido = selector.inner_text(timeout=1000)
                            self.logger.debug(f"\n Valor extraído (inner_text) de '{selector}': '{valor_extraido}'")
                        else:
                            self.logger.debug(f"\n Valor extraído (text_content) de '{selector}': '{valor_extraido}'")
                    else:
                        valor_extraido = selector.inner_text(timeout=1000)
                        self.logger.debug(f"\n Valor extraído (inner_text) de '{selector}': '{valor_extraido}'")
                except Error:
                    self.logger.warning(f"\n No se pudo extraer input_value, text_content ni inner_text de '{selector}'.")
                    valor_extraido = None # Asegurarse de que sea None si todo falla

            if valor_extraido is not None:
                # Stripping whitespace for cleaner results if it's a string
                valor_final = valor_extraido.strip() if isinstance(valor_extraido, str) else valor_extraido
                self.logger.info(f"\n ✅ Valor final obtenido del elemento '{selector}': '{valor_final}'")
                self.tomar_captura(f"{nombre_base}_valor_extraido_exito", directorio)
                return valor_final
            else:
                self.logger.warning(f"\n ❌ No se pudo extraer ningún valor significativo del elemento '{selector}'.")
                self.tomar_captura(f"{nombre_base}_fallo_extraccion_valor_no_encontrado", directorio)
                return None

        except TimeoutError as e:
            mensaje_error = (
                f"\n ❌ FALLO (Timeout): El elemento '{selector}' "
                f"no se volvió visible/habilitado a tiempo para extraer su valor. Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_timeout_extraccion_valor", directorio)
            raise AssertionError(f"\n Elemento no disponible para extracción de valor: {selector}") from e

        except Error as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error de Playwright): Ocurrió un error de Playwright al intentar extraer el valor de '{selector}'. Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_playwright_error_extraccion_valor", directorio)
            raise AssertionError(f"\n Error de Playwright al extraer valor: {selector}") from e

        except Exception as e:
            mensaje_error = (
                f"\n ❌ FALLO (Error Inesperado): Ocurrió un error desconocido al intentar extraer el valor de '{selector}'. Detalles: {e}"
            )
            self.logger.error(mensaje_error)
            self.tomar_captura(f"{nombre_base}_fallo_inesperado_extraccion_valor", directorio)
            raise AssertionError(f"\n Error inesperado al extraer valor: {selector}") from e
        
    # --- Manejadores y funciones para Alertas y Confirmaciones ---

    # Handler para alertas simples (usado con page.once)
    def _get_simple_alert_handler_for_on(self):
        def handler(dialog: Dialog):
            self._alerta_detectada = True
            self._alerta_mensaje_capturado = dialog.message
            self._alerta_tipo_capturado = dialog.type
            self.logger.info(f"\n--> [LISTENER ON - Simple Alert] Alerta detectada: Tipo='{dialog.type}', Mensaje='{dialog.message}'") # Se agregó log
            dialog.accept()
            self.logger.info("\n--> [LISTENER ON - Simple Alert] Alerta ACEPTADA.") # Se agregó log
        return handler

    # Handler para diálogos de confirmación (usado con page.once)
    def _get_confirmation_dialog_handler_for_on(self, accion: str):
        def handler(dialog: Dialog):
            self._alerta_detectada = True
            self._alerta_mensaje_capturado = dialog.message
            self._alerta_tipo_capturado = dialog.type
            self.logger.info(f"\n--> [LISTENER ON - Dinámico] Confirmación detectada: Tipo='{dialog.type}', Mensaje='{dialog.message}'") # Se agregó log
            if accion == 'accept':
                dialog.accept()
                self.logger.info("\n--> [LISTENER ON - Dinámico] Confirmación ACEPTADA.") # Se agregó log
            elif accion == 'dismiss':
                dialog.dismiss()
                self.logger.info("--> [LISTENER ON - Dinámico] Confirmación CANCELADA.") # Se agregó log
            else:
                self.logger.warning(f"\n--> [LISTENER ON - Dinámico] Acción desconocida '{accion}'. Aceptando por defecto.") # Se cambió a warning y se agregó log
                dialog.accept()
        return handler    
    
    # Handler para diálogos de pregunta (prompt) (usado con page.once)
    def _get_prompt_dialog_handler_for_on(self, input_text: str, accion: str):
        def handler(dialog: Dialog):
            self._alerta_detectada = True
            self._alerta_mensaje_capturado = dialog.message
            self._alerta_tipo_capturado = dialog.type
            self._alerta_input_capturado = input_text
            self.logger.info(f"\n--> [LISTENER ON - Prompt Dinámico] Diálogo detectado: Tipo='{dialog.type}', Mensaje='{dialog.message}'") # Se agregó log

            if accion == 'accept':
                if dialog.type == "prompt":
                    dialog.accept(input_text)
                    self.logger.info(f"\n--> [LISTENER ON - Prompt Dinámico] Texto '{input_text}' introducido y prompt ACEPTADO.") # Se agregó log
                else:
                    dialog.accept()
                    self.logger.info("\n--> [LISTENER ON - Prompt Dinámico] Diálogo ACEPTADO (sin texto, no es prompt).") # Se agregó log
            elif accion == 'dismiss':
                dialog.dismiss()
                self.logger.info("\n--> [LISTENER ON - Prompt Dinámico] Diálogo CANCELADO.") # Se agregó log
            else:
                self.logger.warning(f"\n--> [LISTENER ON - Prompt Dinámico] Acción desconocida '{accion}'. Cancelando por defecto.") # Se cambió a warning y se agregó log
                dialog.dismiss()
        return handler

    # Handler de eventos para cuando se abre una nueva página.
    def _on_new_page(self, page: Page):
        """
        Manejador de eventos para nuevas páginas (popups).
        """
        self._popup_detectado = True
        self._popup_page = page
        self._popup_url_capturado = page.url
        self._popup_title_capturado = page.title()
        self._all_new_pages_opened_by_click.append(page)
        self.logger.info(f"\n 🌐 Nueva página (popup) detectada. URL: {page.url}")
        # Opcional: Si solo te interesa la primera popup o una específica, podrías manejarlo aquí.
        # Por ahora, solo la añadimos a la lista.
        
    # Handler para hacer drag and drop manual
    def _realizar_drag_and_drop_manual(self, elemento_origen, elemento_destino, nombre_base, directorio, nombre_paso, tiempo= 1, timeout_ms: int = 1000):
        """
        Método privado para realizar Drag and Drop usando las acciones de ratón de bajo nivel.
        Se llama como fallback si el método drag_and_drop() falla.
        """
        try:
            elemento_origen.hover(timeout=timeout_ms) # Mueve el ratón sobre el elemento de origen
            self.page.mouse.down() # Presiona el botón izquierdo del ratón
            self.page.wait_for_timeout(tiempo * 1000) # Pausa usando Playwright's wait_for_timeout (en milisegundos)
            elemento_destino.hover(timeout=timeout_ms) # Mueve el ratón sobre el elemento de destino
            self.page.wait_for_timeout(tiempo * 1000) # Pausa usando Playwright's wait_for_timeout (en milisegundos)
            self.page.mouse.up() # Suelta el botón izquierdo del ratón

        except Error as e:
            error_msg = (
                f"\n ❌ FALLO (Playwright Error - Manual) - {nombre_paso}: Ocurrió un error de Playwright al intentar realizar 'Drag and Drop' manualmente.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_manual_drag_and_drop_playwright", directorio)
            raise
        
        except Exception as e:
            error_msg = (
                f"\n ❌ FALLO (Inesperado - Manual) - {nombre_paso}: Ocurrió un error inesperado al intentar realizar 'Drag and Drop' manualmente.\n"
                f"Detalles: {e}"
            )
            self.logger.error(error_msg)
            self.tomar_captura(f"{nombre_base}_error_inesperado_manual_drag_and_drop", directorio)
            raise